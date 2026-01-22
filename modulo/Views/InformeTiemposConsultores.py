from datetime import datetime

from collections import defaultdict

from django.shortcuts import render, HttpResponse
from django.db.models import Sum

from modulo.forms import inforTiempoEmpleadosFilterForm
from modulo.models import Consultores, Tiempos_Cliente, Facturacion_Consultores

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required


def filtrar_datos(request):
    """
    Helper genérico que filtra Tiempos_Cliente a partir de parámetros GET.
    (Se deja por compatibilidad con otros posibles usos.)
    """
    documentos = request.GET.getlist('Documento')
    anios = request.GET.getlist('Anio')
    lineas = request.GET.getlist('LineaId')
    meses = request.GET.getlist('Mes')

    tiempos = Tiempos_Cliente.objects.all()

    if documentos:
        tiempos = tiempos.filter(Documento__in=documentos)
    else:
        # Filtrar directamente por documentos válidos en la consulta principal
        tiempos = tiempos.filter(
            Documento__in=Consultores.objects.values_list('Documento', flat=True)
        )

    if anios:
        tiempos = tiempos.filter(Anio__in=anios)
    if meses:
        tiempos = tiempos.filter(Mes__in=meses)
    if lineas:
        tiempos = tiempos.filter(LineaId__in=lineas)

    # Anotar el total por mes, línea y documento
    tiempos = tiempos.values('Mes', 'LineaId', 'Documento').annotate(
        total=Sum('Horas')
    )
    return tiempos


def Filtrar_datos(forms, tiempos):
    """
    Helper que aplica los filtros del formulario sobre el queryset de tiempos.
    Mantiene la misma firma/uso que tu código original.
    """
    documento = forms.cleaned_data.get('documento')
    anio = forms.cleaned_data.get('Anio')
    mes = forms.cleaned_data.get('Mes')
    linea = forms.cleaned_data.get('LineaId')

    if documento:
        tiempos = tiempos.filter(Documento__in=documento)
    if anio:
        tiempos = tiempos.filter(Anio=anio)
    if mes:
        tiempos = tiempos.filter(Mes__in=mes)
    if linea:
        tiempos = tiempos.filter(LineaId__in=linea)

    return tiempos


@login_required
@verificar_permiso('can_view_informe_tiempos_consultores')
def tiempos_clientes_filtrado(request):
    """
    Vista principal del informe de tiempos de consultores.
    - Aplica filtros.
    - Arma la estructura para la tabla (incluyendo facturaciones dinámicas).
    - Deja en sesión los datos para la exportación a Excel.
    """
    tiempos_info = []
    show_data = False
    max_facturaciones = 0

    # Form siempre disponible en el contexto
    forms = inforTiempoEmpleadosFilterForm(request.GET or None)

    if request.method == 'GET':
        documentos_validos = set(
            Consultores.objects.values_list('Documento', flat=True)
        )

        # Base de tiempos con select_related para evitar N+1
        tiempos_qs = Tiempos_Cliente.objects.select_related(
            'ClienteId',
            'LineaId',
        ).filter(Documento__in=documentos_validos)

        if forms.is_valid():
            # Aplicar filtros del formulario
            tiempos_qs = Filtrar_datos(forms, tiempos_qs)

            # Evaluar queryset una sola vez y reutilizar
            tiempos_lista = list(tiempos_qs)

            # Cargar consultores en memoria (evita consultas dentro del loop)
            consultores = {
                c.Documento: c
                for c in Consultores.objects.select_related('ModuloId').all()
            }

            # ==============================
            # PRE-CARGA DE FACTURACIONES
            # ==============================
            # Se arma un queryset único de facturaciones y se agrupa en memoria
            docs = {t.Documento for t in tiempos_lista}
            anios = {t.Anio for t in tiempos_lista}
            clientes = {t.ClienteId_id for t in tiempos_lista if t.ClienteId_id}
            lineas = {t.LineaId_id for t in tiempos_lista if t.LineaId_id}
            meses = {t.Mes for t in tiempos_lista}

            fact_qs = Facturacion_Consultores.objects.filter(
                Documento__in=docs,
                Anio__in=anios,
                ClienteId__in=clientes,
                LineaId__in=lineas,
                Periodo_Cobrado__in=meses,
            ).order_by('Cta_Cobro')

            # Mapa: (Documento, Anio, ClienteId_id, LineaId_id, Periodo_Cobrado) -> [facturaciones...]
            facturaciones_map = defaultdict(list)
            for f in fact_qs:
                key = (
                    f.Documento,
                    f.Anio,
                    f.ClienteId_id,
                    f.LineaId_id,
                    f.Periodo_Cobrado,
                )
                facturaciones_map[key].append(f)

            # ==============================
            # ARMADO DE REGISTROS
            # ==============================
            for t in tiempos_lista:
                # Si el consultor no existe, lo saltamos (mismo criterio que tu código)
                if t.Documento not in consultores:
                    continue

                cons = consultores[t.Documento]

                # Clave para facturaciones asociada al tiempo actual
                fact_key = (
                    t.Documento,
                    t.Anio,
                    t.ClienteId_id,
                    t.LineaId_id,
                    t.Mes,
                )
                facturaciones = facturaciones_map.get(fact_key, [])

                registro = {
                    'Documento': t.Documento,
                    'Nombre': cons.Nombre if cons else '',
                    'Empresa': cons.Empresa if cons else '',
                    'Anio': t.Anio,
                    'Mes': t.Mes,
                    'Linea': t.LineaId.Linea if t.LineaId else '',
                    'Cliente': t.ClienteId.Nombre_Cliente if t.ClienteId else '',
                    'Modulo': (
                        cons.ModuloId.Modulo if cons and cons.ModuloId else ''
                    ),
                    'Horas': float(t.Horas) if t.Horas else 0.0,
                    'Facturaciones': [],
                }

                for facturacion in facturaciones:
                    registro['Facturaciones'].append({
                        'Cta_Cobro': facturacion.Cta_Cobro or '',
                        'Horas_Facturacion': float(facturacion.Horas) if facturacion.Horas else 0.0,
                    })

                num_facturaciones = len(registro['Facturaciones'])
                if num_facturaciones > max_facturaciones:
                    max_facturaciones = num_facturaciones

                tiempos_info.append(registro)

            # Ordenar por nombre de consultor (igual que antes)
            tiempos_info.sort(key=lambda x: x['Nombre'])
            show_data = bool(tiempos_info)

        # Guardar en sesión para exportar
        request.session['tiempos_info'] = tiempos_info
        request.session['max_facturaciones'] = max_facturaciones

    context = {
        'form': forms,
        'tiempos_info': tiempos_info,
        'show_data': show_data,
        'max_facturaciones': max_facturaciones,
        'mensaje': (
            "No se encontraron resultados para los filtros aplicados"
            if not tiempos_info
            else ""
        ),
    }

    return render(
        request,
        'Informes/InformestiemposconsultoresIndex.html',
        context,
    )


@login_required
@verificar_permiso('can_view_informe_tiempos_consultores')
def exportar_tiempos_clientes_excel(request):
    """
    Exporta a Excel los datos ya calculados y guardados en sesión por
    tiempos_clientes_filtrado.
    """
    tiempos_info = request.session.get('tiempos_info', [])
    max_facturaciones = request.session.get('max_facturaciones', 0)

    if not tiempos_info:
        return HttpResponse("No hay datos para exportar.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Tiempos Consultores"

    # Encabezados base
    encabezados_base = [
        "Documento",
        "Nombre",
        "Empresa",
        "Línea",
        "Cliente",
        "Módulo",
        "Año",
        "Mes",
        "Horas",
    ]

    # Columnas dinámicas según max_facturaciones
    for i in range(1, max_facturaciones + 1):
        encabezados_base.extend([
            f"Cta_Cobro_{i}",
            f"Horas_Facturacion_{i}",
        ])

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Escribir encabezados
    for col_num, header in enumerate(encabezados_base, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escribir filas
    for row_num, registro in enumerate(tiempos_info, 2):
        base_data = [
            registro.get('Documento', ''),
            registro.get('Nombre', ''),
            registro.get('Empresa', ''),
            registro.get('Linea', ''),
            registro.get('Cliente', ''),
            registro.get('Modulo', ''),
            registro.get('Anio', ''),
            registro.get('Mes', ''),
            registro.get('Horas', 0),
        ]

        facturaciones_data = []
        for fact in registro.get('Facturaciones', []):
            facturaciones_data.extend([
                fact.get('Cta_Cobro', ''),
                fact.get('Horas_Facturacion', 0),
            ])

        # Rellenar hasta el máximo de columnas dinámicas
        while len(facturaciones_data) < max_facturaciones * 2:
            facturaciones_data.extend(['', ''])

        row_data = base_data + facturaciones_data[: max_facturaciones * 2]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border

            header_lower = encabezados_base[col_num - 1].lower()
            if isinstance(value, (int, float)) and header_lower.startswith('horas'):
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(value, str):
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

    # Autoajuste de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Respuesta HTTP con el Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"tiempos_consultores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response
