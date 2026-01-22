# Informe de Empleados

# Librerías estándar y utilitarias
from datetime import datetime
from collections import defaultdict, Counter

# Django
from django.http import HttpResponse
from django.db.models import Q, Count
from django.shortcuts import render

# Formularios y modelos del módulo
from modulo.forms import EmpleadoFilterForm
from modulo.models import Empleado

# Librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required


def filtrar_empleados(form, empleados):
    """
    Aplica los filtros del formulario sobre el queryset de empleados:
    - Nombre exacto
    - Línea, Cargo, Perfil
    - Certificado SAP (valor booleano)
    - Estado Activo (True/False)
    """
    nombre = form.cleaned_data.get('Nombre')
    linea = form.cleaned_data.get('LineaId')
    cargo = form.cleaned_data.get('Cargo')
    certificaciones = form.cleaned_data.get('CertificadoSAP')
    perfil = form.cleaned_data.get('PerfilId')
    activo = form.cleaned_data.get('Activo')

    if nombre:
        empleados = empleados.filter(Nombre=nombre)
    if linea:
        empleados = empleados.filter(LineaId=linea)
    if cargo:
        empleados = empleados.filter(CargoId=cargo)
    if certificaciones:
        empleados = empleados.filter(CertificadoSAP=bool(int(certificaciones)))
    if perfil:
        empleados = empleados.filter(PerfilId=perfil)
    if activo:
        empleados = empleados.filter(Activo=activo == 'True')

    return empleados


def obtener_informe_empleados(empleados):
    """
    Recorre el queryset de empleados y construye una lista de diccionarios
    con los datos a mostrar y exportar. Incluye datos personales, académicos
    y de certificación.
    """
    empleado_info = []

    # El select_related se hace en la vista para evitar N+1 queries
    for empleado in empleados:
        datos_empleado = {
            'TipoDocumento': empleado.TipoDocumento.Nombre,
            'Documento': empleado.Documento,
            'Nombre': empleado.Nombre,
            'FechaNacimiento': empleado.FechaNacimiento,
            'FechaIngreso': empleado.FechaIngreso,
            'FechaOperacion': empleado.FechaOperacion,
            'Modulo': empleado.ModuloId.Modulo,
            'Perfil': empleado.PerfilId.Perfil,
            'Linea': empleado.LineaId.Linea,
            'CargoId': empleado.CargoId.Cargo,
            'TituloProfesional': empleado.TituloProfesional,
            'FechaGrado': empleado.FechaGrado,
            'Universidad': empleado.Universidad,
            'ProfesionRealizada': empleado.ProfesionRealizada,
            'AcademiaSAP': 'SI' if empleado.AcademiaSAP else 'NO',
            'CertificadoSAP': 'SI' if empleado.CertificadoSAP else 'NO',
            'OtrasCertificaciones': 'SI' if empleado.OtrasCertificaciones else 'NO',
            'Postgrados': 'SI' if empleado.Postgrados else 'NO',
            'Activo': 'SI' if empleado.Activo else 'NO',
            'FechaRetiro': empleado.FechaRetiro,
            'Direccion': empleado.Direccion if empleado.Direccion else '',
            'Ciudad': empleado.Ciudad if empleado.Ciudad else '',
            'Departamento': empleado.Departamento if empleado.Departamento else '',
            'DireccionAlterna': empleado.DireccionAlterna if empleado.DireccionAlterna else '',
            'Telefono1': empleado.Telefono1 if empleado.Telefono1 else '',
            'Telefono2': empleado.Telefono2 if empleado.Telefono2 else '',
        }

        empleado_info.append(datos_empleado)

    return empleado_info


@login_required
@verificar_permiso('can_view_informe_empleado')
def informe_empleados(request):
    """
    Vista que construye el informe de empleados:
    - Aplica filtros si se recibe GET con parámetros.
    - Genera estructura de datos para tabla y cards.
    - Prepara diccionarios para gráficos por línea, módulo, perfil, etc.
    """
    empleados_info = []
    # QuerySet vacío por defecto
    empleados = Empleado.objects.none()
    show_data = False
    busqueda_realizada = False

    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)

        if request.GET:
            busqueda_realizada = True

            if form.is_valid():
                # Se hace select_related para evitar N+1 en relaciones usadas en el reporte
                empleados = (
                    Empleado.objects
                    .select_related(
                        'TipoDocumento',
                        'ModuloId',
                        'PerfilId',
                        'LineaId',
                        'CargoId',
                    )
                )

                # Aplicar filtros
                empleados = filtrar_empleados(form, empleados)

                # Obtener la información del informe solo si hay resultados
                if empleados.exists():
                    empleados_info = obtener_informe_empleados(empleados)
                    show_data = True
    else:
        form = EmpleadoFilterForm()

    # =====================
    # Métricas para cards
    # =====================
    empleados_totales = 0
    empleados_activos = 0
    empleados_inactivos = 0
    total_empleados = 0
    activos = 0

    if show_data:
        # Un solo hit a BD para totales
        agregados = empleados.aggregate(
            total=Count('pk'),
            activos=Count('pk', filter=Q(Activo=True)),
        )
        empleados_totales = agregados['total'] or 0
        empleados_activos = agregados['activos'] or 0
        empleados_inactivos = empleados_totales - empleados_activos

        total_empleados = len(empleados_info)
        # Activos en el reporte (equivalente a empleados_activos)
        activos = empleados_activos

    # Diccionarios para acumulación por línea
    certificados_sap_por_linea = defaultdict(int)
    otras_certificaciones_por_linea = defaultdict(int)
    postgrados_por_linea = defaultdict(int)

    # Agrupaciones para gráficos
    modulos_data = Counter()
    perfiles_data = Counter()
    lineas_data = Counter()

    for e in empleados_info:
        linea = e['Linea']

        # Acumuladores por línea
        if e['CertificadoSAP'] == 'SI':
            certificados_sap_por_linea[linea] += 1
        if e['OtrasCertificaciones'] == 'SI':
            otras_certificaciones_por_linea[linea] += 1
        if e['Postgrados'] == 'SI':
            postgrados_por_linea[linea] += 1

        # Contadores por agrupación
        modulos_data[e['Modulo']] += 1
        perfiles_data[e['Perfil']] += 1
        lineas_data[linea] += 1

    # Totales para las cards de certificados/postgrados
    total_certificados_sap = sum(certificados_sap_por_linea.values())
    total_otras_certificaciones = sum(otras_certificaciones_por_linea.values())
    total_postgrados = sum(postgrados_por_linea.values())

    context = {
        'form': form,
        'empleados_info': empleados_info,
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'mensaje': (
            "No se encontraron resultados para los filtros aplicados."
            if busqueda_realizada and not show_data
            else "No se ha realizado ninguna búsqueda aún."
        ),

        # Cards
        'total_empleados': total_empleados,
        'activos': activos,
        'empleados_totales': empleados_totales,
        'empleados_activos': empleados_activos,
        'empleados_inactivos': empleados_inactivos,
        'certificados_sap': total_certificados_sap,
        'otras_certificaciones': total_otras_certificaciones,
        'postgrados': total_postgrados,

        # Diccionarios para gráficas
        'certificados_sap_por_linea': dict(certificados_sap_por_linea),
        'otras_certificaciones_por_linea': dict(otras_certificaciones_por_linea),
        'postgrados_por_linea': dict(postgrados_por_linea),

        'certificados_sap_labels': list(certificados_sap_por_linea.keys()),
        'certificados_sap_data': list(certificados_sap_por_linea.values()),

        'otras_certificaciones_labels': list(otras_certificaciones_por_linea.keys()),
        'otras_certificaciones_data': list(otras_certificaciones_por_linea.values()),

        'postgrados_labels': list(postgrados_por_linea.keys()),
        'postgrados_data': list(postgrados_por_linea.values()),

        'activos_inactivos_labels': ['Activos', 'Inactivos'],
        'activos_inactivos_data': [empleados_activos, empleados_inactivos],

        'modulos_labels': list(modulos_data.keys()),
        'modulos_data': list(modulos_data.values()),

        'perfiles_labels': list(perfiles_data.keys()),
        'perfiles_data': list(perfiles_data.values()),

        'lineas_labels': list(lineas_data.keys()),
        'lineas_data': list(lineas_data.values()),
    }

    return render(request, 'Informes/InformesEmpleadoIndex.html', context)


@login_required
@verificar_permiso('can_view_informe_empleado')
def exportar_empleados_excel(request):
    """
    Genera un archivo Excel con los datos del informe de empleados:
    - Aplica los mismos filtros del formulario.
    - Formatea y ajusta el archivo usando openpyxl.
    - Devuelve un archivo descargable con timestamp.
    """
    empleado_info = []

    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)

        if form.is_valid():
            # Inicializar empleados con select_related para evitar N+1
            empleados = (
                Empleado.objects
                .select_related(
                    'TipoDocumento',
                    'ModuloId',
                    'PerfilId',
                    'LineaId',
                    'CargoId',
                )
            )

            # Filtrar empleados
            empleados = filtrar_empleados(form, empleados)

            # Obtener información de los empleados
            empleado_info = obtener_informe_empleados(empleados)

        # Mantengo el mismo flujo de control que tenías:
        # primero se revisa si hay datos, luego la validez del form.
        if not empleado_info:
            return HttpResponse("No se encontraron datos para exportar.", status=404)

        if not form.is_valid():
            return HttpResponse(
                "Filtros no válidos. Por favor, revisa los criterios ingresados.",
                status=400,
            )

    # Preparar datos para exportación
    data = []
    for empleado in empleado_info:
        fila = {
            'Línea': empleado['Linea'],
            'Documento Colaborador': empleado['Documento'],
            'Nombre Colaborador': empleado['Nombre'],
            'Perfil': empleado['Perfil'],
            'Módulo': empleado['Modulo'],
            'Certificado SAP': empleado['CertificadoSAP'],
            'Fecha Ingreso': empleado['FechaIngreso'],
            'Documento': empleado['Documento'],
            'Nombre': empleado['Nombre'],
            'FechaNacimiento': empleado['FechaNacimiento'],
            'FechaIngreso': empleado['FechaIngreso'],  # columna adicional como en el original
            'FechaOperacion': empleado['FechaOperacion'],
            'TituloProfesional': empleado['TituloProfesional'],
            'FechaGrado': empleado['FechaGrado'],
            'Universidad': empleado['Universidad'],
            'ProfesionRealizada': empleado['ProfesionRealizada'],
            'AcademiaSAP': empleado['AcademiaSAP'],
            'OtrasCertificaciones': empleado['OtrasCertificaciones'],
            'Postgrados': empleado['Postgrados'],
            'Activo': empleado['Activo'],
            'FechaRetiro': empleado['FechaRetiro'],
            'Direccion': empleado['Direccion'],
            'Ciudad': empleado['Ciudad'],
            'Departamento': empleado['Departamento'],
            'DireccionAlterna': empleado['DireccionAlterna'],
            'Telefono1': empleado['Telefono1'],
            'Telefono2': empleado['Telefono2'],
        }
        data.append(fila)

    # Crear y formatear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe de Empleados"

    # Agregar encabezados
    headers = list(data[0].keys())
    for col_idx, col in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Agregar datos
    for r_idx, row in enumerate(data, start=2):
        for c_idx, value in enumerate(row.values(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')

    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.border = thin_border

    # Crear respuesta de Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"empleados_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    # 👉 línea corregida (sin comillas mezcladas)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)

    return response
