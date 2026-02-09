# modulo/Views/InformeFacturacionCentroCostos.py

from collections import defaultdict
from decimal import Decimal
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q, IntegerField
from django.db.models.functions import Cast
from django.http import HttpResponse
from django.shortcuts import render

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from modulo.decorators import verificar_permiso
from modulo.forms import FacturacionClientesFilterForm
from modulo.models import (
    FacturacionClientes,
    Linea,
    CentrosCostos,
    Clientes,
    Modulo,
    Tiempos_Cliente,
    Nomina,
    Tarifa_Consultores,
    Tarifa_Clientes,
    Horas_Habiles,
    IND,
)


# ----------------------------
# Helpers de cálculo
# ----------------------------
def _D(x, default="0"):
    try:
        return Decimal(str(x if x is not None else default))
    except Exception:
        return Decimal(default)


def _build_month_sets(meses_filtrados):
    """
    Retorna:
      - meses_int: lista única de ints [1..12] o los seleccionados
      - meses_query: lista con ints + strings ("1" y "01") para soportar DB con o sin cero a la izquierda
    """
    if meses_filtrados:
        meses_int = sorted({int(str(m)) for m in meses_filtrados})
    else:
        meses_int = list(range(1, 13))

    meses_query = set()
    for m in meses_int:
        meses_query.add(m)
        meses_query.add(str(m))
        meses_query.add(f"{m:02d}")

    return meses_int, list(meses_query)


def _pick_effective_entry(entries, anio_int: int, mes_int: int):
    """
    entries: lista ordenada desc [(anio, mes, ...)]
    Retorna la primera entrada con (anio, mes) <= (anio_int, mes_int)
    """
    target = (anio_int, mes_int)
    for e in entries:
        if (e[0], e[1]) <= target:
            return e
    return None


def _has_year_selected(form: FacturacionClientesFilterForm) -> bool:
    """
    ✅ Solo permite ejecutar consultas si el usuario seleccionó Año.
    - Si el form no está bound => False
    - Si está bound pero Anio vacío => False
    """
    if not form or not getattr(form, "is_bound", False):
        return False

    raw = (form.data.get("Anio") or "").strip()
    return raw not in ("", "None", "Todos")


# ----------------------------
# Horas hábiles / IND
# ----------------------------
def obtener_horas_habiles(anio: str) -> dict:
    """Obtiene días y horas hábiles por mes para un año específico (en Decimal para cálculos seguros)."""
    registros = Horas_Habiles.objects.filter(Anio=anio)
    data = {}
    for r in registros:
        dias = _D(r.Dias_Habiles, "0")
        horas_diarias = _D(r.Horas_Laborales, "0")
        data[str(r.Mes)] = {
            "dias": dias,
            "horas_mes": dias * horas_diarias if dias and horas_diarias else Decimal("0"),
            "horas_diarias": horas_diarias,
        }
    return data


def obtener_indice_ind(anio=None, mes=None) -> Decimal:
    """
    Devuelve el IND como PORCENTAJE (ej: 1.56 significa 1.56%).
    Luego, para aplicar al salario: salario * (1 + IND/100)

    Reglas:
    - Si existe IND exacto para (anio, mes): usarlo (soporta mes "01" vs "1").
    - Si no existe: usar el IND más reciente <= (anio, mes).
    - Si no hay ninguno anterior: usar el más cercano >= (anio, mes).
    - Si no hay anio/mes (None): usar el último IND existente.
    - Si no hay registros: retorna 0 (para que el factor sea 1).
    """

    qs = IND.objects.all().annotate(
        anio_i=Cast("Anio", IntegerField()),
        mes_i=Cast("Mes", IntegerField()),
    )

    # Caso sin fecha: último existente
    if anio is None or mes is None:
        reg = qs.order_by("-anio_i", "-mes_i").first()
        return _D(reg.Indice, "0") if reg else Decimal("0")

    anio_int = int(anio)
    mes_int = int(mes)

    # 1) Exacto (resuelve "01" vs "1")
    anio_candidates = [str(anio_int), anio_int]
    mes_candidates = [f"{mes_int:02d}", str(mes_int), mes_int]
    reg = IND.objects.filter(Anio__in=anio_candidates, Mes__in=mes_candidates).first()
    if reg and reg.Indice is not None:
        return _D(reg.Indice, "0")

    # 2) Más reciente anterior o igual (<=)
    prior = qs.filter(
        Q(anio_i__lt=anio_int) | (Q(anio_i=anio_int) & Q(mes_i__lte=mes_int))
    ).order_by("-anio_i", "-mes_i").first()
    if prior and prior.Indice is not None:
        return _D(prior.Indice, "0")

    # 3) Más cercano posterior o igual (>=)
    after = qs.filter(
        Q(anio_i__gt=anio_int) | (Q(anio_i=anio_int) & Q(mes_i__gte=mes_int))
    ).order_by("anio_i", "mes_i").first()
    if after and after.Indice is not None:
        return _D(after.Indice, "0")

    # 4) Último recurso: último existente
    last = qs.order_by("-anio_i", "-mes_i").first()
    return _D(last.Indice, "0") if last else Decimal("0")


# ----------------------------
# Precarga / Costos
# ----------------------------
def precargar_datos(anio: str, meses_query: list, clientes: list, lineas: list):
    """Precarga todos los datos necesarios para optimizar consultas."""
    clientes_ids = [c.ClienteId for c in clientes]
    lineas_ids = [l.LineaId for l in lineas]

    tiempos = Tiempos_Cliente.objects.filter(
        Anio=anio,
        Mes__in=meses_query,
        ClienteId__in=clientes_ids,
        LineaId__in=lineas_ids,
    ).values(
        "Anio",
        "Mes",
        "Documento",
        "ClienteId",
        "LineaId",
        "Horas",
        "ModuloId",
        "centrocostosId",
    )

    # Nóminas
    nominas = Nomina.objects.all().values("Documento", "Anio", "Mes", "Salario")
    nominas_dict = defaultdict(list)
    for n in nominas:
        nominas_dict[n["Documento"]].append(
            (int(n["Anio"]), int(n["Mes"]), _D(n["Salario"], "0"))
        )
    for doc in nominas_dict:
        nominas_dict[doc].sort(key=lambda x: (-x[0], -x[1]))

    # Tarifas consultores
    tarifas_consultores = Tarifa_Consultores.objects.all().values(
        "documentoId", "anio", "mes", "valorHora", "valorDia", "valorMes", "moduloId"
    )
    tarifas_consultores_dict = defaultdict(lambda: defaultdict(list))
    for t in tarifas_consultores:
        tarifas_consultores_dict[t["documentoId"]][t["moduloId"]].append(
            (
                int(t["anio"]),
                int(t["mes"]),
                _D(t["valorHora"], "0"),
                _D(t["valorDia"], "0"),
                _D(t["valorMes"], "0"),
            )
        )
    for doc in tarifas_consultores_dict:
        for modulo in tarifas_consultores_dict[doc]:
            tarifas_consultores_dict[doc][modulo].sort(key=lambda x: (-x[0], -x[1]))

    tarifas_clientes = Tarifa_Clientes.objects.filter(clienteId__in=clientes_ids)
    tarifas_por_cliente = defaultdict(list)
    for tarifa in tarifas_clientes:
        tarifas_por_cliente[tarifa.clienteId_id].append(tarifa)

    return {
        "tiempos_data": list(tiempos),
        "nominas_dict": nominas_dict,
        "tarifas_consultores_dict": tarifas_consultores_dict,
        "tarifas_por_cliente": tarifas_por_cliente,
    }


def calcular_costos_por_ceco_modulo_mes(
    anio: str,
    meses_int: list,
    lineas_ids: list,
    datos_precargados: dict,
    horas_habiles: dict,
):
    """
    Calcula costos agrupados por Centro de Costos (código), Módulo (nombre) y Mes (string "1","2"...).

    Reglas:
    - EMPLEADOS: salario nómina incrementado por % IND del mes:
        salario_ajustado = salario * (1 + IND/100)
      Luego:
        costo_hora = salario_ajustado / horas_habiles_mes
    - CONSULTORES: tarifa (hora/día/mes) sin IND.
    """

    costos_por_ceco_modulo_mes = defaultdict(lambda: defaultdict(lambda: defaultdict(Decimal)))
    lineas_ids_set = set(lineas_ids or [])

    ind_factor_cache = {}
    ceco_codigo_cache = {}
    modulo_nombre_cache = {}

    def get_hh_mes(mes_int: int):
        for k in (str(mes_int), f"{mes_int:02d}"):
            if k in horas_habiles:
                return horas_habiles[k]
        return None

    def get_ind_factor(anio_int: int, mes_int: int) -> Decimal:
        key = (anio_int, mes_int)
        if key in ind_factor_cache:
            return ind_factor_cache[key]

        ind_val = _D(obtener_indice_ind(anio_int, mes_int), "0")  # EJ: 1.56
        # ✅ IND es FACTOR (multiplicador). Si no hay IND, usar 1.
        factor = ind_val if ind_val > 0 else Decimal("1")

        ind_factor_cache[key] = factor
        return factor

    for mes_int in (meses_int or []):
        hh_mes = get_hh_mes(int(mes_int))
        if not hh_mes:
            continue

        horas_mes_total = _D(hh_mes.get("horas_mes", 0), "0")
        horas_diarias = _D(hh_mes.get("horas_diarias", 0), "0")

        ind_factor_mes = get_ind_factor(int(anio), int(mes_int))

        for tiempo in datos_precargados.get("tiempos_data", []):
            if str(tiempo.get("Anio")) != str(anio):
                continue

            try:
                tiempo_mes_int = int(str(tiempo.get("Mes")))
            except Exception:
                continue
            if tiempo_mes_int != int(mes_int):
                continue

            if tiempo.get("LineaId") not in lineas_ids_set:
                continue

            documento = str(tiempo.get("Documento", "")).strip()
            anio_int = int(str(tiempo.get("Anio")))
            modulo_id = tiempo.get("ModuloId")
            ceco_id = tiempo.get("centrocostosId")
            horas_trabajadas = _D(tiempo.get("Horas", 0), "0")

            costo_hora = Decimal("0")

            # EMPLEADOS
            nominas_dict = datos_precargados.get("nominas_dict", {})
            if documento in nominas_dict:
                entrada = _pick_effective_entry(nominas_dict.get(documento, []), anio_int, int(mes_int))
                if entrada:
                    salario = _D(entrada[2], "0")
                    if horas_mes_total > 0:
                        salario_ajustado = salario * ind_factor_mes
                        costo_hora = salario_ajustado / horas_mes_total

            # CONSULTORES
            else:
                tarifas_dict = datos_precargados.get("tarifas_consultores_dict", {})
                entradas_tarifa = tarifas_dict.get(documento, {}).get(modulo_id, [])
                entrada = _pick_effective_entry(entradas_tarifa, anio_int, int(mes_int))
                if entrada:
                    valorHora = _D(entrada[2], "0")
                    valorDia = _D(entrada[3], "0")
                    valorMes = _D(entrada[4], "0")

                    if valorHora > 0:
                        costo_hora = valorHora
                    elif valorDia > 0 and horas_diarias > 0:
                        costo_hora = valorDia / horas_diarias
                    elif valorMes > 0 and horas_mes_total > 0:
                        costo_hora = valorMes / horas_mes_total

            costo_total = horas_trabajadas * costo_hora

            # código CeCo (cache)
            codigo_ceco = "SIN_CECO"
            if ceco_id:
                if ceco_id in ceco_codigo_cache:
                    codigo_ceco = ceco_codigo_cache[ceco_id]
                else:
                    codigo = CentrosCostos.objects.filter(id=ceco_id).values_list("codigoCeCo", flat=True).first()
                    codigo_ceco = codigo or "SIN_CECO"
                    ceco_codigo_cache[ceco_id] = codigo_ceco

            # nombre módulo (cache)
            nombre_modulo = "SIN_MODULO"
            if modulo_id is not None:
                if modulo_id in modulo_nombre_cache:
                    nombre_modulo = modulo_nombre_cache[modulo_id]
                else:
                    nombre = Modulo.objects.filter(ModuloId=modulo_id).values_list("Modulo", flat=True).first()
                    nombre_modulo = nombre or "SIN_MODULO"
                    modulo_nombre_cache[modulo_id] = nombre_modulo

            costos_por_ceco_modulo_mes[codigo_ceco][nombre_modulo][str(int(mes_int))] += costo_total

    return costos_por_ceco_modulo_mes


# ----------------------------
# Filtros / Vistas
# ----------------------------
def filtrar_datos(form=None):
    """
    ✅ Importante:
    - Si no hay año seleccionado => NO consultamos nada y devolvemos vacío.
    - Si el form está bound pero inválido => devolvemos vacío (para no mostrar todo).
    """
    if not form:
        return ([], Linea.objects.none(), None, {})

    # Si no hay año seleccionado, no hacemos consultas
    if not _has_year_selected(form):
        return ([], Linea.objects.none(), None, {})

    # Si intentó buscar (año) pero el form es inválido: no mostrar todo
    if form.is_bound and not form.is_valid():
        return ([], Linea.objects.none(), None, {})

    # A partir de aquí: hay año y form válido
    facturas = FacturacionClientes.objects.all().select_related("LineaId", "ModuloId")

    anio = form.cleaned_data.get("Anio")
    lineas = form.cleaned_data.get("LineaId")
    meses = form.cleaned_data.get("Mes")
    ceco = form.cleaned_data.get("Ceco")

    # filtros (anio ya existe sí o sí)
    facturas = facturas.filter(Anio=anio)

    if lineas:
        facturas = facturas.filter(LineaId__in=lineas)

    if meses:
        meses_int, meses_query = _build_month_sets(meses)
        facturas = facturas.filter(Mes__in=meses_query)

    if ceco:
        cecos_filtro = list(ceco) if isinstance(ceco, (list, tuple)) else [ceco]
        cecos_filtro = [str(c).strip() for c in cecos_filtro]
        facturas = facturas.filter(Ceco__in=cecos_filtro)
        if facturas.count() == 0:
            # fallback regex si el campo trae variaciones
            facturas = FacturacionClientes.objects.filter(Anio=anio)
            facturas = facturas.filter(Ceco__iregex=r"^(%s)$" % "|".join(cecos_filtro))

    # CALCULAR COSTOS (solo con año)
    meses_int, meses_query = _build_month_sets(meses)
    lineas_ids = [l.LineaId for l in (lineas or Linea.objects.all())]
    horas_habiles = obtener_horas_habiles(anio)

    datos_precargados = precargar_datos(
        anio=anio,
        meses_query=meses_query,
        clientes=Clientes.objects.all(),
        lineas=lineas or Linea.objects.all(),
    )

    costos_por_ceco_modulo_mes = calcular_costos_por_ceco_modulo_mes(
        anio=anio,
        meses_int=meses_int,
        lineas_ids=lineas_ids,
        datos_precargados=datos_precargados,
        horas_habiles=horas_habiles,
    )

    # Descripciones CeCo
    cecos_codigos = facturas.values_list("Ceco", flat=True).distinct()
    descripciones_cecos = dict(
        CentrosCostos.objects.filter(codigoCeCo__in=cecos_codigos).values_list("codigoCeCo", "descripcionCeCo")
    )

    facturas_agrupadas = (
        facturas.values(
            "ConsecutivoId",
            "Mes",
            "Ceco",
            "LineaId__Linea",
            "ModuloId__Modulo",
        )
        .annotate(total_valor=Sum("Valor"))
        .order_by("Ceco", "LineaId__Linea", "ModuloId__Modulo", "Mes")
    )

    for item in facturas_agrupadas:
        item["descripcion_ceco"] = descripciones_cecos.get(item.get("Ceco"), "")

    # ✅ corrección: filtrar por PK de Linea
    lineas_obj = Linea.objects.filter(
        LineaId__in=facturas.values_list("LineaId", flat=True).distinct()
    ).distinct()

    return (
        facturas_agrupadas,
        lineas_obj,
        form.cleaned_data.get("Mes"),
        costos_por_ceco_modulo_mes,
    )


def convert_to_regular_dict(d):
    """Convierte defaultdict a dict regular para el template."""
    if isinstance(d, defaultdict):
        d = {k: convert_to_regular_dict(v) for k, v in d.items()}
    return d


@login_required
@verificar_permiso("can_view_informe_facturacion_centrocostos")
def informe_facturacion_CentroCostos(request):
    # ✅ IMPORTANTE: si no hay GET, form UNBOUND (para que no valide y no filtre)
    form = FacturacionClientesFilterForm(request.GET) if request.GET else FacturacionClientesFilterForm()

    meses_completos = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]

    facturas, lineas_filtradas, meses_filtrados, costos_por_ceco_modulo_mes = filtrar_datos(form)
    meses = [par for par in meses_completos if not meses_filtrados or str(par[0]) in meses_filtrados]

    datos = {}

    totales_por_linea = defaultdict(float)
    totales_por_ceco = defaultdict(float)
    totales_por_modulo = defaultdict(float)
    totales_por_mes = defaultdict(float)
    totales_por_linea_mes = defaultdict(lambda: defaultdict(float))
    totales_por_ceco_linea = defaultdict(lambda: defaultdict(float))
    totales_por_ceco_linea_mes = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    totales_por_ceco_linea_modulo = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    totales_por_ceco_modulo = defaultdict(lambda: defaultdict(float))

    costos_por_ceco = defaultdict(float)
    costos_por_modulo = defaultdict(float)
    costos_por_ceco_modulo = defaultdict(lambda: defaultdict(float))
    costos_por_mes = defaultdict(float)
    costo_global = 0.0

    costos_por_ceco_linea_mes = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    costos_por_ceco_linea = defaultdict(lambda: defaultdict(float))

    # ✅ NUEVO: Totales por CeCo y Mes (para fila TOTAL CECO en tabla)
    totales_por_ceco_mes = defaultdict(lambda: defaultdict(float))
    costos_por_ceco_mes = defaultdict(lambda: defaultdict(float))

    total_global = 0.0

    # Agrupar facturas por ceco + modulo + mes para evitar duplicados
    facturas_agrupadas_por_ceco_modulo_mes = defaultdict(float)
    for item in facturas:
        ceco = item.get("Ceco", "Sin Ceco")
        modulo = item.get("ModuloId__Modulo", "Sin módulo")
        mes = int(item["Mes"]) if item.get("Mes") else 0
        mes_str = str(mes)

        valor = float(item.get("total_valor") or 0)
        if meses_filtrados and str(mes) not in meses_filtrados:
            continue

        key = f"{ceco}_{modulo}_{mes_str}"
        facturas_agrupadas_por_ceco_modulo_mes[key] += valor

    # Procesar facturas agrupadas
    for key, valor in facturas_agrupadas_por_ceco_modulo_mes.items():
        ceco, modulo, mes_str = key.split("_")
        mes = int(mes_str)

        primera_factura = next(
            (
                item for item in facturas
                if item.get("Ceco") == ceco
                and item.get("ModuloId__Modulo") == modulo
                and int(item.get("Mes", 0)) == mes
            ),
            None,
        )
        if not primera_factura:
            continue

        descripcion = primera_factura.get("descripcion_ceco", "")
        linea = primera_factura.get("LineaId__Linea", "Sin línea")

        datos.setdefault(ceco, {}).setdefault(descripcion, {}).setdefault(linea, {}).setdefault(modulo, {})
        datos[ceco][descripcion][linea][modulo].setdefault(mes, {"total_valor": 0.0, "costo": 0.0})

        datos[ceco][descripcion][linea][modulo][mes]["total_valor"] += valor

        # Costo (una sola vez por ceco + modulo + mes)
        costo_mes = float(costos_por_ceco_modulo_mes.get(ceco, {}).get(modulo, {}).get(mes_str, Decimal("0.00")))
        datos[ceco][descripcion][linea][modulo][mes]["costo"] = costo_mes

        # Acumular costos por línea y mes
        costos_por_ceco_linea_mes[ceco][linea][mes] += costo_mes
        costos_por_ceco_linea[ceco][linea] += costo_mes

        # ✅ NUEVO: Totales por CeCo y Mes (facturación y costo)
        totales_por_ceco_mes[ceco][mes] += valor
        costos_por_ceco_mes[ceco][mes] += costo_mes

        # Totales facturación
        totales_por_linea[linea] += valor
        totales_por_ceco[ceco] += valor
        totales_por_modulo[modulo] += valor
        totales_por_mes[mes] += valor
        totales_por_linea_mes[linea][mes] += valor
        totales_por_ceco_linea_modulo[ceco][linea][modulo] += valor
        totales_por_ceco_modulo[ceco][modulo] += valor
        totales_por_ceco_linea[ceco][linea] += valor
        totales_por_ceco_linea_mes[ceco][linea][mes] += valor
        total_global += valor

        # Totales costo
        costos_por_ceco[ceco] += costo_mes
        costos_por_modulo[modulo] += costo_mes
        costos_por_ceco_modulo[ceco][modulo] += costo_mes
        costos_por_mes[mes] += costo_mes
        costo_global += costo_mes

    cecos_con_totales = {}
    for ceco_k, descripciones in datos.items():
        total_filas = 0
        for _, lineas_dict in descripciones.items():
            for _, modulos in lineas_dict.items():
                total_filas += len(modulos)
        cecos_con_totales[ceco_k] = total_filas

    graficos = generar_graficos_facturacion(datos)

    context = {
        "form": form,
        "datos": datos,
        "lineas": lineas_filtradas,
        "meses": meses,
        "cecos_con_totales": cecos_con_totales,

        "totales_por_linea": convert_to_regular_dict(totales_por_linea),
        "totales_por_ceco": convert_to_regular_dict(totales_por_ceco),
        "totales_por_modulo": convert_to_regular_dict(totales_por_modulo),
        "totales_por_mes": convert_to_regular_dict(totales_por_mes),
        "totales_por_linea_mes": convert_to_regular_dict(totales_por_linea_mes),
        "graficos": graficos,
        "totales_por_ceco_linea": convert_to_regular_dict(totales_por_ceco_linea),
        "totales_por_ceco_linea_mes": convert_to_regular_dict(totales_por_ceco_linea_mes),
        "totales_por_ceco_linea_modulo": convert_to_regular_dict(totales_por_ceco_linea_modulo),
        "totales_por_ceco_modulo": convert_to_regular_dict(totales_por_ceco_modulo),
        "total_global": total_global,

        "costos_por_ceco": convert_to_regular_dict(costos_por_ceco),
        "costos_por_modulo": convert_to_regular_dict(costos_por_modulo),
        "costos_por_ceco_modulo": convert_to_regular_dict(costos_por_ceco_modulo),
        "costos_por_mes": convert_to_regular_dict(costos_por_mes),
        "costo_global": costo_global,
        "margen_global": total_global - costo_global,

        "costos_por_ceco_linea": convert_to_regular_dict(costos_por_ceco_linea),
        "costos_por_ceco_linea_mes": convert_to_regular_dict(costos_por_ceco_linea_mes),

        # ✅ NUEVO: para fila TOTAL CECO (por mes)
        "totales_por_ceco_mes": convert_to_regular_dict(totales_por_ceco_mes),
        "costos_por_ceco_mes": convert_to_regular_dict(costos_por_ceco_mes),
    }

    return render(request, "Informes/InformeFacturacionCentroCostosIndex.html", context)


@login_required
@verificar_permiso("can_view_informe_facturacion_centrocostos")
def descargar_reporte_excel_facturacion_clientes(request):
    # ✅ Requerir año para descargar
    anio_raw = (request.GET.get("Anio") or "").strip()
    if anio_raw in ("", "None", "Todos"):
        return HttpResponse("Debes seleccionar un año para descargar el reporte.")

    form = FacturacionClientesFilterForm(request.GET)
    if not form.is_valid():
        return HttpResponse("Parámetros inválidos para generar el reporte (verifica el Año).")

    meses_completos = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]

    try:
        facturas, _lineas_filtradas, meses_filtrados, costos_por_ceco_modulo_mes = filtrar_datos(form)
    except Exception as e:
        return HttpResponse(f"Error al obtener datos: {str(e)}")

    if not facturas:
        return HttpResponse("No hay datos disponibles para los filtros seleccionados.")

    meses = meses_completos
    if meses_filtrados:
        meses = [par for par in meses_completos if str(par[0]) in meses_filtrados]

    # Agrupar facturas por ceco + modulo + mes para evitar duplicados
    facturas_agrupadas_por_ceco_modulo_mes = defaultdict(float)
    for item in facturas:
        ceco = item.get("Ceco", "Sin Ceco")
        modulo = item.get("ModuloId__Modulo", "Sin módulo")
        mes = int(item["Mes"]) if item.get("Mes") else 0
        mes_str = str(mes)

        valor = float(item.get("total_valor") or 0)
        if meses_filtrados and str(mes) not in meses_filtrados:
            continue

        key = f"{ceco}_{modulo}_{mes_str}"
        facturas_agrupadas_por_ceco_modulo_mes[key] += valor

    datos = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict))))
    totales_por_mes = defaultdict(float)
    costos_por_mes = defaultdict(float)
    total_global = 0.0
    costo_global = 0.0

    # ✅ NUEVO: totales por CECO (para fila TOTAL CECO en Excel)
    totales_por_ceco = defaultdict(float)
    costos_por_ceco = defaultdict(float)
    totales_por_ceco_mes = defaultdict(lambda: defaultdict(float))
    costos_por_ceco_mes = defaultdict(lambda: defaultdict(float))

    for key, valor in facturas_agrupadas_por_ceco_modulo_mes.items():
        ceco, modulo, mes_str = key.split("_")
        mes = int(mes_str)

        primera_factura = next(
            (
                item for item in facturas
                if item.get("Ceco") == ceco
                and item.get("ModuloId__Modulo") == modulo
                and int(item.get("Mes", 0)) == mes
            ),
            None,
        )
        if not primera_factura:
            continue

        descripcion = primera_factura.get("descripcion_ceco", "")
        linea = primera_factura.get("LineaId__Linea", "Sin línea")

        costo = float(costos_por_ceco_modulo_mes.get(ceco, {}).get(modulo, {}).get(mes_str, Decimal("0.00")))

        if mes not in datos[ceco][descripcion][linea][modulo]:
            datos[ceco][descripcion][linea][modulo][mes] = {"facturacion": 0.0, "costo": 0.0, "margen": 0.0}

        datos[ceco][descripcion][linea][modulo][mes]["facturacion"] += valor
        datos[ceco][descripcion][linea][modulo][mes]["costo"] = costo
        datos[ceco][descripcion][linea][modulo][mes]["margen"] = valor - costo

        totales_por_mes[mes] += valor
        costos_por_mes[mes] += costo
        total_global += valor
        costo_global += costo

        # ✅ NUEVO acumulado por CECO
        totales_por_ceco[ceco] += valor
        costos_por_ceco[ceco] += costo
        totales_por_ceco_mes[ceco][mes] += valor
        costos_por_ceco_mes[ceco][mes] += costo

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturación Centro Costos"

    header_style = {
        "font": Font(bold=True, color="FFFFFF", size=12),
        "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }
    data_style = {
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "number_format": "#,##0.00",
    }
    global_total_style = {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "number_format": "#,##0.00",
    }
    margen_style = {
        "font": Font(bold=True, color="000000"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "number_format": "#,##0.00",
    }

    # ✅ NUEVO estilo TOTAL CECO
    ceco_total_style = {
        "font": Font(bold=True, color="000000"),
        "fill": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "number_format": "#,##0.00",
    }
    ceco_total_label_style = {
        **ceco_total_style,
        "alignment": Alignment(horizontal="center", vertical="center"),
    }

    columns_fijas = [("Ceco", 15), ("Descripción", 30), ("Línea", 20), ("Módulo", 20)]

    columnas_por_mes = []
    for mes_num, mes_nombre in meses:
        columnas_por_mes.extend(
            [
                (f"{mes_nombre} Facturación", 15),
                (f"{mes_nombre} Costo", 15),
                (f"{mes_nombre} Margen", 15),
            ]
        )

    # Encabezados
    col_num = 1
    for col_title, width in columns_fijas:
        cell = ws.cell(row=1, column=col_num, value=col_title)
        for attr, value in header_style.items():
            setattr(cell, attr, value)
        ws.column_dimensions[get_column_letter(col_num)].width = width
        col_num += 1

    for mes_col_title, width in columnas_por_mes:
        cell = ws.cell(row=1, column=col_num, value=mes_col_title)
        for attr, value in header_style.items():
            setattr(cell, attr, value)
        ws.column_dimensions[get_column_letter(col_num)].width = width
        col_num += 1

    total_columns = [("Total Facturación", 15), ("Total Costo", 15), ("Total Margen", 15)]
    for col_title, width in total_columns:
        cell = ws.cell(row=1, column=col_num, value=col_title)
        for attr, value in header_style.items():
            setattr(cell, attr, value)
        ws.column_dimensions[get_column_letter(col_num)].width = width
        col_num += 1

    # Datos
    row_num = 2
    for ceco, descripciones in datos.items():
        ceco_start_row = row_num

        for descripcion, lineas in descripciones.items():
            desc_start_row = row_num

            for linea, modulos in lineas.items():
                linea_start_row = row_num

                for modulo, meses_data in modulos.items():
                    ws.cell(row=row_num, column=1, value=ceco if row_num == ceco_start_row else "")
                    ws.cell(row=row_num, column=2, value=descripcion if row_num == desc_start_row else "")
                    ws.cell(row=row_num, column=3, value=linea if row_num == linea_start_row else "")
                    ws.cell(row=row_num, column=4, value=modulo)

                    for i in range(1, 5):
                        cell = ws.cell(row=row_num, column=i)
                        for attr, value in data_style.items():
                            if attr != "number_format":
                                setattr(cell, attr, value)

                    col_num_data = 5
                    total_facturacion_fila = 0.0
                    total_costo_fila = 0.0
                    total_margen_fila = 0.0

                    for mes_num, _mes_nombre in meses:
                        mes_data = meses_data.get(mes_num, {"facturacion": 0.0, "costo": 0.0, "margen": 0.0})
                        facturacion = float(mes_data["facturacion"])
                        costo = float(mes_data["costo"])
                        margen = float(mes_data["margen"])

                        cell = ws.cell(row=row_num, column=col_num_data, value=facturacion)
                        for attr, value in data_style.items():
                            setattr(cell, attr, value)
                        col_num_data += 1

                        cell = ws.cell(row=row_num, column=col_num_data, value=costo)
                        for attr, value in data_style.items():
                            setattr(cell, attr, value)
                        col_num_data += 1

                        cell = ws.cell(row=row_num, column=col_num_data, value=margen)
                        for attr, value in margen_style.items():
                            setattr(cell, attr, value)
                        col_num_data += 1

                        total_facturacion_fila += facturacion
                        total_costo_fila += costo
                        total_margen_fila += margen

                    cell = ws.cell(row=row_num, column=col_num_data, value=total_facturacion_fila)
                    for attr, value in data_style.items():
                        setattr(cell, attr, value)
                    col_num_data += 1

                    cell = ws.cell(row=row_num, column=col_num_data, value=total_costo_fila)
                    for attr, value in data_style.items():
                        setattr(cell, attr, value)
                    col_num_data += 1

                    cell = ws.cell(row=row_num, column=col_num_data, value=total_margen_fila)
                    for attr, value in margen_style.items():
                        setattr(cell, attr, value)

                    row_num += 1

        # ✅ NUEVO: TOTAL POR CECO (después de terminar todas las filas del CECO)
        ws.cell(row=row_num, column=1, value=f"TOTAL {ceco}")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)

        for c in range(1, 5):
            cell = ws.cell(row=row_num, column=c)
            for attr, value in ceco_total_label_style.items():
                setattr(cell, attr, value)

        col_num_data = 5
        total_fact_ceco = 0.0
        total_cost_ceco = 0.0

        for mes_num, _mes_nombre in meses:
            fact_mes = float(totales_por_ceco_mes[ceco].get(mes_num, 0.0))
            cost_mes = float(costos_por_ceco_mes[ceco].get(mes_num, 0.0))
            marg_mes = fact_mes - cost_mes

            cell = ws.cell(row=row_num, column=col_num_data, value=fact_mes)
            for attr, value in ceco_total_style.items():
                setattr(cell, attr, value)
            col_num_data += 1

            cell = ws.cell(row=row_num, column=col_num_data, value=cost_mes)
            for attr, value in ceco_total_style.items():
                setattr(cell, attr, value)
            col_num_data += 1

            cell = ws.cell(row=row_num, column=col_num_data, value=marg_mes)
            for attr, value in ceco_total_style.items():
                setattr(cell, attr, value)
            col_num_data += 1

            total_fact_ceco += fact_mes
            total_cost_ceco += cost_mes

        total_marg_ceco = total_fact_ceco - total_cost_ceco

        cell = ws.cell(row=row_num, column=col_num_data, value=total_fact_ceco)
        for attr, value in ceco_total_style.items():
            setattr(cell, attr, value)
        col_num_data += 1

        cell = ws.cell(row=row_num, column=col_num_data, value=total_cost_ceco)
        for attr, value in ceco_total_style.items():
            setattr(cell, attr, value)
        col_num_data += 1

        cell = ws.cell(row=row_num, column=col_num_data, value=total_marg_ceco)
        for attr, value in ceco_total_style.items():
            setattr(cell, attr, value)

        row_num += 1

    # Total global
    if datos:
        ws.cell(row=row_num, column=1, value="TOTAL GLOBAL")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)

        # Estilo label total global
        for c in range(1, 5):
            cell = ws.cell(row=row_num, column=c)
            for attr, value in global_total_style.items():
                if attr != "number_format":
                    setattr(cell, attr, value)

        col_num_data = 5

        for mes_num, _mes_nombre in meses:
            cell = ws.cell(row=row_num, column=col_num_data, value=totales_por_mes.get(mes_num, 0.0))
            for attr, value in global_total_style.items():
                setattr(cell, attr, value)
            col_num_data += 1

            cell = ws.cell(row=row_num, column=col_num_data, value=costos_por_mes.get(mes_num, 0.0))
            for attr, value in global_total_style.items():
                setattr(cell, attr, value)
            col_num_data += 1

            margen_mes = totales_por_mes.get(mes_num, 0.0) - costos_por_mes.get(mes_num, 0.0)
            cell = ws.cell(row=row_num, column=col_num_data, value=margen_mes)
            for attr, value in global_total_style.items():
                setattr(cell, attr, value)
            col_num_data += 1

        cell = ws.cell(row=row_num, column=col_num_data, value=total_global)
        for attr, value in global_total_style.items():
            setattr(cell, attr, value)
        col_num_data += 1

        cell = ws.cell(row=row_num, column=col_num_data, value=costo_global)
        for attr, value in global_total_style.items():
            setattr(cell, attr, value)
        col_num_data += 1

        cell = ws.cell(row=row_num, column=col_num_data, value=total_global - costo_global)
        for attr, value in global_total_style.items():
            setattr(cell, attr, value)

    ws.freeze_panes = "A2"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"facturacion_clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ----------------------------
# Gráficos
# ----------------------------
def generar_graficos_facturacion(datos):
    """Genera SOLO 2 gráficos de barras: totales por CeCo y totales por módulo."""
    graficos = []

    # 1) Totales por CeCo
    cecos_labels = []
    cecos_data = []

    for ceco, descripciones in datos.items():
        descripcion = next(iter(descripciones.keys())) if descripciones else ""
        total_ceco = 0.0

        for linea, modulos in descripciones.get(descripcion, {}).items():
            for _modulo, meses_data in modulos.items():
                for _mes, valores in meses_data.items():
                    total_ceco += float(valores.get("total_valor", 0))

        if total_ceco > 0:
            cecos_labels.append(f"{ceco} - {descripcion}")
            cecos_data.append(total_ceco)

    if cecos_data:
        graficos.append(
            {
                "tipo": "cecos",
                "nombre": "Totales por Centro de Costos",
                "config": {
                    "tipo_grafico": "bar",
                    "labels": cecos_labels,
                    "datasets": [
                        {
                            "label": "Total Facturado",
                            "data": cecos_data,
                            "backgroundColor": "rgba(54, 162, 235, 0.7)",
                            "borderColor": "rgba(54, 162, 235, 1)",
                            "borderWidth": 1,
                        }
                    ],
                },
            }
        )

    # 2) Totales por Módulo con colores por descripción
    modulos_data = []
    modulos_labels = []
    modulos_descripciones = []
    descripciones_colores = {}
    colores_disponibles = [
        "rgba(255, 99, 132, 0.7)",
        "rgba(54, 162, 235, 0.7)",
        "rgba(255, 206, 86, 0.7)",
        "rgba(75, 192, 192, 0.7)",
        "rgba(153, 102, 255, 0.7)",
        "rgba(255, 159, 64, 0.7)",
        "rgba(199, 199, 199, 0.7)",
    ]

    descripciones_unicas = set()
    for _ceco, descripciones in datos.items():
        descripcion = next(iter(descripciones.keys())) if descripciones else "Sin descripción"
        descripciones_unicas.add(descripcion)

    for i, descripcion in enumerate(descripciones_unicas):
        descripciones_colores[descripcion] = colores_disponibles[i % len(colores_disponibles)]

    background_colors = []
    for _ceco, descripciones in datos.items():
        descripcion = next(iter(descripciones.keys())) if descripciones else "Sin descripción"
        color_desc = descripciones_colores.get(descripcion) or colores_disponibles[len(descripciones_colores) % len(colores_disponibles)]
        descripciones_colores.setdefault(descripcion, color_desc)

        for _linea, modulos in descripciones.get(descripcion, {}).items():
            for modulo, meses_data in modulos.items():
                total = sum(float(m.get("total_valor", 0)) for m in meses_data.values())
                if total > 0:
                    modulos_labels.append(modulo)
                    modulos_data.append(total)
                    modulos_descripciones.append(descripcion)
                    background_colors.append(descripciones_colores[descripcion])

    if modulos_data:
        datasets = [
            {
                "label": "Total Facturado",
                "data": modulos_data,
                "backgroundColor": background_colors,
                "borderColor": [c.replace("0.7", "1") for c in background_colors],
                "borderWidth": 1,
                "descripciones": modulos_descripciones,
            }
        ]

        legend_datasets = []
        for descripcion, color in descripciones_colores.items():
            legend_datasets.append(
                {
                    "label": descripcion,
                    "data": [],
                    "backgroundColor": color,
                    "borderColor": color.replace("0.7", "1"),
                    "borderWidth": 1,
                }
            )

        graficos.append(
            {
                "tipo": "modulos",
                "nombre": "Totales por Módulo (Colores por Descripción)",
                "config": {
                    "tipo_grafico": "bar",
                    "labels": modulos_labels,
                    "descripciones_colores": descripciones_colores,
                    "datasets": datasets,
                    "legend_datasets": legend_datasets,
                },
            }
        )

    return graficos
