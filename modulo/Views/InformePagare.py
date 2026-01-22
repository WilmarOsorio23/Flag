import logging
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

from dateutil.relativedelta import relativedelta

from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_date, parse_datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border

from modulo.forms import EmpleadoConPagareFilterForm
from modulo.models import Empleado, Pagare
from modulo.decorators import verificar_permiso

logger = logging.getLogger(__name__)


# ----------------------------
# Helpers robustos
# ----------------------------
INVALID_DATE_STRINGS = {
    "0",
    "0000-00-00",
    "0000-00-00 00:00:00",
    "0000-00-00T00:00:00",
    "0000-00-00T00:00:00Z",
    "00/00/0000",
    "0000/00/00",
    "null",
    "none",
    "N/A",
    "-",
}

def to_date(value):
    """
    Convierte date/datetime/str a date.
    Tolerante con basura tipo 0000-00-00.
    """
    if not value:
        return None

    # ya es date (pero no datetime)
    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    # datetime -> date
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None

        low = s.lower()
        if low in INVALID_DATE_STRINGS:
            return None

        # ISO date
        try:
            d = parse_date(s)
            if d:
                return d
        except (ValueError, TypeError):
            pass

        # ISO datetime
        try:
            dt = parse_datetime(s)
            if dt:
                return dt.date()
        except (ValueError, TypeError):
            pass

        # formatos comunes
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue

    return None


def fmt_date(value, fmt="%Y-%m-%d"):
    d = to_date(value)
    return d.strftime(fmt) if d else ""


def to_decimal(value, default=Decimal("0")):
    try:
        if value is None:
            return default
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return default


def calcular_meses_condonados(fecha_inicio):
    """
    Calcula meses condonados según la fecha inicio condonación.
    Tolera fecha como str.
    """
    fecha_inicio = to_date(fecha_inicio)
    if not fecha_inicio:
        return 0

    hoy = date.today()
    if fecha_inicio > hoy:
        return 0

    diferencia = relativedelta(hoy, fecha_inicio)
    meses = diferencia.years * 12 + diferencia.months

    # Si el día actual >= día inicio, cuenta mes actual
    if hoy.day >= fecha_inicio.day:
        meses += 1

    return max(0, meses)


def filtrar_pagares(form, pagares_qs):
    empleados = form.cleaned_data.get("empleados_con_pagare")
    anio = form.cleaned_data.get("Anio")
    lineas = form.cleaned_data.get("LineaId")
    tipos = form.cleaned_data.get("tipo_pagare")
    estado = form.cleaned_data.get("estado_pagare")

    if empleados:
        pagares_qs = pagares_qs.filter(Documento__in=empleados.values_list("Documento", flat=True))

    if anio:
        pagares_qs = pagares_qs.filter(Fecha_Creacion_Pagare__year=anio)

    if lineas:
        docs_linea = Empleado.objects.filter(LineaId__in=lineas).values_list("Documento", flat=True)
        pagares_qs = pagares_qs.filter(Documento__in=docs_linea)

    if tipos:
        pagares_qs = pagares_qs.filter(Tipo_Pagare__in=tipos)

    # ✅ estado robusto (case-insensitive)
    if estado:
        pagares_qs = pagares_qs.filter(estado__iexact=estado)

    return pagares_qs


@login_required
@verificar_permiso("can_view_informe_pagares")
def informe_pagares(request):
    pagares_info = []
    show_data = False
    fecha_actual = datetime.now().strftime("%Y-%m-%d")

    form = EmpleadoConPagareFilterForm(request.GET or None)

    pagares_qs = Pagare.objects.select_related("Tipo_Pagare").all()

    if form.is_valid():
        pagares_qs = filtrar_pagares(form, pagares_qs)

        documentos = list(pagares_qs.values_list("Documento", flat=True).distinct())
        empleados = (
            Empleado.objects.filter(Documento__in=documentos)
            .select_related("LineaId", "CargoId")
        )
        empleados_dict = {str(emp.Documento): emp for emp in empleados}

        for pagare in pagares_qs:
            empleado = empleados_dict.get(str(pagare.Documento))
            if not empleado:
                continue

            # Condonación
            fecha_inicio_cond = to_date(pagare.Fecha_inicio_Condonacion)
            meses_total = int(pagare.Meses_de_condonacion or 0)

            meses_condonados = 0
            if fecha_inicio_cond and meses_total > 0:
                meses_condonados = min(calcular_meses_condonados(fecha_inicio_cond), meses_total)

            meses_por_condonar = max(0, meses_total - meses_condonados)

            valor_pagare = to_decimal(pagare.Valor_Pagare)
            valor_condonado = Decimal("0")
            if meses_total > 0:
                valor_condonado = (valor_pagare / Decimal(meses_total)) * Decimal(meses_condonados)

            deuda_pagare = max(Decimal("0"), valor_pagare - valor_condonado)

            registro = {
                "Documento": empleado.Documento,
                "Nombre": empleado.Nombre,
                "TipoPagare": pagare.Tipo_Pagare.Desc_Tipo_Pagare if pagare.Tipo_Pagare else "N/A",
                "Descripcion": pagare.descripcion,
                "Estado": pagare.estado,

                "Linea": empleado.LineaId.Linea if getattr(empleado, "LineaId", None) else "N/A",
                "Cargo": empleado.CargoId.Cargo if getattr(empleado, "CargoId", None) else "N/A",

                # ✅ fechas robustas (evita .strftime en str)
                "FechaIngreso": fmt_date(empleado.FechaIngreso),
                "FechaOperacion": fmt_date(empleado.FechaOperacion),
                "FechaRetiro": fmt_date(empleado.FechaRetiro),

                "ConsecutivoPagare": pagare.Pagare_Id,
                "FechaPagare": fmt_date(pagare.Fecha_Creacion_Pagare),
                "FechaInicioCondonacion": fmt_date(pagare.Fecha_inicio_Condonacion),
                "FechaFinCondonacion": fmt_date(pagare.Fecha_fin_Condonacion),

                "ValorPagare": float(valor_pagare),
                "MesesCondonacion": meses_total,
                "MesesCondonados": meses_condonados,
                "MesesPorCondonar": meses_por_condonar,

                # % ejecución: usa el campo real del pagaré (por horas), si existe
                "PorcentajeEjecucion": float(pagare.porcentaje_ejecucion or 0),

                "ValorCondonado": float(valor_condonado),
                "DeudaPagare": float(deuda_pagare),

                "pagare_id": pagare.Pagare_Id,
            }
            pagares_info.append(registro)

        show_data = bool(pagares_info)
        request.session["pagares_info"] = pagares_info
    else:
        logger.debug("Errores en el formulario Informe Pagaré: %s", form.errors)

    context = {
        "form": form,
        "pagares_info": pagares_info,
        "show_data": show_data,
        "mensaje": "No se encontraron resultados para los filtros aplicados" if not pagares_info else "",
        "fecha_actual": fecha_actual,
    }
    return render(request, 'Pagare/InformePagare.html', context)


# ----------------------------
# Exportar Excel (se mantiene tu lógica; no toqué lo grande salvo lo necesario)
# ----------------------------
@login_required
@verificar_permiso("can_view_informe_pagares")
def exportar_pagares_excel(request):
    pagares_info = request.session.get("pagares_info", [])
    logger.debug("Pagarés en sesión: %s", len(pagares_info))

    if not pagares_info:
        return HttpResponse("No hay datos para exportar.")

    export_all = request.GET.get("export_all", "0") == "1"
    selected_ids_str = request.GET.get("selected_ids", "")
    selected_ids = selected_ids_str.split(",") if selected_ids_str else []
    fecha_informe = request.GET.get("fecha_informe", datetime.now().strftime("%Y-%m-%d"))

    if not export_all and selected_ids:
        pagares_info = [p for p in pagares_info if str(p.get("ConsecutivoPagare", "")) in selected_ids]

    if not pagares_info:
        return HttpResponse("No hay datos para exportar.")

    wb = Workbook()
    if len(pagares_info) > 1 or export_all:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    columnas = [
        "Documento", "Nombre", "Tipo Pagaré", "Descripción", "Línea", "Cargo",
        "Fecha Ingreso", "Fecha Inicio Operación", "Consecutivo Pagare",
        "Fecha Pagare", "Fecha Inicio Condonación", "Fecha Fin Condonación",
        "Valor del Pagaré", "Valor Condonado", "Deuda Pagare",
        "Meses Condonación", "Meses Condonados", "Meses por Condonar",
        "%Ejecución", "Fecha Retiro",
    ]

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    def crear_hoja_resumen(ws, registros, fecha):
        ws.cell(row=1, column=1, value="Fecha del informe:").font = header_font
        ws.cell(row=1, column=2, value=fecha)

        for col_num, header in enumerate(columnas, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num = 4
        for r in registros:
            vals = [
                r.get("Documento",""), r.get("Nombre",""), r.get("TipoPagare",""), r.get("Descripcion",""),
                r.get("Linea",""), r.get("Cargo",""), r.get("FechaIngreso",""), r.get("FechaOperacion",""),
                r.get("ConsecutivoPagare",""), r.get("FechaPagare",""), r.get("FechaInicioCondonacion",""),
                r.get("FechaFinCondonacion",""), r.get("ValorPagare",0), r.get("ValorCondonado",0),
                r.get("DeudaPagare",0), r.get("MesesCondonacion",0), r.get("MesesCondonados",0),
                r.get("MesesPorCondonar",0), r.get("PorcentajeEjecucion",0), r.get("FechaRetiro",""),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=c, value=v)
                cell.border = border
            row_num += 1

        auto_width(ws)
        return ws

    # ✅ Mantengo tu lógica de “una hoja vs muchas”
    if len(pagares_info) == 1 and not export_all:
        ws = wb.active
        ws.title = "Pagare"
        crear_hoja_resumen(ws, pagares_info, fecha_informe)
    else:
        ws_resumen = wb.create_sheet(title="Resumen")
        crear_hoja_resumen(ws_resumen, pagares_info, fecha_informe)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="informe_pagares_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response
