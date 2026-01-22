# Informe de Certificación de Empleados

# Librerías estándar y de terceros
from datetime import datetime

# Django
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

# Librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

# Modelos y formularios del módulo
from modulo.forms import EmpleadoFilterForm
from modulo.models import Certificacion, Detalle_Certificacion, Empleado

# Utilidad para conteo de elementos
from collections import Counter
from modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required

def filtrar_empleado(form, empleados, certificaciones, detalles):
    """
    Aplica los filtros del formulario a los conjuntos de empleados y certificaciones.
    - Filtra por nombre de empleado (como texto parcial).
    - Filtra por línea (ID).
    - Filtra por certificación (ID).
    Devuelve los empleados con certificaciones asociadas según los filtros.
    """

    linea = form.cleaned_data.get('LineaId')
    certificacion = form.cleaned_data.get('Certificacion')
    nombre = form.cleaned_data.get('Nombre')

    if nombre:
        empleados = empleados.filter(Nombre__icontains=nombre)
    if linea:
        empleados = empleados.filter(LineaId=linea)
    if certificacion:
        detalles = detalles.filter(CertificacionId=certificacion)

    empleados_con_certificacion = empleados.filter(
        Documento__in=detalles.values_list('Documento__Documento', flat=True)
    )
    
    return empleados_con_certificacion, certificaciones, detalles

def obtener_informe_certificacion(certificaciones, empleados, detalles):
    """
    Une información entre empleados y detalles de certificaciones.
    Devuelve una lista de diccionarios con los datos organizados para visualización.
    """
    certificaciones_info = []
    detalles = detalles.select_related('Documento', 'CertificacionId')
    empleados_dict = {empleado.Documento: empleado for empleado in empleados}

    for detalle in detalles:
        empleado = empleados_dict.get(detalle.Documento.Documento)
        if empleado:
            certificaciones_info.append({
                'Documento': empleado.Documento,
                'Nombre': empleado.Nombre,
                'Linea': empleado.LineaId.Linea if empleado.LineaId else '',
                'Cargo': empleado.CargoId.Cargo if empleado.CargoId else '',
                'Perfil': empleado.PerfilId.Perfil if empleado.PerfilId else '',
                'Modulo': empleado.ModuloId.Modulo if empleado.ModuloId else '',
                'CertificacionId': detalle.CertificacionId.CertificacionId,
                'Certificacion': detalle.CertificacionId.Certificacion,
                'Fecha_Certificacion': detalle.Fecha_Certificacion,
            })
    return certificaciones_info

@login_required
@verificar_permiso('can_view_informe_certificacion')
def empleado_filtrado(request):
    """
    Vista que construye el informe HTML con filtros, tabla de datos y resumenes (cards).
    - Usa GET para aplicar filtros.
    - Genera contexto con lista filtrada y datos agregados.
    """
    empleados = Empleado.objects.all()
    certificaciones = Certificacion.objects.all()
    detalles = Detalle_Certificacion.objects.all()
    certificaciones_info = []
    show_data = False
    busqueda_realizada = False

    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)

        if request.GET:
            busqueda_realizada = True
            if form.is_valid():
                empleados, certificaciones, detalles = filtrar_empleado(form, empleados, certificaciones, detalles)
                certificaciones_info = obtener_informe_certificacion(certificaciones, empleados, detalles)
                show_data = bool(certificaciones_info)
    else:
        form = EmpleadoFilterForm()

    #Cálculo de Cards
    total_certificaciones = len(certificaciones_info)
    modulos = [cert['Modulo'] for cert in certificaciones_info if cert['Modulo']]
    conteo_modulos = dict(Counter(modulos))
    lineas = [cert['Linea'] for cert in certificaciones_info if cert['Linea']]
    conteo_lineas = dict(Counter(lineas))

    context = {
        'form': form,
        'certificaciones_info': certificaciones_info,
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'total_certificaciones': total_certificaciones,
        'conteo_modulos': conteo_modulos,
        'conteo_lineas': conteo_lineas,
        'modulos_labels': list(conteo_modulos.keys()),
        'modulos_data': list(conteo_modulos.values()),
        'lineas_labels': list(conteo_lineas.keys()),
        'lineas_data': list(conteo_lineas.values()),
        'mensaje': "No se encontraron resultados para los filtros aplicados." if busqueda_realizada and not show_data else "No se ha realizado ninguna búsqueda aún."
    }

    return render(request, 'Informes/InformescertificacionIndex.html', context)

@login_required
@verificar_permiso('can_view_informe_certificacion')
@csrf_exempt
def exportar_certificaciones_excel(request):
    """
    Exporta los datos filtrados a un archivo Excel.
    - Aplica filtros mediante el mismo formulario.
    - Reutiliza la lógica del informe.
    - Formatea y decora el archivo generado con OpenPyXL.
    """
    certificaciones_info = []  
    

    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)
        if form.is_valid():
            certificaciones = Certificacion.objects.all()
            empleados = Empleado.objects.all()
            detalles = Detalle_Certificacion.objects.all()

            empleados, certificaciones, detalles = filtrar_empleado(form, empleados, certificaciones, detalles)
            certificaciones_info = obtener_informe_certificacion(certificaciones, empleados, detalles)

        if not certificaciones_info:
            return HttpResponse("No se encontraron datos para exportar.", status=404)

        if not form.is_valid():
            return HttpResponse("Filtros no válidos. Por favor, revisa los criterios ingresados.", status=400)

    # Prepara datos para exportación
    data = []
    for cert in certificaciones_info:
        fila = {
            'Nombre Colaborador': cert['Nombre'],
            'Documento': cert['Documento'],
            'Módulo': cert['Modulo'],
            'Perfil': cert['Perfil'],
            'Línea': cert['Linea'],
            'Cargo': cert['Cargo'],
            'ID Certificación': cert['CertificacionId'],
            'Certificación': cert['Certificacion'],
            'Fecha Certificación': cert['Fecha_Certificacion'],
        }
        data.append(fila)

    # Crear el archivo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificaciones"

    # Agregar encabezados
    for col in data[0].keys():
        cell = ws.cell(row=1, column=list(data[0].keys()).index(col) + 1, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Agregar datos al Excel
    for r_idx, row in enumerate(data, 2):
        for c_idx, value in enumerate(row.values(), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')

    # Ancho automático y bordes
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 2
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Generar archivo para descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"certificaciones_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response