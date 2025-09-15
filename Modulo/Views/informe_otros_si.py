#Informe Contratos Otros Si

# Librerías estándar y de terceros
from collections import defaultdict
from datetime import datetime

# Django
from django.http import HttpResponse
from django.shortcuts import render
from django.http import JsonResponse

# Formularios y modelos del módulo
from Modulo.forms import OtrosSiFilterForm
from Modulo.models import Clientes, ContratosOtrosSi, ClientesContratos
from Modulo.models import Clientes

# Librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from Modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required


def get_cliente_id_by_nombre(request):
    """
    Vista auxiliar que devuelve el ID de un cliente dado su nombre.

    Parámetros:
        request (HttpRequest): La solicitud GET con el nombre del cliente.

    Retorna:
        JsonResponse: Contiene el ID del cliente si se encuentra, o un error 404 si no existe.
    """
    nombre = request.GET.get('nombre')
    cliente = Clientes.objects.filter(Nombre_Cliente=nombre).first()
    if cliente:
        return JsonResponse({'cliente_id': cliente.ClienteId})
    return JsonResponse({'error': 'Cliente no encontrado'}, status=404)

# Filtrar OtrosSí por cliente y fecha de inicio
def filtrar_otros_si(form, clientes, otros_si):
    """
    Aplica filtros al queryset de clientes y contratos Otros Sí basados en los datos del formulario.

    Parámetros:
        form (OtrosSiFilterForm): Formulario con los datos de filtro.
        clientes (QuerySet): QuerySet inicial de clientes.
        otros_si (QuerySet): QuerySet inicial de contratos Otros Sí.

    Retorna:
        tuple: Tupla con los clientes y contratos Otros Sí filtrados.
    """
    clientes = clientes.order_by('Nombre_Cliente')
    nombre = form.cleaned_data.get('Nombre_Cliente')
    fecha_inicio = form.cleaned_data.get('FechaInicio')
    contrato = form.cleaned_data.get('Contrato')
    contrato_vigente = form.cleaned_data.get('ContratoVigente') 

    if nombre:
        clientes = clientes.filter(Nombre_Cliente=nombre)
    if fecha_inicio:
        otros_si = otros_si.filter(FechaInicio=fecha_inicio)
    if contrato:
        otros_si = otros_si.filter(Contrato__iexact=contrato.strip())

    if contrato_vigente in ('True', 'False'):
        # Obtener contratos de ClientesContratos que tengan ese estado
        contratos_filtrados = ClientesContratos.objects.filter(
            ContratoVigente=(contrato_vigente == 'True')
        ).values_list('Contrato', flat=True)

        # Filtrar ContratosOtrosSi por los contratos que coincidan
        otros_si = otros_si.filter(Contrato__in=contratos_filtrados)
    
    cliente_ids = otros_si.values_list('ClienteId', flat=True)
    clientes = clientes.filter(ClienteId__in=cliente_ids)

    return clientes, otros_si

# Obtener información del informe
def obtener_otros_si(clientes, otros_si):
    """
    Agrupa los contratos Otros Sí por cliente y estructura la información para la vista.

    Parámetros:
        clientes (QuerySet): Lista de clientes filtrados.
        otros_si (QuerySet): Lista de contratos Otros Sí filtrados.

    Retorna:
        list: Lista de diccionarios con los datos del cliente y sus contratos Otros Sí.
    """
    cliente_otros_si_info = []
    otros_si_por_cliente = defaultdict(list)

    for registro in otros_si:
        otros_si_por_cliente[registro.ClienteId_id].append(registro)

    for cliente in clientes:
        otros_cliente = otros_si_por_cliente.get(cliente.ClienteId, [])

        datos_cliente = {
            'documento': cliente.DocumentoId,
            'nombre': cliente.Nombre_Cliente,
            'otros_si': [{
                'FechaInicio': registro.FechaInicio,
                'FechaFin': registro.FechaFin,
                'por_vencer': registro.FechaFin and (registro.FechaFin - datetime.today().date()).days <= 30,
                'NumeroOtroSi': registro.NumeroOtroSi,
                'ValorOtroSi': registro.ValorOtroSi,
                'ValorIncluyeIva': registro.ValorIncluyeIva,
                'Polizas': registro.Polizas,
                'PolizasDesc': registro.PolizasDesc,
                'FirmadoFlag': registro.FirmadoFlag,
                'FirmadoCliente': registro.FirmadoCliente,
                'Contrato': registro.Contrato,
                'moneda': registro.monedaId.Nombre if registro.monedaId else "Sin Moneda"
            } for registro in otros_cliente]
        }

        cliente_otros_si_info.append(datos_cliente)

    return cliente_otros_si_info

@login_required
@verificar_permiso('can_manage_informe_otros_si')
def informe_otros_si(request):
    """
    Vista principal que renderiza el informe HTML de contratos Otros Sí.

    Parámetros:
        request (HttpRequest): Solicitud GET con los filtros.

    Retorna:
        HttpResponse: Renderiza el template con el contexto que incluye resultados, métricas y formularios.
    """
    cliente_otros_si_info = []
    show_data = False
    busqueda_realizada = False

    if request.method == 'GET':
        form = OtrosSiFilterForm(request.GET)

        if request.GET:
            busqueda_realizada = True
            if form.is_valid():
                clientes = Clientes.objects.only('Nombre_Cliente', 'ClienteId').filter(Activo=True)
                otros_si = ContratosOtrosSi.objects.select_related('ClienteId', 'monedaId').all()

                clientes, otros_si = filtrar_otros_si(form, clientes, otros_si)
                cliente_otros_si_info = obtener_otros_si(clientes, otros_si)
                show_data = bool(cliente_otros_si_info)
        else:
            print("Errores del formulario:", form.errors)
    else:
        form = OtrosSiFilterForm()

    total_contratos = 0
    valor_total = 0
    contratos_firmados = 0
    contratos_firmados_cliente = 0
    contratos_con_polizas = 0
    total_usd = total_cop = total_mxn = 0

    for cliente in cliente_otros_si_info:
        for otro in cliente['otros_si']:
            total_contratos += 1
            valor_total += otro['ValorOtroSi'] or 0
            if otro['FirmadoFlag']:
                contratos_firmados += 1
            if otro['FirmadoCliente']:
                contratos_firmados_cliente += 1
            if otro['Polizas']:
                contratos_con_polizas += 1

    porcentaje_firmado_cliente = porcentaje_firmado_interno = porcentaje_con_polizas = 0
    if total_contratos > 0:
        porcentaje_firmado_cliente = round((contratos_firmados_cliente / total_contratos) * 100)
        porcentaje_firmado_interno = round((contratos_firmados / total_contratos) * 100)
        porcentaje_con_polizas = round((contratos_con_polizas / total_contratos) * 100)

    for cliente in cliente_otros_si_info:
        for otro in cliente['otros_si']:
            moneda = otro['moneda'].upper()
            if moneda == "USD":
                total_usd += 1
            elif moneda == "COP":
                total_cop += 1
            elif moneda == "MXN":
                total_mxn += 1

    porcentaje_usd = porcentaje_cop = porcentaje_mxn = 0
    if total_contratos > 0:
        porcentaje_usd = round((total_usd / total_contratos) * 100)
        porcentaje_cop = round((total_cop / total_contratos) * 100)
        porcentaje_mxn = round((total_mxn / total_contratos) * 100)

    # Ajustes para visualización de la barra de progreso
    min_display_percentage = 5  # Mínimo porcentaje visual para que la barra sea visible
    porcentaje_usd_display = porcentaje_usd if porcentaje_usd == 0 else max(porcentaje_usd, min_display_percentage)
    porcentaje_cop_display = porcentaje_cop if porcentaje_cop == 0 else max(porcentaje_cop, min_display_percentage)
    porcentaje_mxn_display = porcentaje_mxn if porcentaje_mxn == 0 else max(porcentaje_mxn, min_display_percentage)

    total_display_percentage = porcentaje_usd_display + porcentaje_cop_display + porcentaje_mxn_display
    if total_display_percentage > 100:
        scale_factor = 100 / total_display_percentage
        porcentaje_usd_display = round(porcentaje_usd_display * scale_factor)
        porcentaje_cop_display = round(porcentaje_cop_display * scale_factor)
        porcentaje_mxn_display = round(porcentaje_mxn_display * scale_factor)

    total_monedas = total_usd + total_cop + total_mxn
    # Asegurar un mínimo de visibilidad visual
    min_display_percentage = 10 

    porcentaje_usd_display = porcentaje_usd if porcentaje_usd == 0 else max(porcentaje_usd, min_display_percentage)
    porcentaje_cop_display = porcentaje_cop if porcentaje_cop == 0 else max(porcentaje_cop, min_display_percentage)
    porcentaje_mxn_display = porcentaje_mxn if porcentaje_mxn == 0 else max(porcentaje_mxn, min_display_percentage)

    # Sumar y ajustar si superan el 100%
    total_display_percentage = porcentaje_usd_display + porcentaje_cop_display + porcentaje_mxn_display
    if total_display_percentage > 100:
        scale_factor = 100 / total_display_percentage
        porcentaje_usd_display = round(porcentaje_usd_display * scale_factor)
        porcentaje_cop_display = round(porcentaje_cop_display * scale_factor)
        porcentaje_mxn_display = round(porcentaje_mxn_display * scale_factor)
    context = {
        'form': form,
        'cliente_otros_si_info': cliente_otros_si_info,
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'mensaje': "No se encontraron resultados para los filtros aplicados." if busqueda_realizada and not show_data else "No se ha realizado ninguna búsqueda aún.",
        'total_contratos': total_contratos,
        'valor_total': valor_total,
        'contratos_firmados': contratos_firmados,
        'contratos_firmados_cliente': contratos_firmados_cliente,
        'contratos_con_polizas': contratos_con_polizas,
        'porcentaje_firmado_cliente': porcentaje_firmado_cliente,
        'porcentaje_firmado_interno': porcentaje_firmado_interno,
        'porcentaje_con_polizas': porcentaje_con_polizas,
        'total_usd': total_usd,
        'total_cop': total_cop,
        'total_mxn': total_mxn,
        'porcentaje_usd': porcentaje_usd,
        'porcentaje_cop': porcentaje_cop,
        'porcentaje_mxn': porcentaje_mxn,
        'porcentaje_usd_display': porcentaje_usd_display,
        'porcentaje_cop_display': porcentaje_cop_display,
        'porcentaje_mxn_display': porcentaje_mxn_display,
        'total_monedas': total_monedas,
        'porcentaje_usd_display': porcentaje_usd_display,
        'porcentaje_cop_display': porcentaje_cop_display,
        'porcentaje_mxn_display': porcentaje_mxn_display,
    }

    return render(request, 'informes/informes_otros_si_index.html', context)


@login_required
@verificar_permiso('can_manage_informe_otros_si')
# Exportar a Excel
def exportar_otros_si_excel(request):
    """
    Vista que genera un archivo Excel con los contratos Otros Sí filtrados.

    Parámetros:
        request (HttpRequest): Solicitud GET con los filtros.

    Retorna:
        HttpResponse: Respuesta HTTP con el archivo Excel generado para descarga.
    """
    cliente_otros_si_info = []

    if request.method == 'GET':
        form = OtrosSiFilterForm(request.GET)

        if form.is_valid():
            clientes = Clientes.objects.only('Nombre_Cliente', 'ClienteId').filter(Activo=True)
            otros_si = ContratosOtrosSi.objects.select_related('ClienteId', 'monedaId').all()

            clientes, otros_si = filtrar_otros_si(form, clientes, otros_si)
            cliente_otros_si_info = obtener_otros_si(clientes, otros_si)

            if not cliente_otros_si_info:
                return HttpResponse("No se encontraron resultados para los filtros aplicados.")
        else:
            return HttpResponse("No se encontraron resultados para los filtros aplicados.")

        data = []
        for cliente in cliente_otros_si_info:
            for registro in cliente['otros_si']:
                data.append([
                    cliente['documento'],
                    cliente['nombre'],
                    registro['FechaInicio'],
                    registro['FechaFin'],
                    registro['NumeroOtroSi'],
                    registro['Contrato'],
                    registro['ValorOtroSi'],
                    'SI' if registro['ValorIncluyeIva'] else 'NO',
                    'SI' if registro['Polizas'] else 'NO',
                    registro['PolizasDesc'],
                    'SI' if registro['FirmadoFlag'] else 'NO',
                    'SI' if registro['FirmadoCliente'] else 'NO',
                    registro['moneda'],
                ])

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Otros Sí"

        encabezados = [
            'Documento Cliente',
            'Nombre Cliente',
            'Fecha Inicio',
            'Fecha Fin',
            'Número OtroSí',
            'Contrato',
            'Valor OtroSí',
            'Incluye IVA',
            'Polizas',
            'Descripción Polizas',
            'Firmado (Interno)',
            'Firmado Cliente',
            'Moneda',
        ]

        for col_num, header in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_num, row in enumerate(data, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"otros_si_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response
