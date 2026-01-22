# Informe de Salarios

# Librerías estándar y de terceros
from datetime import date, datetime
from collections import defaultdict, Counter
from dateutil.relativedelta import relativedelta

# Django
from django.http import HttpResponse
from django.shortcuts import render

# Formularios y modelos del módulo
from Modulo.forms import EmpleadoFilterForm
from Modulo.models import Empleado, Nomina

# Librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from Modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required


# Función para filtrar empleados y nóminas (helper, sin decoradores)
def filtrar_empleados_y_nominas(form, empleados, nominas):
    """
    Aplica filtros del formulario sobre empleados y nóminas.
    - Filtra por nombre, línea, cargo y año.
    - Devuelve los empleados y nóminas que cumplen los filtros aplicados.
    """
    empleados = empleados.order_by('Documento')
    nominas = nominas.order_by('-Anio')

    nombre = form.cleaned_data.get('Nombre')
    linea = form.cleaned_data.get('LineaId')
    anio = form.cleaned_data.get('Anio')
    cargo = form.cleaned_data.get('Cargo')

    if nombre:
        empleados = empleados.filter(Nombre__icontains=nombre)
    if linea:
        empleados = empleados.filter(LineaId=linea)
    if cargo:
        empleados = empleados.filter(CargoId=cargo)
    if anio:
        nominas = nominas.filter(Anio=anio)

    documentos_nominas = nominas.values_list('Documento', flat=True)
    empleados = empleados.filter(Documento__in=documentos_nominas)

    return empleados, nominas


# Función para calcular la información del empleado (helper, sin decoradores)
def obtener_empleado_info(empleados, nominas, meses):
    """
    Une la información de empleados con sus registros de nómina.
    - Agrupa la información de salarios y clientes mes a mes.
    - Calcula los años en la empresa.
    Devuelve una lista de diccionarios con los datos completos por empleado.
    """

    def limpiar_documento(doc):
        # Elimina puntos, espacios y convierte a string sin espacios adicionales
        return str(doc).replace('.', '').replace(' ', '').strip()

    empleado_info = []

    # Mapear nóminas por (documento_limpio, año_str) -> {mes_str: nomina}
    nominas_por_documento_y_mes = defaultdict(dict)
    # Mapear documento_limpio -> set(años_str)
    anios_por_documento = defaultdict(set)

    for nomina in nominas:
        # nomina.Documento es FK a modelo Documento, su campo Documento es el valor real
        doc_limpio = limpiar_documento(nomina.Documento.Documento)
        anio_str = str(nomina.Anio)
        mes_str = str(nomina.Mes).zfill(2)

        nominas_por_documento_y_mes[(doc_limpio, anio_str)][mes_str] = nomina
        anios_por_documento[doc_limpio].add(anio_str)

    fecha_actual = date.today()

    for empleado in empleados:
        doc_limpio = limpiar_documento(empleado.Documento)
        anios_nominas = sorted(anios_por_documento.get(doc_limpio, []))

        # Años en Flag solo depende del empleado, no del año de nómina
        años_en_flag = relativedelta(fecha_actual, empleado.FechaIngreso).years

        for anio_str in anios_nominas:
            salarios_meses = []

            for mes in range(1, 13):
                mes_key = str(mes).zfill(2)
                mes_nomina = nominas_por_documento_y_mes.get(
                    (doc_limpio, anio_str), {}
                ).get(mes_key)

                if mes_nomina:
                    salario = mes_nomina.Salario
                    cliente = (
                        mes_nomina.Cliente.Nombre_Cliente
                        if mes_nomina.Cliente
                        else ""
                    )
                else:
                    salario = ""
                    cliente = ""

                salarios_meses.append({'salario': salario, 'cliente': cliente})

            empleado_info.append({
                'nombre_linea': empleado.LineaId.Linea,
                'documento_colaborador': empleado.Documento,
                'nombre_colaborador': empleado.Nombre,
                'cargo': empleado.CargoId.Cargo,
                'perfil': empleado.PerfilId.Perfil,
                'modulo': empleado.ModuloId.Modulo,
                'certificado': 'SI' if empleado.CertificadoSAP else 'NO',
                'fecha_ingreso': empleado.FechaIngreso,
                'años_en_flag': años_en_flag,
                'año': int(anio_str),
                'meses': salarios_meses,
            })

    return empleado_info


@login_required
@verificar_permiso('can_view_informe_salarios')
# Informe de Nómina de Empleados
def empleado_nomina_filtrado(request):
    """
    Vista para mostrar el informe de nómina de empleados.
    - Aplica filtros GET desde el formulario.
    - Genera contexto con información de salarios por mes.
    """
    meses = "Enero Febrero Marzo Abril Mayo Junio Julio Agosto Septiembre Octubre Noviembre Diciembre".split()
    empleado_info = []
    show_data = False
    busqueda_realizada = False

    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)

        if request.GET:
            busqueda_realizada = True

            if form.is_valid():
                # Inicializar empleados y nóminas con select_related para evitar N+1
                empleados = Empleado.objects.select_related(
                    'LineaId',
                    'CargoId',
                    'PerfilId',
                    'ModuloId',
                )
                nominas = Nomina.objects.select_related(
                    'Documento',
                    'Cliente',
                )

                # Aplicar filtros
                empleados, nominas = filtrar_empleados_y_nominas(form, empleados, nominas)

                # Obtener información de los empleados
                empleado_info = obtener_empleado_info(empleados, nominas, meses)
                show_data = bool(empleado_info)
    else:
        form = EmpleadoFilterForm()

    # Lógica para cards
    total_certificados = sum(1 for e in empleado_info if e['certificado'] == 'SI')

    lineas = [e['nombre_linea'] for e in empleado_info if e['nombre_linea']]
    conteo_lineas = dict(Counter(lineas))

    context = {
        "meses": meses,
        'form': form,
        'empleado_info': empleado_info,
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'total_certificados': total_certificados,
        'conteo_lineas': conteo_lineas,
        'lineas_labels': list(conteo_lineas.keys()),
        'lineas_data': list(conteo_lineas.values()),
        'mensaje': (
            "No se encontraron resultados para los filtros aplicados."
            if busqueda_realizada and not show_data
            else "No se ha realizado ninguna búsqueda aún."
        ),
    }

    return render(request, 'Informes/Informessalariosindex.html', context)


@login_required
@verificar_permiso('can_view_informe_salarios')
# Funcionalidad para descargar los resultados en Excel
def exportar_nomina_excel(request):
    """
    Exporta a Excel la información de nómina filtrada por el formulario.
    - Recupera los mismos filtros de la vista.
    - Genera un archivo Excel con la información mensual de salarios y clientes.
    """
    meses = "Enero Febrero Marzo Abril Mayo Junio Julio Agosto Septiembre Octubre Noviembre Diciembre".split()
    empleado_info = []

    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)

        if form.is_valid():
            # Inicializar empleados y nóminas con select_related para evitar N+1
            empleados = Empleado.objects.select_related(
                'LineaId',
                'CargoId',
                'PerfilId',
                'ModuloId',
            )
            nominas = Nomina.objects.select_related(
                'Documento',
                'Cliente',
            )

            # Aplicar filtros
            empleados, nominas = filtrar_empleados_y_nominas(form, empleados, nominas)

            # Obtener información de los empleados
            empleado_info = obtener_empleado_info(empleados, nominas, meses)

        if not empleado_info:
            return HttpResponse("No se encontraron datos para exportar.", status=404)

        if not form.is_valid():
            return HttpResponse(
                "Filtros no válidos. Por favor, revisa los criterios ingresados.",
                status=400,
            )

    # Preparar datos para Excel
    data = []
    for empleado in empleado_info:
        fila = {
            'Nombre Línea': empleado['nombre_linea'],
            'Documento Colaborador': empleado['documento_colaborador'],
            'Nombre Colaborador': empleado['nombre_colaborador'],
            'Cargo': empleado['cargo'],
            'Perfil': empleado['perfil'],
            'Módulo': empleado['modulo'],
            'Certificado SAP': empleado['certificado'],
            'Fecha Ingreso': empleado['fecha_ingreso'],
            'Años en Flag': empleado['años_en_flag'],
            'Año': empleado['año'],
        }

        # Agregar salarios y clientes por mes
        for i, mes_data in enumerate(empleado['meses'], 1):
            fila[f'Salario {meses[i - 1]}'] = mes_data['salario']
            fila[f'Cliente {meses[i - 1]}'] = mes_data['cliente']

        data.append(fila)

    # Crear libro y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe de Nómina"

    # Encabezados
    headers = list(data[0].keys())
    for col_idx, col in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for r_idx, row in enumerate(data, start=2):
        for c_idx, value in enumerate(row.values(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.border = thin_border

    # Crear respuesta de Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"nomina_Empleados_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)

    return response
