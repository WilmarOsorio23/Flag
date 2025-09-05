from django.shortcuts import render, HttpResponse
import csv
from collections import defaultdict
from django.db.models import Sum
from Modulo.forms import inforTiempoEmpleadosFilterForm
from Modulo.models import Consultores, Tiempos_Cliente, Linea, Clientes, Facturacion_Consultores
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from datetime import datetime
from Modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required

@login_required
@verificar_permiso('can_manage_informe_tiempos_consultores')

def filtrar_datos(request):
    documento = request.GET.getlist('Documento')
    anio = request.GET.getlist('Anio')
    linea = request.GET.getlist('LineaId')
    mes = request.GET.getlist('Mes')

    tiempos = Tiempos_Cliente.objects.all()

    #Filtrar por documentos que existan en la tabla consultores
    if documento:
        tiempos=tiempos.filter(Documento__in=documento)
    else:
        # Filtrar directamente por documentos válidos en la consulta principal
        tiempos = Tiempos_Cliente.objects.filter(Documento__in=Consultores.objects.values_list('Documento', flat=True))
    if anio:
        tiempos = tiempos.filter(Anio=anio)
    if mes:
        tiempos=tiempos.filter(Mes=mes)
    if linea:
        tiempos= tiempos.filter(lineaId__in=linea)
    


     # Anotar el total por mes, línea y cliente
    tiempos = tiempos.values('Mes', 'LineaId', 'Documento').annotate(total=Sum('Horas'))
    return tiempos

def Filtrar_datos(forms,tiempos):
    documento = forms.cleaned_data.get('documento')
    anio = forms.cleaned_data.get('Anio')
    mes = forms.cleaned_data.get('Mes')
    linea = forms.cleaned_data.get('LineaId')

    #aplicar filtros solo si se seleccionan

    if documento:
        tiempos = tiempos.filter(Documento__in= documento)
    if anio:
        tiempos=tiempos.filter(Anio=anio)
    if mes:
        tiempos=tiempos.filter(Mes__in=mes)
    if linea:
        tiempos= tiempos.filter(LineaId__in=linea)

    return tiempos
def tiempos_clientes_filtrado(request):
    tiempos_info = []
    show_data = False
    max_facturaciones = 0

    if request.method == 'GET':
        forms = inforTiempoEmpleadosFilterForm(request.GET)
        documentos_validos = set(Consultores.objects.values_list('Documento', flat=True))
        tiempos = Tiempos_Cliente.objects.select_related('ClienteId', 'LineaId').filter(Documento__in=documentos_validos)

        if forms.is_valid():
            tiempos = Filtrar_datos(forms, tiempos)
            consultores = {c.Documento: c for c in Consultores.objects.select_related('ModuloId').all()}
            
            for t in tiempos:
                if not Consultores.objects.filter(Documento=t.Documento).exists():
                    continue
                
                facturaciones = Facturacion_Consultores.objects.filter(
                    Anio=t.Anio,
                    Documento=t.Documento,
                    ClienteId=t.ClienteId,
                    LineaId=t.LineaId,
                    Periodo_Cobrado=t.Mes
                ).order_by('Cta_Cobro')
                
                # Debug: Imprimir información básica
                print(f"Procesando registro - Documento: {t.Documento}, Cliente: {t.ClienteId}, Mes: {t.Mes}")
                print(f"Facturaciones encontradas: {facturaciones.count()}")
                
                registro = {
                    'Documento': t.Documento,
                    'Nombre': consultores[t.Documento].Nombre if t.Documento in consultores else '',
                    'Empresa': consultores[t.Documento].Empresa if t.Documento in consultores else '',
                    'Anio': t.Anio,
                    'Mes': t.Mes,
                    'Linea': t.LineaId.Linea if t.LineaId else '',
                    'Cliente': t.ClienteId.Nombre_Cliente if t.ClienteId else '',
                    'Modulo': consultores[t.Documento].ModuloId.Modulo if t.Documento in consultores and consultores[t.Documento].ModuloId else '',
                    'Horas': float(t.Horas) if t.Horas else 0.0,
                    'Facturaciones': []
                }
                
                for facturacion in facturaciones:
                    print(f"Facturación encontrada: Cta_Cobro: {facturacion.Cta_Cobro}, Horas: {facturacion.Horas}")
                    registro['Facturaciones'].append({
                        'Cta_Cobro': facturacion.Cta_Cobro if facturacion.Cta_Cobro else '',
                        'Horas_Facturacion': float(facturacion.Horas) if facturacion.Horas else 0.0
                    })
                
                if facturaciones.count() > 0:
                    print(f"Primera facturación en registro: {registro['Facturaciones'][0]}")
                
                num_facturaciones = len(registro['Facturaciones'])
                if num_facturaciones > max_facturaciones:
                    max_facturaciones = num_facturaciones
                
                tiempos_info.append(registro)
            
            tiempos_info.sort(key=lambda x: x['Nombre'])
            show_data = bool(tiempos_info)
            
            # Debug: Imprimir resumen
            print(f"\nResumen de datos:")
            print(f"Total registros: {len(tiempos_info)}")
            print(f"Máximo facturaciones: {max_facturaciones}")
            print(f"Primer registro completo: {tiempos_info[0] if tiempos_info else 'No hay datos'}\n")
        
        request.session['tiempos_info'] = tiempos_info
        request.session['max_facturaciones'] = max_facturaciones

    context = {
        'form': forms,
        'tiempos_info': tiempos_info,
        'show_data': show_data,
        'max_facturaciones': max_facturaciones,
        'mensaje': "No se encontraron resultados para los filtros aplicados" if not tiempos_info else ""
    }

    return render(request, 'informes/informes_tiempos_consultores_index.html', context)
            
def exportar_tiempos_clientes_excel(request):
    # Obtener datos de la sesión
    tiempos_info = request.session.get('tiempos_info', [])
    max_facturaciones = request.session.get('max_facturaciones', 0)

    if not tiempos_info:
        return HttpResponse("No hay datos para exportar.")

    # Generar el Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Tiempos Consultores"

    # Definir encabezados base
    encabezados_base = [
        "Documento", "Nombre", "Empresa", "Línea", "Cliente", "Módulo", "Año", "Mes", "Horas"
    ]
    
    # Agregar columnas dinámicas de facturación (basado en max_facturaciones)
    for i in range(1, max_facturaciones + 1):
        encabezados_base.extend([f"Cta_Cobro_{i}", f"Horas_Facturacion_{i}"])

    # Estilo para bordes y encabezados
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar encabezados
    for col_num, header in enumerate(encabezados_base, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Insertar datos
    for row_num, registro in enumerate(tiempos_info, 2):
        # Datos base
        base_data = [
            registro.get('Documento', ''),
            registro.get('Nombre', ''),
            registro.get('Empresa', ''),
            registro.get('Linea', ''),
            registro.get('Cliente', ''),  # Nota: Hay un typo aquí ('Cliente' vs 'Cliente')
            registro.get('Modulo', ''),
            registro.get('Anio', ''),
            registro.get('Mes', ''),
            registro.get('Horas', 0)
        ]
        
        # Datos de facturación
        facturaciones_data = []
        for fact in registro.get('Facturaciones', []):
            facturaciones_data.extend([
                fact.get('Cta_Cobro', ''),
                fact.get('Horas_Facturacion', 0)
            ])
        
        # Rellenar con valores vacíos si hay menos facturaciones
        while len(facturaciones_data) < max_facturaciones * 2:
            facturaciones_data.extend(['', ''])
        
        # Combinar todos los datos
        row_data = base_data + facturaciones_data[:max_facturaciones * 2]

        # Escribir en Excel
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            
            # Formato numérico para horas
            if isinstance(value, (int, float)) and encabezados_base[col_num-1].lower().startswith('horas'):
                cell.number_format = '0.00'
            elif isinstance(value, str):
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

    # Ajustar automáticamente el ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Configurar respuesta
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"tiempos_consultores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response