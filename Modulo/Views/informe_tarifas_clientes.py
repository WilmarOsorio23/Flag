from django.shortcuts import render
from collections import defaultdict
from django.http import HttpResponse
from Modulo.forms import TarifaClienteFilterForm
from Modulo.models import Clientes, Tarifa_Clientes
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from Modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required

#Función para filtrar Tarifas de clientes
def filtrar_tarifas_clientes(form, clientes, clientes_tarifas):
    clientes = clientes.order_by('Nombre_Cliente')
    nombre = form.cleaned_data.get('Nombre_Cliente')
    linea = form.cleaned_data.get('LineaId')
    modulo = form.cleaned_data.get('ModuloId')
    Activo = form.cleaned_data.get('Activo')
    anios = form.cleaned_data.get('anio')

    if nombre:
        clientes = clientes.filter(Nombre_Cliente=nombre)
    if linea:
        clientes_tarifas = clientes_tarifas.filter(lineaId=linea)
    if modulo:
        clientes_tarifas = clientes_tarifas.filter(moduloId=modulo)
    if Activo:
        clientes = clientes.filter(Activo=bool(int(Activo)))
    if anios:
        clientes_tarifas = clientes_tarifas.filter(anio__in=anios)
    
    clientesId_tarifas = clientes_tarifas.values_list('clienteId', flat=True)
    clientes = clientes.filter(ClienteId__in=clientesId_tarifas)

    return clientes, clientes_tarifas

#Función para obtener información de las tarifas de clientes
def obtener_tarifas_clientes(clientes, tarifas, anios):
    cliente_tarifas_info = []
    tarifas_por_cliente = defaultdict(list)

    # Organizar las tarifas por clienteId para optimizar consultas
    for tarifa in tarifas:
        tarifas_por_cliente[tarifa.clienteId_id].append(tarifa)

    for cliente in clientes:
        # Obtener tarifas del cliente
        tarifas_cliente = tarifas_por_cliente.get(cliente.ClienteId, [])

        # Crear el diccionario con los datos del cliente y sus tarifas
        datos_cliente = {
           
            'documento': cliente.DocumentoId,
            'nombre_cliente': cliente.Nombre_Cliente,
            'activo': 'Activo' if cliente.Activo  else 'Inactivo',
            'tarifas': {anio: [] for anio in anios}
        }

        for tarifa in tarifas_cliente:
            if tarifa.anio in anios:
                datos_cliente['tarifas'][tarifa.anio].append({
                    'linea': tarifa.lineaId.Linea,
                    'modulo': tarifa.moduloId.Modulo,
                    'mes': tarifa.mes,
                    'tarifa_hora': tarifa.valorHora,
                    'tarifa_dia': tarifa.valorDia,
                    'tarifa_mes': tarifa.valorMes,
                    'tarifa_bolsa': tarifa.bolsaMes,
                    'referencia': tarifa.referenciaId.descripcionReferencia,
                    'centro_costos': tarifa.centrocostosId.descripcionCeCo,
                    'iva': tarifa.iva,
                    'sitio_trabajo': tarifa.sitioTrabajo,
                    'moneda': tarifa.monedaId.Nombre,

                })

        cliente_tarifas_info.append(datos_cliente)

    return cliente_tarifas_info

@login_required
@verificar_permiso('can_view_informe_tarifas_clientes')
# Función para generar el informe de tarifas de clientes
def tarifas_clientes_filtrado(request):
    cliente_tarifas_info = []
    show_data = False  

    if request.method == 'GET':
        form = TarifaClienteFilterForm(request.GET)

        if form.is_valid():
            # Filtrar clientes y tarifas
            clientes = Clientes.objects.only('Nombre_Cliente', 'ClienteId').filter(Activo=True)
            anios = form.cleaned_data.get('anio')
            tarifas = Tarifa_Clientes.objects.select_related('lineaId', 'moduloId', 'referenciaId', 'centrocostosId', 'monedaId').filter(anio__in=anios)
            clientes, tarifas = filtrar_tarifas_clientes(form, clientes, tarifas)

            # Obtener los años disponibles
            anios_disponibles = Tarifa_Clientes.objects.values_list('anio', flat=True).distinct()

            # Obtener información de las tarifas de los clientes
            cliente_tarifas_info = obtener_tarifas_clientes(clientes, tarifas, anios)

            show_data = bool(cliente_tarifas_info)  # Mostrar datos si hay resultados       
    else:
        form = TarifaClienteFilterForm()
    
    context = {
        'form': form,
        'cliente_tarifas_info':  cliente_tarifas_info,
        'show_data': show_data,
        'mensaje': "No se encontraron resultados para los filtros aplicados." if not cliente_tarifas_info else "",
        'anios_disponibles': anios_disponibles,
    }

    return render(request, 'informes/informes_tarifas_clientes_index.html', context)

@login_required
@verificar_permiso('can_view_informe_tarifas_clientes')
#Función para exportar el informe de tarifas de clientes a Excel
def exportar_tarifas_clientes_excel(request):
    cliente_tarifas_info = []

    if request.method == 'GET':
        form = TarifaClienteFilterForm(request.GET)

        if form.is_valid():
            # Inicializar clientes y tarifas
            clientes = Clientes.objects.all()
            tarifas = Tarifa_Clientes.objects.all()

            # Obtener información de los clientes y tarifas filtrados
            clientes, tarifas = filtrar_tarifas_clientes(form, clientes, tarifas)

            # Obtener los años seleccionados por el usuario
            anios = form.cleaned_data.get('anio')

            # Obtener información de las tarifas de los clientes
            cliente_tarifas_info = obtener_tarifas_clientes(clientes, tarifas, anios)

            if not cliente_tarifas_info:
                return HttpResponse("No se encontraron resultados para los filtros aplicados.")
        else:
            return HttpResponse("No se encontraron resultados para los filtros aplicados.")
    
        # Preparar los datos para Excel
        data = []
        for cliente in cliente_tarifas_info:
            for anio, tarifas in cliente['tarifas'].items():
                for tarifa in tarifas:
                    data.append([
                        cliente['documento'],
                        cliente['nombre_cliente'],
                        cliente['activo'],
                        anio,
                        tarifa['linea'],
                        tarifa['modulo'],
                        tarifa['mes'],
                        tarifa['tarifa_hora'],
                        tarifa['tarifa_dia'],
                        tarifa['tarifa_mes'],
                        tarifa['tarifa_bolsa'],
                        tarifa['referencia'],
                        tarifa['centro_costos'],
                        tarifa['iva'],
                        tarifa['sitio_trabajo'],
                        tarifa['moneda'],
                    ])
        if data:
            # Crear el libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe tarifas de clientes"

            # Definir los encabezados
            encabezados = [
                'Documento',
                'Nombre Cliente',
                'Activo',
                'Año',
                'Linea',
                'Modulo',
                'Mes',
                'Tarifa Hora',
                'Tarifa Dia',
                'Tarifa Mes',
                'Tarifa Bolsa',
                'Referencia',
                'Centro Costos',
                'IVA',
                'Sitio Trabajo',
                'Moneda',
            ]
            for col_num, header in enumerate(encabezados, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Agregar datos a la hoja
            for row_num, row in enumerate(data, 2):
                    for col_num, value in enumerate(row, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.alignment = Alignment(horizontal='center')  # Centrar datos

                # Ajustar el ancho de las columnas
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border

                # Crear respuesta de Excel
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = f"tarifas_clientes_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'

                # Guardar el libro de trabajo en la respuesta  
                wb.save(response)

                return response
        