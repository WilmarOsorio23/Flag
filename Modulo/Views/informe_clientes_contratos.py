from django.shortcuts import render
from collections import defaultdict
from django.http import HttpResponse
from Modulo.forms import ClientesContratosFilterForm
from Modulo.models import Clientes, ClientesContratos
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from Modulo.models import ContratosOtrosSi
from django.db import models

#Función para filtrar Contratos de clientes
def filtrar_clientes_contratos(form, clientes, clientes_contratos):
    clientes = clientes.order_by('Nombre_Cliente')
    nombre = form.cleaned_data.get('Nombre_Cliente')
    fecha_inicio = form.cleaned_data.get('FechaInicio')
    contrato_vigente = form.cleaned_data.get('ContratoVigente')
    
    if nombre:
        clientes = clientes.filter(Nombre_Cliente=nombre)
    if fecha_inicio:
        clientes_contratos = clientes_contratos.filter(FechaInicio=fecha_inicio)
    if contrato_vigente:
        clientes_contratos = clientes_contratos.filter(ContratoVigente=(contrato_vigente == 'True'))
    
    
    ClienteId_contratos = clientes_contratos.values_list('ClienteId', flat=True)
    clientes = clientes.filter(ClienteId__in=ClienteId_contratos)

    return clientes, clientes_contratos

#Función para obtener información de los contratos de clientes
def obtener_clientes_contratos(clientes, contratos):
    cliente_contratos_info = []
    contratos_por_cliente = defaultdict(list)


    # Organizar los contratos por ClienteId para optimizar consultas
    for contrato in contratos:
        contratos_por_cliente[contrato.ClienteId_id].append(contrato)

    # Obtener conteo de otros sí por contrato
    otros_si_counts = ContratosOtrosSi.objects.values('Contrato').annotate(total=models.Count('ContratosOtrosSiId'))
    contrato_to_otros_si_count = {item['Contrato']: item['total'] for item in otros_si_counts}

    for cliente in clientes:
        # Obtener contratos del cliente
        contratos_cliente = contratos_por_cliente.get(cliente.ClienteId, [])

        # Crear el diccionario con los datos del cliente y sus contratos
        datos_cliente = {
            'documento': cliente.DocumentoId,
            'nombre': cliente.Nombre_Cliente,
            'contratos': [{
                'ClienteId': contrato.ClienteId,
                'FechaInicio': contrato.FechaInicio,
                'FechaFin': contrato.FechaFin,
                'Contrato': contrato.Contrato,
                'ContratoVigente': contrato.ContratoVigente,
                'OC_Facturar': contrato.OC_Facturar,
                'Parafiscales': contrato.Parafiscales,
                'HorarioServicio': contrato.HorarioServicio,
                'FechaFacturacion': contrato.FechaFacturacion,
                'TipoFacturacion': contrato.TipoFacturacion,
                'Observaciones': contrato.Observaciones,
                'Polizas': contrato.Polizas,
                'PolizasDesc': contrato.PolizasDesc,
                'ContratoValor': contrato.ContratoValor,
                'IncluyeIvaValor': contrato.IncluyeIvaValor,
                'ContratoDesc': contrato.ContratoDesc,
                'ServicioRemoto': contrato.ServicioRemoto,
                'documento': cliente.DocumentoId,
                'Nombre_Cliente': cliente.Nombre_Cliente,
                'monedaId': contrato.monedaId,
                'otros_si_count': contrato_to_otros_si_count.get(contrato.Contrato, 0)
                }
                for contrato in contratos_cliente]  # Corrected variable name here
        }

        cliente_contratos_info.append(datos_cliente)

    return cliente_contratos_info

# Función para generar el informe de contratos de clientes
def clientes_contratos_filtrado(request):
    cliente_contratos_info = []
    show_data = False  

    if request.method == 'GET':
        form = ClientesContratosFilterForm(request.GET)

        if form.is_valid():
            # Filtrar clientes y contratos
            clientes = Clientes.objects.only('Nombre_Cliente', 'ClienteId').filter(Activo=True)
            contratos = ClientesContratos.objects.select_related('ClienteId', 'monedaId').all()
            clientes, contratos = filtrar_clientes_contratos(form, clientes, contratos)
            
            print("Clientes filtrados:", clientes)
            print("Contratos filtrados:", contratos)

            # Obtener información de los contratos de los clientes
            cliente_contratos_info = obtener_clientes_contratos(clientes, contratos)

            print("Información de contratos de clientes:", cliente_contratos_info)

            show_data = bool(cliente_contratos_info)  # Mostrar datos si hay resultados       
    else:
        form = ClientesContratosFilterForm()
    
    context = {
        'form': form,
        'cliente_contratos_info': cliente_contratos_info,
        'show_data': show_data,
        'mensaje': "No se encontraron resultados para los filtros aplicados." if not cliente_contratos_info else "",
    }

    return render(request, 'informes/informes_clientes_contratos_index.html', context)

# Función para exportar el informe de contratos de clientes a Excel
def exportar_clientes_contratos_excel(request):
    cliente_contratos_info = []

    if request.method == 'GET':
        form = ClientesContratosFilterForm(request.GET)

        if form.is_valid():
            # Filtrar clientes y contratos
            clientes = Clientes.objects.only('Nombre_Cliente', 'ClienteId').filter(Activo=True)
            contratos = ClientesContratos.objects.select_related('ClienteId').all()
            clientes, contratos = filtrar_clientes_contratos(form, clientes, contratos)

            # Obtener información de los contratos de los clientes
            cliente_contratos_info = obtener_clientes_contratos(clientes, contratos)

            if not cliente_contratos_info:
                return HttpResponse("No se encontraron resultados para los filtros aplicados.")
            
        else:
                print("Errores del formulario:", form.errors)
                return HttpResponse("No se encontraron resultados para los filtros aplicados.")

        data = []    
        for cliente in cliente_contratos_info:
            for contrato in cliente['contratos']:
                data.append([
                    cliente['nombre'],                                      # 1. Nombre Cliente
                    contrato['Contrato'],                                   # 2. Contrato
                    contrato['ContratoDesc'],                               # 3. Descripción Contrato
                    contrato['FechaInicio'],                                # 4. Fecha Inicio
                    contrato['FechaFin'],                                   # 5. Fecha Fin
                    'SI' if contrato['ContratoVigente'] else 'NO',          # 6. Contrato Vigente
                    contrato['ContratoValor'],                              # 7. Valor Contrato
                    contrato['monedaId'].Nombre if contrato['monedaId'] else "Sin Moneda",  # 8. Moneda
                    'SI' if contrato['IncluyeIvaValor'] else 'NO',          # 9. Incluye IVA
                    contrato['otros_si_count'],                             # 10. # Otros Sí
                    'SI' if contrato['Polizas'] else 'NO',                  # 11. Pólizas
                    contrato['PolizasDesc'],                                # 12. Descripción Pólizas
                    'SI' if contrato['OC_Facturar'] else 'NO',              # 13. OC Facturar
                    'SI' if contrato['Parafiscales'] else 'NO',             # 14. Parafiscales
                    contrato['HorarioServicio'],                            # 15. Horario Servicio
                    contrato['FechaFacturacion'],                           # 16. Fecha Facturación
                    contrato['TipoFacturacion'],                            # 17. Tipo Facturación
                    'SI' if contrato['ServicioRemoto'] else 'NO',           # 18. Servicio Remoto
                ])

        if data:
            # Crear el libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Contratos Clientes"

            # Definir los encabezados
            encabezados = [
                'Nombre Cliente',            # 1
                'Contrato',                  # 2
                'Descripción Contrato',      # 3
                'Fecha Inicio',              # 4
                'Fecha Fin',                 # 5
                'Contrato Vigente',          # 6
                'Valor Contrato',            # 7
                'Moneda',                    # 8
                'Incluye IVA Valor',         # 9
                '# Otros Sí',                # 10
                'Polizas',                   # 11
                'Polizas Desc',              # 12
                'OC Facturar',               # 13
                'Parafiscales',              # 14
                'Horario Servicio',          # 15
                'Fecha Facturación',         # 16
                'Tipo Facturación',          # 17
                'Servicio Remoto',           # 18
            ]
            for col_num, header in enumerate(encabezados, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Agregar datos a la hoja
            for row_num, row in enumerate(data, 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal='center')  # Centrar datos

            # Ajustar el ancho de las columnas
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # Crear respuesta de Excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f"contratos_clientes_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            # Guardar el libro de trabajo en la respuesta
            wb.save(response)

        return response