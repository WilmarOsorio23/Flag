import random
from django import template
from django.shortcuts import render, HttpResponse
from collections import defaultdict
from django.db.models import Sum
from Modulo.forms import InformeFacturacionForm
from Modulo.models import FacturacionClientes, Linea, Clientes
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from datetime import datetime

# Función para filtrar datos
def filtrar_datos(form):
    anio = form.cleaned_data.get('Anio')
    lineas_seleccionadas = form.cleaned_data.get('LineaId')
    clientes_seleccionados = form.cleaned_data.get('ClienteId')

    facturas = FacturacionClientes.objects.all()

    if anio:
        facturas = facturas.filter(Anio=anio)
    
    if lineas_seleccionadas:
        facturas = facturas.filter(LineaId__in=lineas_seleccionadas)
        lineas = lineas_seleccionadas
    else:
        lineas = Linea.objects.all()
    
    if clientes_seleccionados:
        facturas = facturas.filter(ClienteId__in=clientes_seleccionados)
        clientes = clientes_seleccionados
    else:
        clientes = Clientes.objects.all()

    facturas = facturas.values('Mes', 'LineaId', 'ClienteId').annotate(total=Sum('Valor'))
    
    return facturas, lineas, clientes

# Función para organizar los datos
def organizar_datos(facturas, lineas, clientes):
    meses = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    totales = {
        'por_linea': defaultdict(lambda: defaultdict(float)),
        'general': defaultdict(float),
        'clientes': defaultdict(lambda: defaultdict(lambda: defaultdict(float))),
        'footer': {
            'por_linea': defaultdict(float),
            'general': 0,
            'clientes': defaultdict(lambda: defaultdict(float))
        }
    }

    for f in facturas:
        mes = f['Mes']
        linea_id = f['LineaId']
        cliente_id = f['ClienteId']
        total = float(f['total'] or 0)  # Convertir a float

        # Acumular por mes y línea
        totales['por_linea'][mes][linea_id] += total
        totales['footer']['por_linea'][linea_id] += total

        # Acumular por mes (general)
        totales['general'][mes] += total
        totales['footer']['general'] += total

        # Acumular por cliente y línea
        totales['clientes'][cliente_id][mes][linea_id] += total
        totales['footer']['clientes'][cliente_id][linea_id] += total

    rows = []
    for mes_num, mes_nombre in meses:
        row = {
            'mes': mes_nombre,
            'totales_linea': {},
            'total_general': float(totales['general'].get(mes_num, 0)),  # Convertir a float
            'clientes': {}
        }

        # Totales por línea para este mes
        for linea in lineas:
            linea_id = linea.LineaId
            row['totales_linea'][linea_id] = float(totales['por_linea'][mes_num].get(linea_id, 0))  # Convertir a float

        # Totales por cliente para este mes
        for cliente in clientes:
            cliente_id = cliente.ClienteId
            row['clientes'][cliente_id] = {
                'totales': {}
            }
            for linea in lineas:
                linea_id = linea.LineaId
                row['clientes'][cliente_id]['totales'][linea_id] = float(totales['clientes'][cliente_id][mes_num].get(linea_id, 0))  # Convertir a float

        rows.append(row)

    footer = {
        'totales_linea': {},
        'total_general': float(totales['footer']['general']),  # Convertir a float
        'clientes': {}
    }

    # Footer: totales por línea
    for linea in lineas:
        linea_id = linea.LineaId
        footer['totales_linea'][linea_id] = float(totales['footer']['por_linea'].get(linea_id, 0))  # Convertir a float

    # Footer: totales por cliente
    for cliente in clientes:
        cliente_id = cliente.ClienteId
        footer['clientes'][cliente_id] = {
            'totales': {}
        }
        for linea in lineas:
            linea_id = linea.LineaId
            footer['clientes'][cliente_id]['totales'][linea_id] = float(totales['footer']['clientes'][cliente_id].get(linea_id, 0))  # Convertir a float

    lineas_activas = [linea for linea in lineas if footer['totales_linea'].get(linea.LineaId, 0) > 0]

    clientes_activos = []
    for cliente in clientes:
        for linea in lineas_activas:
            total = footer['clientes'][cliente.ClienteId]['totales'].get(linea.LineaId, 0)
            if total > 0:
                clientes_activos.append((cliente, linea))

    return rows, footer, lineas, clientes, meses, lineas_activas, clientes_activos


# Función para generar el reporte en el frontend
def informes_facturacion_index(request):
    form = InformeFacturacionForm(request.GET or None)
    mensaje = None
    rows = None
    footer = None
    lineas = None
    clientes = None
    meses = None

    # Verificar si el formulario se ha enviado
    if 'buscar' in request.GET:
        if form.is_valid():
            facturas, lineas, clientes = filtrar_datos(form)
            if facturas.exists():
                rows, footer, lineas, clientes, meses, lineas_activas, clientes_activos = organizar_datos(facturas, lineas, clientes)
            else:
                mensaje = "No se encontraron datos con los filtros seleccionados."
        else:
            mensaje = "Por favor, corrige los errores en el formulario."
    else:
        # Primera carga: traer todos los datos y procesarlos como en filtrar_datos
        facturas = FacturacionClientes.objects.values('Mes', 'LineaId', 'ClienteId').annotate(total=Sum('Valor'))
        lineas = Linea.objects.all()
        clientes = Clientes.objects.all()
        if facturas.exists():
            rows, footer, lineas, clientes, meses, lineas_activas, clientes_activos = organizar_datos(facturas, lineas, clientes)
        else:
            mensaje = "No se encontraron datos."

    context = {
        'form': form,
        'rows': rows,
        'footer': footer,
        'lineas': lineas,
        'clientes': clientes,
        'meses': meses,
        'mensaje': mensaje,
        'lineas_activas': lineas_activas,
        'clientes_activos': clientes_activos
    }

    return render(request, 'Informes/informes_facturacion_index.html', context)


def generar_graficos_por_linea(rows, lineas_activas, clientes_activos, footer):
    """Genera datos para gráficos por línea y cliente, incluyendo Total General y Total Línea"""
    graficos = []

    colores = [
        'rgba(75, 192, 192, 0.5)',  # Azul claro
        'rgba(255, 99, 132, 0.5)',  # Rojo claro
        'rgba(255, 206, 86, 0.5)',  # Amarillo claro
        'rgba(54, 162, 235, 0.5)',  # Azul oscuro
        'rgba(153, 102, 255, 0.5)', # Morado claro
        'rgba(201, 203, 207, 0.5)'  # Gris claro
    ]

    # Gráfico para Total General
    labels = [row['mes'] for row in rows]
    total_general = [row['total_general'] for row in rows]

    graficos.append({
        'nombre': 'Total General',
        'labels': labels,
        'datasets': [
            {
                'label': 'Total General',
                'data': total_general,
                'type': 'bar',
                'backgroundColor': colores[0],
                'borderColor': colores[0].replace('0.5', '1'),
                'borderWidth': 1
            }
        ]
    })

    # Gráfico para Total Línea (primeras 4 columnas: Desarrollo, Consultoría, Seguridad, Soporte)
    columnas_interes = ['Desarrollo', 'Consultoria', 'Seguridad', 'Soporte']
    labels_total_linea = columnas_interes  # Eje X: nombres de las columnas
    data_total_linea = []

    for columna in columnas_interes:
        # Verificar que la columna exista en las líneas activas
        total_columna = sum(
            footer['totales_linea'].get(linea.LineaId, 0)
            for linea in lineas_activas if linea.Linea.strip().lower() == columna.strip().lower()
        )
        # Calcular el porcentaje de participación
        porcentaje = (total_columna / footer['total_general'] * 100) if footer['total_general'] > 0 else 0
        data_total_linea.append(porcentaje)

    graficos.append({
        'nombre': 'Total Línea',
        'labels': labels_total_linea,
        'datasets': [
            {
                'label': 'Total Línea',
                'data': data_total_linea,
                'type': 'bar',
                'backgroundColor': colores[1],
                'borderColor': colores[1].replace('0.5', '1'),
                'borderWidth': 1
            }
        ]
    })

    # Gráficos por línea
    for index, linea in enumerate(lineas_activas):
        total_linea = footer['totales_linea'].get(linea.LineaId, 0)
        if total_linea <= 0:
            continue

        clientes_data = []
        participacion_data = []

        for cliente, linea_cliente in clientes_activos:
            if linea_cliente.LineaId != linea.LineaId:
                continue

            total_cliente = footer['clientes'][cliente.ClienteId]['totales'].get(linea.LineaId, 0)
            participacion = (total_cliente / total_linea) * 100 if total_linea > 0 else 0

            clientes_data.append(cliente.Nombre_Cliente)
            participacion_data.append(participacion)

        if not clientes_data:
            continue

        color_index = (index + 2) % len(colores)  # Ciclar colores si hay más líneas que colores
        graficos.append({
            'nombre': f'% Participación en {linea.Linea}',
            'labels': clientes_data,
            'datasets': [
                {
                    'label': f'% Participación en {linea.Linea}',
                    'data': participacion_data,
                    'type': 'bar',
                    'backgroundColor': colores[color_index],
                    'borderColor': colores[color_index].replace('0.5', '1'),
                    'borderWidth': 1
                }
            ]
        })

    return graficos

def informes_facturacion_index(request):
    form = InformeFacturacionForm(request.GET or None)
    mensaje = None
    rows = None
    footer = None
    lineas = None
    clientes = None
    meses = None
    graficos_por_linea = []

    if 'buscar' in request.GET:
        if form.is_valid():
            facturas, lineas, clientes = filtrar_datos(form)
            if facturas.exists():
                rows, footer, lineas, clientes, meses, lineas_activas, clientes_activos = organizar_datos(facturas, lineas, clientes)
                # Pasar footer como nuevo parámetro
                graficos_por_linea = generar_graficos_por_linea(rows, lineas_activas, clientes_activos, footer)
            else:
                mensaje = "No se encontraron datos con los filtros seleccionados."
        else:
            mensaje = "Por favor, corrige los errores en el formulario."
    else:
        facturas = FacturacionClientes.objects.values('Mes', 'LineaId', 'ClienteId').annotate(total=Sum('Valor'))
        lineas = Linea.objects.all()
        clientes = Clientes.objects.all()
        if facturas.exists():
            rows, footer, lineas, clientes, meses, lineas_activas, clientes_activos = organizar_datos(facturas, lineas, clientes)
            # Pasar footer como nuevo parámetro
            graficos_por_linea = generar_graficos_por_linea(rows, lineas_activas, clientes_activos, footer)
        else:
            mensaje = "No se encontraron datos."

    context = {
        'form': form,
        'rows': rows,
        'footer': footer,
        'lineas': lineas,
        'clientes': clientes,
        'meses': meses,
        'mensaje': mensaje,
        'lineas_activas': lineas_activas,
        'clientes_activos': clientes_activos,
        'graficos_por_linea': graficos_por_linea
    }

    return render(request, 'Informes/informes_facturacion_index.html', context)


# Función para descargar el reporte en Excel
def descargar_reporte_excel(request):
    # Obtener parámetros y convertir a enteros
    anio = request.GET.get('Anio')
    lineas_ids = [int(id) for id in request.GET.getlist('LineaId') if id.isdigit()]
    clientes_ids = [int(id) for id in request.GET.getlist('ClienteId') if id.isdigit()]

    # Filtrar datos
    facturas = FacturacionClientes.objects.all()
    
    # Aplicar filtro de año
    if anio and anio.isdigit():
        facturas = facturas.filter(Anio=int(anio))
    
    # Aplicar filtro de líneas
    if lineas_ids:
        lineas = Linea.objects.filter(LineaId__in=lineas_ids)
        facturas = facturas.filter(LineaId__in=lineas_ids)
    else:
        lineas = Linea.objects.all()
    
    # Aplicar filtro de clientes
    if clientes_ids:
        clientes = Clientes.objects.filter(ClienteId__in=clientes_ids)
        facturas = facturas.filter(ClienteId__in=clientes_ids)
    else:
        clientes = Clientes.objects.all()

    # Procesar datos
    facturas = facturas.values('Mes', 'LineaId', 'ClienteId').annotate(total=Sum('Valor'))
    
    if facturas.exists():
        # Capturar todos los valores retornados
        rows, footer, lineas_queryset, clientes_queryset, meses, lineas_activas, clientes_activos = organizar_datos(facturas, lineas, clientes)
        
        # Usar los valores ya obtenidos de organizar_datos
        active_client_lines = clientes_activos

        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe de Facturación"

        # =========================================================================
        # ESTILOS PERSONALIZADOS
        # =========================================================================
        header_font = Font(bold=True, color="FFFFFF")
        blue_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        light_blue_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
        footer_fill = PatternFill(start_color="336699", end_color="336699", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # =========================================================================
        # CONSTRUIR CABECERAS
        # =========================================================================
        # Primera fila de cabecera
        first_row = [""]  # Celda A1 vacía para "Meses"
        first_row.append("Totales por Línea")
        first_row.extend([""] * (len(lineas_activas) - 1))  # Celdas para merge
        first_row.append("")

        # Agrupar clientes únicos manteniendo el orden
        seen_clients = set()
        unique_clients = []
        client_lineas_count = defaultdict(int)
        
        for cliente, linea in active_client_lines:
            client_lineas_count[cliente] += 1
            if cliente not in seen_clients:
                seen_clients.add(cliente)
                unique_clients.append(cliente)

        # Añadir nombres de clientes con espacios para merge
        for cliente in unique_clients:
            count = client_lineas_count[cliente]
            first_row.append(cliente.Nombre_Cliente)
            if count > 1:
                first_row.extend([""] * (count - 1))

        ws.append(first_row)

        # Segunda fila de cabecera
        second_row = ["Meses"]
        second_row.extend([linea.Linea for linea in lineas_activas])
        second_row.append("Total General")
        second_row.extend([linea.Linea for _, linea in active_client_lines])
        ws.append(second_row)

        # Combinar celdas
        # Merge "Totales por Línea" (B1:G1 para 6 líneas)
        start_col = 2
        end_col = start_col + len(lineas_activas) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        # Merge para nombres de clientes
        current_col = end_col + 2  # Columna después de Total General
        for cliente in unique_clients:
            num_lineas = client_lineas_count[cliente]
            if num_lineas > 1:
                end_merge = current_col + num_lineas - 1
                ws.merge_cells(
                    start_row=1,
                    start_column=current_col,
                    end_row=1,
                    end_column=end_merge
                )
            current_col += num_lineas

        # Agrupar líneas por cliente para combinar celdas
        client_groups = defaultdict(list)
        for cliente, linea in active_client_lines:
            client_groups[cliente].append(linea)

        # Combinar celdas para clientes
        start_col = len(lineas_activas) + 3  # Columna inicial después de Total General
        current_col = start_col
        for cliente, lineas in client_groups.items():
            num_columnas = len(lineas)
            if num_columnas > 1:
                end_col = current_col + num_columnas - 1
                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=end_col)
            current_col += num_columnas

        # Aplicar estilos a cabeceras
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font = header_font
                cell.fill = blue_fill if cell.row == 1 else light_blue_fill
                cell.border = border
                cell.alignment = alignment

        # =========================================================================
        # LLENAR DATOS
        # =========================================================================
        for row in rows:
            fila = [
                row['mes'],
                # Usar LineaId como clave del diccionario
                *[row['totales_linea'][linea.LineaId] for linea in lineas_activas],
                row['total_general'],
                # Acceder a los datos usando ClienteId y LineaId
                *[row['clientes'][cliente.ClienteId]['totales'][linea.LineaId] 
                for cliente, linea in clientes_activos]
            ]
            ws.append(fila)

        # =========================================================================
        # FILA DE TOTALES
        # =========================================================================
        fila_footer = [
            "Total",
            # Usar LineaId para acceder a los totales
            *[footer['totales_linea'][linea.LineaId] for linea in lineas_activas],
            footer['total_general'],
            # Acceder a los totales de clientes con ClienteId y LineaId
            *[footer['clientes'][cliente.ClienteId]['totales'][linea.LineaId] 
            for cliente, linea in clientes_activos]
        ]
        ws.append(fila_footer)

        for cell in ws[ws.max_row]:
            cell.fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")  # Color amarillo
            cell.font = Font(bold=True, color="000000")  # Texto en negro
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # =========================================================================
        # FILA DE % PARTICIPACIÓN
        # =========================================================================
        fila_participacion = [
            "% Participación",
            # Calcular % participación para líneas activas
            *[
                (footer['totales_linea'][linea.LineaId] / footer['total_general']) if footer['total_general'] > 0 else 0
                for linea in lineas_activas
            ],
            1.00,  # Total General siempre será 100% (como porcentaje)
            # Calcular % participación para clientes activos
            *[
                (footer['clientes'][cliente.ClienteId]['totales'][linea.LineaId] / footer['totales_linea'][linea.LineaId]) if footer['totales_linea'][linea.LineaId] > 0 else 0
                for cliente, linea in clientes_activos
            ]
        ]
        ws.append(fila_participacion)

        # Aplicar estilo a la fila de % Participación
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")  # Color azul claro
            cell.font = Font(bold=True, color="000000")  # Texto en negro
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):  # Aplicar formato de porcentaje
                cell.number_format = '0.00%'

        # =========================================================================
        # FORMATO NUMÉRICO Y AJUSTES
        # =========================================================================
        # Formato de números con separadores de miles
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row-1):
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,#0.00'

        # Ajustar ancho de columnas
        for col_idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        # Respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"informe_facturacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    else:
        return HttpResponse("No se encontraron datos para exportar.")