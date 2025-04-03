from django import forms
from django.shortcuts import render
from datetime import date, datetime
from collections import defaultdict
from Modulo.forms import HistorialCargosFilterForm
from Modulo.models import Empleado, Linea, Cargos, Historial_Cargos
from django.http import HttpResponse
from django.db.models import Q
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def filtrar_historial_cargos(request):
    # Obtener los filtros de la URL (GET)
    empleados = request.GET.getlist('Empleado')
    lineas = request.GET.getlist('Linea')
    cargos = request.GET.getlist('Cargo')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Construir la consulta con los filtros aplicados
    filtro = Q()
    if empleados:
        filtro &= Q(documentoId__in=empleados)
    if lineas:
        filtro &= Q(documentoId__LineaId__in=lineas)
    if cargos:
        filtro &= Q(cargoId__in=cargos)
    if fecha_inicio:
        filtro &= Q(FechaInicio__gte=fecha_inicio)
    if fecha_fin:
        filtro &= Q(FechaFin__lte=fecha_fin)

    # Obtener los datos del historial de cargos con los filtros aplicados
    historial = Historial_Cargos.objects.filter(filtro).select_related('documentoId', 'cargoId')

# Función para filtrar informe historial de cargos
def filtrar_historial_cargos(form, historial):
    documento = form.cleaned_data.get('Empleado')
    linea = form.cleaned_data.get('Linea')
    cargo = form.cleaned_data.get('Cargo')

    if documento:
        historial = historial.filter(documentoId__Documento__in=documento)
    if linea:
        historial = historial.filter(documentoId__LineaId__in=linea)
    if cargo:
        historial = historial.filter(cargoId__in=cargo)

    return historial

# Vista para mostrar el informe historial de cargos
def historial_cargos_filtrado(request):
    historial_info = []
    show_data = False  
    

    if request.method == 'GET':
        form = HistorialCargosFilterForm(request.GET)
        historial = Historial_Cargos.objects.select_related('documentoId', 'cargoId', 'documentoId__LineaId')

        if form.is_valid():
            historial = filtrar_historial_cargos(form, historial)
            historial_info = [
                {
                    'Documento': h.documentoId.Documento,
                    'Nombre': h.documentoId.Nombre,
                    'Linea': h.documentoId.LineaId.Linea if h.documentoId.LineaId else '',
                    'Cargo': h.cargoId.CargoId if h.cargoId else '',
                    'DescripcionCargo':h.cargoId.Cargo if h.cargoId else '',
                    'FechaInicio': h.FechaInicio.strftime('%Y-%m-%d') if h.FechaInicio else '',
                    'FechaFin': h.FechaFin.strftime('%Y-%m-%d') if h.FechaFin else 'Activo',
                } 
                for h in historial
            ]
            show_data = bool(historial_info)
    else:
        form = HistorialCargosFilterForm()
        

    context = {
        'form': form,
        'historial_cargos_info': historial_info,
        'show_data': show_data,
        'mensaje': "No se encontraron resultados para los filtros aplicados." if not historial_info else ""
    }

    return render(request, 'informes/informes_historial_cargos_index.html', context)

def exportar_historial_cargos_excel(request):
    historial_info = []

    if request.method == 'GET':
        # Extraer los valores correctamente y asegurarse de que sean individuales
        empleados = request.GET.getlist('Empleado')
        lineas = request.GET.getlist('Linea')
        cargos = request.GET.getlist('Cargo')

        # Convertir IDs numéricos a enteros si es necesario
        lineas = [int(l) for l in lineas if l.isdigit()]
        cargos = [int(c) for c in cargos if c.isdigit()]

        # Crear el diccionario con valores correctos
        data = {
            'Empleado': empleados,  
            'Linea': lineas,  
            'Cargo': cargos  
        }
        form = HistorialCargosFilterForm(data)

        if form.is_valid():
            historial = Historial_Cargos.objects.select_related('documentoId', 'documentoId__LineaId', 'cargoId')

            empleados = form.cleaned_data.get('Empleado') or []  
            lineas = form.cleaned_data.get('Linea') or []
            cargos = form.cleaned_data.get('Cargo') or []

            if empleados:
                historial = historial.filter(documentoId__in=empleados)
            if lineas:
                historial = historial.filter(documentoId__LineaId__in=lineas)
            if cargos:
                historial = historial.filter(cargoId__in=cargos)

            if historial.count() == 0:
                return HttpResponse("No se encontraron resultados para los filtros aplicados.")

            for registro in historial:
                historial_info.append({
                    'Documento': registro.documentoId.Documento,
                    'Nombre': registro.documentoId.Nombre,
                    'Linea': registro.documentoId.LineaId.Linea if registro.documentoId.LineaId else '',
                    'Cargo': registro.cargoId.CargoId if registro.cargoId else '',
                    'DescripcionCargo': registro.cargoId.Cargo if registro.cargoId else '',
                    'Fecha Inicio': registro.FechaInicio.strftime('%Y-%m-%d') if registro.FechaInicio else '',
                    'Fecha Fin': registro.FechaFin.strftime('%Y-%m-%d') if registro.FechaFin else 'Activo'
                })
        else:
            print(f"🛠 Errores en el formulario: {form.errors}")  
            return HttpResponse("Error en el formulario.")

    if not historial_info:
        return HttpResponse("No hay datos para exportar.")

    #Generar el Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Cargos"

    # Definir encabezados y su orden
    encabezados = ["Documento", "Nombre", "Linea", "Cargo", "DescripcionCargo", "Fecha Inicio", "Fecha Fin"]

    # Estilo para bordes
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), 
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Aplicar encabezados con estilos
    for col_num, header in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


    # Insertar datos en el Excel con formato adecuado
    for row_num, registro in enumerate(historial_info, 2):
        for col_num, key in enumerate(encabezados, 1):
            value = registro[key]
            cell = ws.cell(row=row_num, column=col_num, value=value)

            # Aplicar bordes a todas las celdas
            cell.border = thin_border

            # Alinear texto en las celdas
            if isinstance(value, str):
                cell.alignment = Alignment(horizontal="left")
            elif isinstance(value, datetime):
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "YYYY-MM-DD"  # Formato de fecha

    # Ajustar automáticamente el ancho de las columnas
    for col_num, col_name in enumerate(encabezados, 1):
        max_length = len(col_name)  # Iniciar con el tamaño del encabezado

        # Revisar todas las filas de la columna para determinar el ancho máximo
        for row_num in range(2, len(historial_info) + 2):  
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        # Ajustar el ancho con un pequeño margen extra
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = max_length + 2

    # Configurar respuesta HTTP con el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"historial_cargos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response