from django.shortcuts import render
from collections import defaultdict
from django.http import HttpResponse
from Modulo.forms import TarifaConsultorFilterForm
from Modulo.models import Consultores, Tarifa_Consultores
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

#Función para filtrar Tarifas de consultores
#Filtros: Nombre Consultor, Linea, Cargo, Años(Checkbox)
def filtrar_tarifas_consultores(form, consultores, consultores_tarifas):
    consultores = consultores.order_by('Nombre')
    nombre = form.cleaned_data.get('Nombre')
    linea = form.cleaned_data.get('LineaId')
    anios = form.cleaned_data.get('anio')

    if nombre:
        consultores = consultores.filter(Nombre=nombre)
    if linea:
        consultores = consultores.filter(LineaId=linea)
    if anios:
        consultores_tarifas = consultores_tarifas.filter(anio__in=anios)

    documentos_tarifas = consultores_tarifas.values_list('documentoId', flat=True)
    consultores = consultores.filter(Documento__in=documentos_tarifas)

    return consultores, consultores_tarifas

#Función para obtener información de las tarifas de consultores
def obtener_tarifas_consultores(consultores, tarifas, anios):
    consultor_tarifas_info = []
    tarifas_por_documento = defaultdict(list)

    # Organizar las tarifas por documentoId para optimizar consultas
    for tarifa in tarifas:
        tarifas_por_documento[tarifa.documentoId_id].append(tarifa)

    for consultor in consultores:
        # Obtener tarifas del consultor
        tarifas_consultor = tarifas_por_documento.get(consultor.Documento, [])

        # Crear el diccionario con los datos del consultor y sus tarifas
        datos_consultor = {
            'linea': consultor.LineaId.Linea,
            'documento': consultor.Documento,
            'nombre': consultor.Nombre,
            'certificado': 'SI' if consultor.Certificado else 'NO',
            'perfil': consultor.PerfilId.Perfil,
            'modulo': consultor.ModuloId.Modulo,
            'tarifas': {anio: [] for anio in anios}
        }

        for tarifa in tarifas_consultor:
            if tarifa.anio in anios:
                datos_consultor['tarifas'][tarifa.anio].append({
                    'mes': tarifa.mes,
                    'clienteID': tarifa.clienteID.Nombre_Cliente,
                    'tarifa_hora': tarifa.valorHora,
                    'tarifa_dia': tarifa.valorDia,
                    'tarifa_mes': tarifa.valorMes,
                    'moneda': tarifa.monedaId.Nombre,
                    'iva': tarifa.iva,
                    'rteFte': tarifa.rteFte,
                })

        # Ordenar las tarifas dentro de cada año, por cliente
        for anio in anios:
            datos_consultor['tarifas'][anio].sort(key=lambda x: x['clienteID'])

        consultor_tarifas_info.append(datos_consultor)

    return consultor_tarifas_info
        

def tarifas_consultores_filtrado(request):
    consultor_tarifas_info = []
    show_data = False  
    busqueda_realizada = False

    if request.method == 'GET':
        form = TarifaConsultorFilterForm(request.GET)

        if request.GET:
            busqueda_realizada = True

            if form.is_valid():
                # Inicializar consultores y tarifas
                consultores = Consultores.objects.all()
                tarifas = Tarifa_Consultores.objects.all()

                # Obtener información de los consultores y tarifas filtrados
                consultores, tarifas = filtrar_tarifas_consultores(form, consultores, tarifas)

                # Obtener los años seleccionados por el usuario
                anios = form.cleaned_data.get('anio')

                #Obtener informacion de las tarifas de los consultores
                consultor_tarifas_info = obtener_tarifas_consultores(consultores, tarifas, anios)
                show_data = bool(consultor_tarifas_info)  # Mostrar datos si hay resultados
        
    else: 
        form = TarifaConsultorFilterForm()

    context = {
        'form': form,
        'consultor_tarifas_info': consultor_tarifas_info,      
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'mensaje': "No se encontraron resultados para los filtros aplicados." if busqueda_realizada and not show_data else "No se ha realizado ninguna búsqueda aún."
    }

    return render(request, 'informes/informes_tarifas_consultores_index.html', context)

def exportar_tarifas_consultores_excel(request):
    consultor_tarifas_info = []

    if request.method == 'GET':
        form = TarifaConsultorFilterForm(request.GET)

        if form.is_valid():
            # Inicializar consultores y tarifas
            consultores = Consultores.objects.all()
            tarifas = Tarifa_Consultores.objects.all()

            # Obtener información de los consultores y tarifas filtrados
            consultores, tarifas = filtrar_tarifas_consultores(form, consultores, tarifas)

            # Obtener los años seleccionados por el usuario
            anios = form.cleaned_data.get('anio')

            # Obtener información de las tarifas de los consultores
            consultor_tarifas_info = obtener_tarifas_consultores(consultores, tarifas, anios)

            if not consultor_tarifas_info:
                return HttpResponse("No se encontraron resultados para los filtros aplicados.")

        else:
                print("Errores del formulario:", form.errors)
                return HttpResponse("No se encontraron resultados para los filtros aplicados.")
        
        # Preparar los datos para Excel
        data = []
        for consultor in consultor_tarifas_info:
            for anio in anios:
                for tarifa in consultor['tarifas'][anio]:
                    # Aquí construyes las filas con la estructura correcta
                    data.append([
                        consultor['linea'],
                        consultor['documento'],
                        consultor['nombre'],
                        consultor['certificado'],
                        consultor['perfil'],
                        consultor['modulo'],
                        anio,
                        tarifa['mes'],
                        tarifa['clienteID'],
                        tarifa['tarifa_hora'],
                        tarifa['tarifa_dia'],
                        tarifa['tarifa_mes'],
                        tarifa['moneda'],
                        tarifa['iva'],
                        tarifa['rteFte']
                    ])
        print("DATA:",data) 
        if data:
            # Crear un libro de trabajo y una hoja
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe tarifas de consultores"

            # Agregar encabezados
            encabezados = [ 
                "Linea", "Documento", "Nombre", "Certificado", "Perfil", "Modulo", 
                "Año", "Mes", "Cliente", "Tarifa por hora", "Tarifa por día", 
                "Tarifa por mes", "Moneda", "iva", "rteFte"
            ]
            for col_num, header in enumerate(encabezados, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Agregar datos a la hoja
            for row_num, row in enumerate(data, 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal='center')  # Centrar datos

            # Ajustar el ancho de las columnas
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # Crear respuesta de Excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f"tarifas_consultores_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            # Guardar el libro de trabajo en la respuesta  
            wb.save(response)

            return response

        else:
            return HttpResponse("No se encontraron datos para exportar.", status=404)
