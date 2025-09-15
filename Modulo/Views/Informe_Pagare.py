import logging
from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Q
from Modulo.forms import EmpleadoConPagareFilterForm
from Modulo.models import Empleado, Pagare
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from Modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required

logger = logging.getLogger(__name__)


def calcular_meses_condonados(fecha_inicio):
    if not fecha_inicio or not isinstance(fecha_inicio, date):
        return 0
    
    hoy = date.today()
    if fecha_inicio > hoy:
        return 0
    
    diferencia = relativedelta(hoy, fecha_inicio)
    meses = diferencia.years * 12 + diferencia.months
    
    # Si el día actual es igual o mayor al día de la fecha de inicio, cuenta como un mes condonado
    if hoy.day >= fecha_inicio.day:
        meses += 1
    
    return max(0, meses)

def filtrar_pagares(form, pagarés):
    empleados = form.cleaned_data.get('empleados_con_pagare')
    anio = form.cleaned_data.get('Anio')
    lineas = form.cleaned_data.get('LineaId')
    tipos = form.cleaned_data.get('tipo_pagare')
    estado = form.cleaned_data.get('estado_pagare')  # Nuevo filtro

    if empleados:
        pagarés = pagarés.filter(Documento__in=empleados.values_list('Documento', flat=True))
    if anio:
        pagarés = pagarés.filter(Fecha_Creacion_Pagare__year=anio)
    if lineas:
        pagarés = pagarés.filter(Documento__in=Empleado.objects.filter(LineaId__in=lineas).values_list('Documento', flat=True))
    if tipos:
        pagarés = pagarés.filter(Tipo_Pagare__in=tipos)
    if estado:  # Aplicar el filtro de estado CORREGIDO
        pagarés = pagarés.filter(estado=estado)
        
    return pagarés

@login_required
@verificar_permiso('can_manage_informe_pagare')
def informe_pagares(request):
    pagarés_info = []
    show_data = False
    fecha_actual = datetime.now().strftime("%Y-%m-%d")

    form = EmpleadoConPagareFilterForm(request.GET or None)
    pagarés = Pagare.objects.select_related('Tipo_Pagare').all()

    print(f"Datos recibidos en el formulario: {request.GET}")

    if form.is_valid():
        pagarés = filtrar_pagares(form, pagarés)
        documentos = pagarés.values_list('Documento', flat=True).distinct()
        empleados = Empleado.objects.filter(Documento__in=documentos).select_related('LineaId', 'CargoId')
        empleados_dict = {emp.Documento: emp for emp in empleados}

        print(f"Pagarés recuperados: {pagarés}")
        print(f"Empleados recuperados: {empleados}")

        for pagare in pagarés:
            empleado = empleados_dict.get(pagare.Documento)

            if empleado:
                fecha_inicio_condonacion = pagare.Fecha_inicio_Condonacion
                meses_condonados = 0
                meses_por_condonar = pagare.Meses_de_condonacion if pagare.Meses_de_condonacion else 0

                # Calcular meses condonados y meses por condonar basándose en la fecha inicio
                if fecha_inicio_condonacion:
                    meses_condonados = calcular_meses_condonados(fecha_inicio_condonacion)
                    meses_condonados = min(meses_condonados, meses_por_condonar)
                    meses_por_condonar = max(0, meses_por_condonar - meses_condonados)

                valor_condonado = 0
                if pagare.Meses_de_condonacion and pagare.Meses_de_condonacion > 0:
                    valor_condonado = (float(pagare.Valor_Pagare) / pagare.Meses_de_condonacion) * meses_condonados

                deuda_pagare = float(pagare.Valor_Pagare) - valor_condonado

                meses_por_condonar = 0
                if pagare.Meses_de_condonacion:
                    meses_por_condonar = max(0, pagare.Meses_de_condonacion - meses_condonados)


                porcentaje_ejecucion = 0.0
                if pagare.Meses_de_condonacion and pagare.Meses_de_condonacion > 0:
                    porcentaje_ejecucion = min((meses_condonados / pagare.Meses_de_condonacion) * 100, 100)


                registro = {
                    'Documento': empleado.Documento,
                    'Nombre': empleado.Nombre,
                    'TipoPagare': pagare.Tipo_Pagare.Desc_Tipo_Pagare if pagare.Tipo_Pagare else 'N/A',
                    'Descripcion':pagare.descripcion,
                    'Estado': pagare.estado,
                    'Linea': empleado.LineaId.Linea if empleado.LineaId else 'N/A',
                    'Cargo': empleado.CargoId.Cargo if empleado.CargoId else 'N/A',
                    'FechaIngreso': empleado.FechaIngreso.strftime("%Y-%m-%d") if empleado.FechaIngreso else "",
                    'FechaOperacion': empleado.FechaOperacion.strftime("%Y-%m-%d") if empleado.FechaOperacion else "",
                    'ConsecutivoPagare': pagare.Pagare_Id,
                    'FechaPagare': pagare.Fecha_Creacion_Pagare.strftime("%Y-%m-%d") if pagare.Fecha_Creacion_Pagare else "",
                    'FechaInicioCondonacion': fecha_inicio_condonacion.strftime("%Y-%m-%d") if fecha_inicio_condonacion else "",
                    'FechaFinCondonacion': pagare.Fecha_fin_Condonacion.strftime("%Y-%m-%d") if pagare.Fecha_fin_Condonacion else "",
                    'ValorPagare': float(pagare.Valor_Pagare),
                    'MesesCondonacion': pagare.Meses_de_condonacion if pagare.Meses_de_condonacion else 0,
                    'PorcentajeEjecucion': float(pagare.porcentaje_ejecucion) if pagare.porcentaje_ejecucion is not None else 0.0,
                    'FechaRetiro': empleado.FechaRetiro.strftime("%Y-%m-%d") if empleado.FechaRetiro else "",
                    'ValorCondonado': valor_condonado,
                    'DeudaPagare': deuda_pagare,
                    'MesesCondonados': meses_condonados,
                    'MesesPorCondonar': meses_por_condonar,
                    'pagare_id': pagare.Pagare_Id
                }
                pagarés_info.append(registro)

        print(f"Datos procesados para la tabla: {pagarés_info}")

        show_data = bool(pagarés_info)
        request.session['pagares_info'] = pagarés_info  # Guardar los datos en la sesión
    else:
        print(f"Errores en el formulario: {form.errors}")

    context = {
        'form': form,
        'pagares_info': pagarés_info,
        'show_data': show_data,
        'mensaje': "No se encontraron resultados para los filtros aplicados" if not pagarés_info else "",
        'fecha_actual': fecha_actual
    }   

    return render(request, 'Pagare/Informe_Pagare.html', context)

@login_required
@verificar_permiso('can_manage_informe_pagare')
def exportar_pagares_excel(request):
    # Obtener datos de la sesión
    pagarés_info = request.session.get('pagares_info', [])
    logger.debug(f"Pagarés en la sesión: {pagarés_info}")

    if not pagarés_info:
        return HttpResponse("No hay datos para exportar.")
    
    # Determinar qué registros exportar
    export_all = request.GET.get('export_all', '0') == '1'
    selected_ids_str = request.GET.get('selected_ids', '')
    selected_ids = selected_ids_str.split(',') if selected_ids_str else []
    fecha_informe = request.GET.get('fecha_informe', datetime.now().strftime("%Y-%m-%d"))
    
    if export_all:
        # Exportar todos los registros filtrados
        pass
    elif selected_ids:
        # Exportar solo los seleccionados
        pagarés_info = [p for p in pagarés_info if str(p.get('ConsecutivoPagare', '')) in selected_ids]
    else:
        # Exportar todo si no hay selección
        pass

    if not pagarés_info:
        return HttpResponse("No hay datos para exportar.")

    wb = Workbook()
    
    # Eliminar la hoja por defecto si vamos a crear varias
    if len(pagarés_info) > 1 or export_all:
        wb.remove(wb.active)
    
    # Configuración de estilos
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))
    
    # Función para crear hoja de resumen
    def crear_hoja_resumen(ws, pagares, fecha):
        # Fecha del informe
        ws.cell(row=1, column=1, value="Fecha del informe:").font = header_font
        ws.cell(row=1, column=2, value=fecha)
        
        # Encabezados de la tabla resumen
        columnas = [
            "Documento", "Nombre", "Tipo Pagaré", "Descripción", "Línea", "Cargo", 
            "Fecha Ingreso", "Fecha Inicio Operación", "Consecutivo Pagare", 
            "Fecha Pagare", "Fecha Inicio Condonación", "Fecha Fin Condonación", 
            "Valor del Pagaré", "Valor Condonado", "Deuda Pagare", 
            "Meses Condonación", "Meses Condonados", "Meses por Condonar", 
            "%Ejecución", "Fecha Retiro"
        ]
        
        # Escribir encabezados (fila 3)
        for col_num, header in enumerate(columnas, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos resumen (a partir de fila 4)
        row_num = 4
        for registro in pagares:
            ws.cell(row=row_num, column=1, value=registro.get('Documento', ''))
            ws.cell(row=row_num, column=2, value=registro.get('Nombre', ''))
            ws.cell(row=row_num, column=3, value=registro.get('TipoPagare', ''))
            ws.cell(row=row_num, column=4, value=registro.get('Descripcion', ''))
            ws.cell(row=row_num, column=5, value=registro.get('Linea', ''))
            ws.cell(row=row_num, column=6, value=registro.get('Cargo', ''))
            ws.cell(row=row_num, column=7, value=registro.get('FechaIngreso', ''))
            ws.cell(row=row_num, column=8, value=registro.get('FechaOperacion', ''))
            ws.cell(row=row_num, column=9, value=registro.get('ConsecutivoPagare', ''))
            ws.cell(row=row_num, column=10, value=registro.get('FechaPagare', ''))
            ws.cell(row=row_num, column=11, value=registro.get('FechaInicioCondonacion', ''))
            ws.cell(row=row_num, column=12, value=registro.get('FechaFinCondonacion', ''))
            ws.cell(row=row_num, column=13, value=registro.get('ValorPagare', 0))
            ws.cell(row=row_num, column=14, value=registro.get('ValorCondonado', 0))
            ws.cell(row=row_num, column=15, value=registro.get('DeudaPagare', 0))
            ws.cell(row=row_num, column=16, value=registro.get('MesesCondonacion', 0))
            ws.cell(row=row_num, column=17, value=registro.get('MesesCondonados', 0))
            ws.cell(row=row_num, column=18, value=registro.get('MesesPorCondonar', 0))
            ws.cell(row=row_num, column=19, value=registro.get('PorcentajeEjecucion', 0))
            ws.cell(row=row_num, column=20, value=registro.get('FechaRetiro', ''))
            
            # Aplicar bordes a todas las celdas
            for col in range(1, 21):
                ws.cell(row=row_num, column=col).border = border
            
            row_num += 1
        
        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        return ws
    
    # Función para crear hoja individual por pagaré
    def crear_hoja_individual(ws, pagare, fecha):
        # Fecha del informe
        ws.cell(row=1, column=1, value="Fecha del informe:").font = header_font
        ws.cell(row=1, column=2, value=fecha)
        
        # Título Resumen Pagaré
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=20)
        title_cell = ws.cell(row=3, column=1, value="RESUMEN PAGARÉ")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Encabezados de la tabla resumen
        columnas = [
            "Documento", "Nombre", "Tipo Pagaré", "Descripción", "Línea", "Cargo", 
            "Fecha Ingreso", "Fecha Inicio Operación", "Consecutivo Pagare", 
            "Fecha Pagare", "Fecha Inicio Condonación", "Fecha Fin Condonación", 
            "Valor del Pagaré", "Valor Condonado", "Deuda Pagare", 
            "Meses Condonación", "Meses Condonados", "Meses por Condonar", 
            "%Ejecución", "Fecha Retiro"
        ]
        
        # Escribir encabezados (fila 5)
        for col_num, header in enumerate(columnas, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos resumen (fila 6)
        ws.cell(row=6, column=1, value=pagare.get('Documento', ''))
        ws.cell(row=6, column=2, value=pagare.get('Nombre', ''))
        ws.cell(row=6, column=3, value=pagare.get('TipoPagare', ''))
        ws.cell(row=6, column=4, value=pagare.get('Descripcion', ''))
        ws.cell(row=6, column=5, value=pagare.get('Linea', ''))
        ws.cell(row=6, column=6, value=pagare.get('Cargo', ''))
        ws.cell(row=6, column=7, value=pagare.get('FechaIngreso', ''))
        ws.cell(row=6, column=8, value=pagare.get('FechaOperacion', ''))
        ws.cell(row=6, column=9, value=pagare.get('ConsecutivoPagare', ''))
        ws.cell(row=6, column=10, value=pagare.get('FechaPagare', ''))
        ws.cell(row=6, column=11, value=pagare.get('FechaInicioCondonacion', ''))
        ws.cell(row=6, column=12, value=pagare.get('FechaFinCondonacion', ''))
        ws.cell(row=6, column=13, value=pagare.get('ValorPagare', 0))
        ws.cell(row=6, column=14, value=pagare.get('ValorCondonado', 0))
        ws.cell(row=6, column=15, value=pagare.get('DeudaPagare', 0))
        ws.cell(row=6, column=16, value=pagare.get('MesesCondonacion', 0))
        ws.cell(row=6, column=17, value=pagare.get('MesesCondonados', 0))
        ws.cell(row=6, column=18, value=pagare.get('MesesPorCondonar', 0))
        ws.cell(row=6, column=19, value=pagare.get('PorcentajeEjecucion', 0))
        ws.cell(row=6, column=20, value=pagare.get('FechaRetiro', ''))
        
        # Aplicar bordes a todas las celdas
        for col in range(1, 21):
            ws.cell(row=6, column=col).border = border
        
        # Espacio antes del detalle
        row_num = 8
        
        # Título Detalle de Cuotas
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        title_cell = ws.cell(row=row_num, column=1, value="DETALLE DE CUOTAS")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        row_num += 1
        
        # Encabezados para detalle
        encabezados_detalle = ["Documento", "Nombre", "No. Cuota", "Fecha Cuota", "Valor cuota", "Estado"]
        
        # Escribir encabezados
        for col_num, header in enumerate(encabezados_detalle, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1
        
        # Generar datos detalle de cuotas
        valor_pagare = float(pagare.get('ValorPagare', 0))
        meses_condonacion = int(pagare.get('MesesCondonacion', 0))
        fecha_inicio_condonacion_str = pagare.get('FechaInicioCondonacion')
        documento = pagare.get('Documento', '')
        nombre = pagare.get('Nombre', '')
        
        # Convertir fecha_inicio_condonacion a un objeto date si es una cadena
        fecha_inicio_condonacion = None
        if fecha_inicio_condonacion_str:
            try:
                fecha_inicio_condonacion = datetime.strptime(fecha_inicio_condonacion_str, "%Y-%m-%d").date()
            except ValueError:
                fecha_inicio_condonacion = None
        
        # Generar cuotas
        if valor_pagare > 0 and meses_condonacion > 0:
            valor_cuota = valor_pagare / meses_condonacion
            
            for i in range(1, meses_condonacion + 1):
                fecha_cuota_str = ""
                estado = ""
                
                if fecha_inicio_condonacion:
                    # Calcular fecha de la cuota
                    fecha_cuota = fecha_inicio_condonacion + relativedelta(months=(i - 1))
                    fecha_cuota_str = fecha_cuota.strftime("%d/%m/%Y")
                    
                    # Determinar estado de la cuota
                    estado = "Pago" if fecha_cuota <= date.today() else "Pendiente"
                
                # Escribir fila en Excel
                ws.cell(row=row_num, column=1, value=documento)
                ws.cell(row=row_num, column=2, value=nombre)
                ws.cell(row=row_num, column=3, value=i)
                ws.cell(row=row_num, column=4, value=fecha_cuota_str)
                ws.cell(row=row_num, column=5, value=valor_cuota)
                ws.cell(row=row_num, column=6, value=estado)
                
                # Aplicar bordes a todas las celdas
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = border
                
                row_num += 1
        
        # Ajustar anchos de columna
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            max_length = 0
            col_idx = ord(col_letter) - 64
            for row in range(1, row_num):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    try:
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        return ws
    
    # Lógica para determinar cómo exportar
    if len(pagarés_info) == 1 and not export_all:
        # Caso 1: Un solo pagaré seleccionado manualmente
        ws = wb.active
        ws.title = "Pagare"
        pagare = pagarés_info[0]
        
        # Agregar resumen
        ws.cell(row=1, column=1, value="Fecha del informe:").font = header_font
        ws.cell(row=1, column=2, value=fecha_informe)
        
        # Título Resumen Pagaré
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=20)
        title_cell = ws.cell(row=3, column=1, value="RESUMEN PAGARÉ")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Encabezados de la tabla resumen
        columnas = [
            "Documento", "Nombre", "Tipo Pagaré", "Descripción", "Línea", "Cargo", 
            "Fecha Ingreso", "Fecha Inicio Operación", "Consecutivo Pagare", 
            "Fecha Pagare", "Fecha Inicio Condonación", "Fecha Fin Condonación", 
            "Valor del Pagaré", "Valor Condonado", "Deuda Pagare", 
            "Meses Condonación", "Meses Condonados", "Meses por Condonar", 
            "%Ejecución", "Fecha Retiro"
        ]
        
        # Escribir encabezados (fila 5)
        for col_num, header in enumerate(columnas, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos resumen (fila 6)
        ws.cell(row=6, column=1, value=pagare.get('Documento', ''))
        ws.cell(row=6, column=2, value=pagare.get('Nombre', ''))
        ws.cell(row=6, column=3, value=pagare.get('TipoPagare', ''))
        ws.cell(row=6, column=4, value=pagare.get('Descripcion', ''))
        ws.cell(row=6, column=5, value=pagare.get('Linea', ''))
        ws.cell(row=6, column=6, value=pagare.get('Cargo', ''))
        ws.cell(row=6, column=7, value=pagare.get('FechaIngreso', ''))
        ws.cell(row=6, column=8, value=pagare.get('FechaOperacion', ''))
        ws.cell(row=6, column=9, value=pagare.get('ConsecutivoPagare', ''))
        ws.cell(row=6, column=10, value=pagare.get('FechaPagare', ''))
        ws.cell(row=6, column=11, value=pagare.get('FechaInicioCondonacion', ''))
        ws.cell(row=6, column=12, value=pagare.get('FechaFinCondonacion', ''))
        ws.cell(row=6, column=13, value=pagare.get('ValorPagare', 0))
        ws.cell(row=6, column=14, value=pagare.get('ValorCondonado', 0))
        ws.cell(row=6, column=15, value=pagare.get('DeudaPagare', 0))
        ws.cell(row=6, column=16, value=pagare.get('MesesCondonacion', 0))
        ws.cell(row=6, column=17, value=pagare.get('MesesCondonados', 0))
        ws.cell(row=6, column=18, value=pagare.get('MesesPorCondonar', 0))
        ws.cell(row=6, column=19, value=pagare.get('PorcentajeEjecucion', 0))
        ws.cell(row=6, column=20, value=pagare.get('FechaRetiro', ''))
        
        # Aplicar bordes a todas las celdas
        for col in range(1, 21):
            ws.cell(row=6, column=col).border = border
        
        # Espacio antes del detalle
        row_num = 8
        
        # Título Detalle de Cuotas
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        title_cell = ws.cell(row=row_num, column=1, value="DETALLE DE CUOTAS")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        row_num += 1
        
        # Encabezados para detalle
        encabezados_detalle = ["Documento", "Nombre", "No. Cuota", "Fecha Cuota", "Valor cuota", "Estado"]
        
        # Escribir encabezados
        for col_num, header in enumerate(encabezados_detalle, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1
        
        # Generar datos detalle de cuotas
        valor_pagare = float(pagare.get('ValorPagare', 0))
        meses_condonacion = int(pagare.get('MesesCondonacion', 0))
        fecha_inicio_condonacion_str = pagare.get('FechaInicioCondonacion')
        documento = pagare.get('Documento', '')
        nombre = pagare.get('Nombre', '')
        
        # Convertir fecha_inicio_condonacion a un objeto date si es una cadena
        fecha_inicio_condonacion = None
        if fecha_inicio_condonacion_str:
            try:
                fecha_inicio_condonacion = datetime.strptime(fecha_inicio_condonacion_str, "%Y-%m-%d").date()
            except ValueError:
                fecha_inicio_condonacion = None
        
        # Generar cuotas
        if valor_pagare > 0 and meses_condonacion > 0:
            valor_cuota = valor_pagare / meses_condonacion
            
            for i in range(1, meses_condonacion + 1):
                fecha_cuota_str = ""
                estado = ""
                
                if fecha_inicio_condonacion:
                    # Calcular fecha de la cuota
                    fecha_cuota = fecha_inicio_condonacion + relativedelta(months=(i - 1))
                    fecha_cuota_str = fecha_cuota.strftime("%d/%m/%Y")
                    
                    # Determinar estado de la cuota
                    estado = "Pago" if fecha_cuota <= date.today() else "Pendiente"
                
                # Escribir fila en Excel
                ws.cell(row=row_num, column=1, value=documento)
                ws.cell(row=row_num, column=2, value=nombre)
                ws.cell(row=row_num, column=3, value=i)
                ws.cell(row=row_num, column=4, value=fecha_cuota_str)
                ws.cell(row=row_num, column=5, value=valor_cuota)
                ws.cell(row=row_num, column=6, value=estado)
                
                # Aplicar bordes a todas las celdas
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = border
                
                row_num += 1
        
        # Ajustar anchos de columna
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            max_length = 0
            col_idx = ord(col_letter) - 64
            for row in range(1, row_num):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    try:
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width
        
    else:
        # Caso 2: Múltiples pagarés o exportar todo
        # Crear hoja de resumen
        ws_resumen = wb.create_sheet(title="Resumen")
        crear_hoja_resumen(ws_resumen, pagarés_info, fecha_informe)
        
        # Crear hojas individuales para cada pagaré
        for pagare in pagarés_info:
            nombre_hoja = f"{pagare['Documento']}-{pagare['ConsecutivoPagare']}"[:31]
            ws_individual = wb.create_sheet(title=nombre_hoja)
            crear_hoja_individual(ws_individual, pagare, fecha_informe)
    
    # Guardar y devolver el Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="informe_pagares_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response