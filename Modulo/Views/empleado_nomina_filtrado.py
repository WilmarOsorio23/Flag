#Informe de Salarios

# Librerías estándar y de terceros
from datetime import date
from datetime import datetime
from collections import defaultdict, Counter
from dateutil.relativedelta import relativedelta

# Django
from django.http import HttpResponse
from django.shortcuts import render

# Formularios y modelos del módulo
from Modulo.forms import EmpleadoFilterForm
from Modulo.models import Empleado, Nomina

# Librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from Modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required

@login_required
@verificar_permiso('can_manage_empleados')

# Función para filtrar empleados y nóminas
def filtrar_empleados_y_nominas(form, empleados, nominas):
    """
    Aplica filtros del formulario sobre empleados y nóminas.
    - Filtra por nombre, línea, cargo y año.
    - Devuelve los empleados y nóminas que cumplen los filtros aplicados.
    """
    empleados = empleados.order_by('Documento')  
    nominas = nominas.order_by('-Anio')  

    nombre = form.cleaned_data.get('Nombre')    
    linea = form.cleaned_data.get('LineaId')
    anio = form.cleaned_data.get('Anio')
    cargo = form.cleaned_data.get('Cargo')

    if nombre:
        empleados = empleados.filter(Nombre__icontains=nombre)
    if linea:
        empleados = empleados.filter(LineaId=linea)
    if cargo:
        empleados = empleados.filter(CargoId=cargo)

    if anio:
        nominas = nominas.filter(Anio=anio)

    documentos_nominas = nominas.values_list('Documento', flat=True)
    empleados = empleados.filter(Documento__in=documentos_nominas)

    return empleados, nominas

# Función para calcular la información del empleado
def obtener_empleado_info(empleados, nominas, meses):
    """
    Une la información de empleados con sus registros de nómina.
    - Agrupa la información de salarios y clientes mes a mes.
    - Calcula los años en la empresa.
    Devuelve una lista de diccionarios con los datos completos por empleado.
    """
    def limpiar_documento(doc):
        #Elimina puntos, espacios y convierte a string sin espacios adicionales
        return str(doc).replace('.', '').replace(' ', '').strip()
    
    empleado_info = []
    nominas_por_documento_y_mes = defaultdict(dict)

    # Organizar las nóminas por Documento, Año y Mes para optimizar consultas
    for nomina in nominas:
        doc_limpio = limpiar_documento(nomina.Documento.Documento)
        nominas_por_documento_y_mes[(doc_limpio, str(nomina.Anio))][str(nomina.Mes).zfill(2)] = nomina

    for empleado in empleados:
        # Obtener años únicos de nómina para el empleado
        anios_nominas = nominas.filter(Documento=empleado.Documento).values_list('Anio', flat=True).distinct()

        for anio in anios_nominas:
            # Salarios y clientes mes a mes para el año específico
            salarios_meses = []
            for mes in range(1, 13):
                doc_limpio = limpiar_documento(empleado.Documento)
                mes_nomina = nominas_por_documento_y_mes.get((empleado.Documento, str(anio)), {}).get(str(mes).zfill(2))

                if mes_nomina:
                    salario = mes_nomina.Salario
                    cliente = mes_nomina.Cliente.Nombre_Cliente if mes_nomina and mes_nomina.Cliente else ""
                else:
                    salario = cliente = ""
                salarios_meses.append({'salario': salario, 'cliente': cliente})

            # Calcular campo "Años en Flag"
            fecha_actual = date.today()
            años_en_flag = relativedelta(fecha_actual, empleado.FechaIngreso).years

            # Agregar datos a la lista
            empleado_info.append({
                'nombre_linea': empleado.LineaId.Linea,
                'documento_colaborador': empleado.Documento,
                'nombre_colaborador': empleado.Nombre,
                'cargo': empleado.CargoId.Cargo,
                'perfil': empleado.PerfilId.Perfil,
                'modulo': empleado.ModuloId.Modulo,
                'certificado': 'SI' if empleado.CertificadoSAP else 'NO',
                'fecha_ingreso': empleado.FechaIngreso,
                'años_en_flag': años_en_flag,
                'año': anio,
                'meses': salarios_meses,
            })
    return empleado_info

# Informe de Nómina de Empleados
def empleado_nomina_filtrado(request):
    """
    Vista para mostrar el informe de nómina de empleados.
    - Aplica filtros GET desde el formulario.
    - Genera contexto con información de salarios por mes.
    """
    meses = "Enero Febrero Marzo Abril Mayo Junio Julio Agosto Septiembre Octubre Noviembre Diciembre".split()
    empleado_info = []  
    show_data = False  
    busqueda_realizada = False

    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)

        if request.GET:
            busqueda_realizada = True
        
            if form.is_valid():
                # Inicializar empleados y nominas
                empleados = Empleado.objects.all()
                nominas = Nomina.objects.all()

                # Obtener información de los empleados filtrada
                empleados, nominas = filtrar_empleados_y_nominas(form, empleados, nominas)
                
                # Obtener información de los empleados
                empleado_info = obtener_empleado_info(empleados, nominas, meses)
                show_data = bool(empleado_info)  # Mostrar datos si hay resultados
            
    else: 
        form = EmpleadoFilterForm()

    #Aplicar lógica para cards
    # Total con Certificación SAP
    total_certificados = sum(1 for e in empleado_info if e['certificado'] == 'SI')

    # Conteo por Línea
    lineas = [e['nombre_linea'] for e in empleado_info if e['nombre_linea']]
    conteo_lineas = dict(Counter(lineas))

    context = {
        "meses": meses,
        'form': form,
        'empleado_info': empleado_info,      
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'total_certificados': total_certificados,
        'conteo_lineas': conteo_lineas,
        'lineas_labels': list(conteo_lineas.keys()),
        'lineas_data': list(conteo_lineas.values()),
        'mensaje': "No se encontraron resultados para los filtros aplicados." if busqueda_realizada and not show_data else "No se ha realizado ninguna búsqueda aún."
    }

    return render(request, 'informes/informes_salarios_index.html', context)

# Funcionalidad para descargar los resultados en Excel
def exportar_nomina_excel(request):
    """
    Exporta a Excel la información de nómina filtrada por el formulario.
    - Recupera los mismos filtros de la vista.
    - Genera un archivo Excel con la información mensual de salarios y clientes.
    """
    meses = "Enero Febrero Marzo Abril Mayo Junio Julio Agosto Septiembre Octubre Noviembre Diciembre".split()
    empleado_info = []  
    
    # Reutilizar la lógica de filtrado de tu vista actual
    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)

        if form.is_valid():
            # Inicializar empleados y nominas
            empleados = Empleado.objects.all()
            nominas = Nomina.objects.all()

            # Obtener información de los empleados filtrada
            empleados, nominas = filtrar_empleados_y_nominas(form, empleados, nominas)
            
            # Obtener información de los empleados
            empleado_info = obtener_empleado_info(empleados, nominas, meses)

        if not empleado_info:
            return HttpResponse("No se encontraron datos para exportar.", status=404)
        
        if not form.is_valid(): 
            return HttpResponse("Filtros no válidos. Por favor, revisa los criterios ingresados.", status=400)

    # Preparar datos para Excel
    data = []
    for empleado in empleado_info:
        fila = {
            'Nombre Línea': empleado['nombre_linea'],
            'Documento Colaborador': empleado['documento_colaborador'],
            'Nombre Colaborador': empleado['nombre_colaborador'],
            'Cargo': empleado['cargo'],
            'Perfil': empleado['perfil'],
            'Módulo': empleado['modulo'],
            'Certificado SAP': empleado['certificado'],
            'Fecha Ingreso': empleado['fecha_ingreso'],
            'Años en Flag': empleado['años_en_flag'],
            'Año': empleado['año']
        }

        # Agregar salarios y clientes por mes
        for i, mes_data in enumerate(empleado['meses'], 1):
            fila[f'Salario {meses[i-1]}'] = mes_data['salario']
            fila[f'Cliente {meses[i-1]}'] = mes_data['cliente']

        data.append(fila)

    # Crear un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe de Nómina"

    # Agregar encabezados
    for col in data[0].keys():
        cell = ws.cell(row=1, column=list(data[0].keys()).index(col) + 1, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Agregar datos del DataFrame a la hoja
    for r_idx, row in enumerate(data, 2):
        for c_idx, value in enumerate(row.values(), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')  # Centrar datos

    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Crear respuesta de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"nomina_Empleados_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

     # Guardar el libro de trabajo en la respuesta  
    wb.save(response)

    return response