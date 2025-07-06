# Informe de Consultores

# Django
from django.shortcuts import render
from django.http import HttpResponse

# Formularios y modelos del módulo
from Modulo.forms import ConsultorFilterForm
from Modulo.models import Consultores

# Librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Librerías estándar y de terceros
from datetime import datetime
from collections import Counter

def filtrar_consultores(form, consultores):
    """
    Aplica los filtros del formulario al conjunto de consultores.
    - Filtra por nombre, línea, módulo, certificación, perfil y estado.
    Devuelve los consultores que cumplen con los criterios seleccionados.
    """
    consultores = consultores.order_by('Nombre')
    nombre = form.cleaned_data.get('Nombre')
    linea = form.cleaned_data.get('LineaId')
    modulo = form.cleaned_data.get('ModuloId')
    certificacion = form.cleaned_data.get('Certificacion')
    perfil = form.cleaned_data.get('PerfilId')
    estado = form.cleaned_data.get('Estado')

    if nombre:
        consultores = consultores.filter(Nombre=nombre)
    if linea:
        consultores = consultores.filter(LineaId=linea)
    if modulo:
        consultores = consultores.filter(ModuloId=modulo)
    if certificacion:
        consultores = consultores.filter(Certificado=bool(int(certificacion)))
    if perfil:
        consultores = consultores.filter(PerfilId=perfil)
    if estado:
        consultores = consultores.filter(Estado=bool(int(estado)))

    return consultores

#Función para obtener información de los consultores
def obtener_consultores(consultores):
    """
    Procesa y organiza los datos de los consultores.
    Devuelve una lista de diccionarios con la información estructurada para presentación o exportación.
    """
    consultor_info = []

    for consultor in consultores:
         # Convertir Fecha_Nacimiento a objeto datetime.date si es una cadena
        fecha_nacimiento = consultor.Fecha_Nacimiento
        if isinstance(fecha_nacimiento, str):
            try:
                fecha_nacimiento = datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
            except ValueError:
                fecha_nacimiento = None

        datos_consultor = {
            'documento': consultor.Documento,
            'tipo_documento': consultor.TipoDocumentoID.Nombre,
            'nombre': consultor.Nombre,
            'empresa': consultor.Empresa,
            'profesion': consultor.Profesion,
            'linea': consultor.LineaId.Linea,
            'modulo': consultor.ModuloId.Modulo,
            'perfil': consultor.PerfilId.Perfil,
            'estado': 'Activo' if consultor.Estado else 'Inactivo',
            'fecha_ingreso': consultor.Fecha_Ingreso.strftime('%Y-%m-%d') if consultor.Fecha_Ingreso else '',
            'fecha_retiro': consultor.Fecha_Retiro.strftime('%Y-%m-%d') if consultor.Fecha_Retiro else '',
            'fecha_operacion': consultor.Fecha_Operacion.strftime('%Y-%m-%d') if consultor.Fecha_Operacion else '',
            'certificado':  'SI' if consultor.Certificado else 'NO',
            'certificaciones': consultor.Certificaciones if consultor.Certificaciones else '',
            'fecha_nacimiento': fecha_nacimiento.strftime('%Y-%m-%d') if fecha_nacimiento else '',
            'anio_evaluacion': consultor.Anio_Evaluacion if consultor.Anio_Evaluacion else '',
            'nota_evaluacion': consultor.NotaEvaluacion if consultor.NotaEvaluacion else '',
            'direccion': consultor.Direccion if consultor.Direccion else '',
            'telefono': consultor.Telefono if consultor.Telefono else '',
        }

        consultor_info.append(datos_consultor)
    return consultor_info

def consultores_filtrado(request):
    """
    Vista para renderizar el informe de consultores.
    - Usa filtros vía GET para aplicar criterios.
    - Construye contexto con métricas agregadas, datos y filtros.
    """
    consultor_info = []
    show_data = False  
    busqueda_realizada = False

    if request.method == 'GET':
        form = ConsultorFilterForm(request.GET)

        if request.GET:
            busqueda_realizada = True

            if form.is_valid():
                # Inicializar consultores
                consultores = Consultores.objects.all()

                # Realizar el filtro de los consultores
                consultores = filtrar_consultores(form, consultores)

                # Obtener información de consultores 
                consultor_info = obtener_consultores(consultores)
                show_data = bool(consultor_info)  # Mostrar datos si hay resultados
        
    else: 
        form = ConsultorFilterForm()

    # Cálculo de métricas y resumenes (cards)
    total_consultores = len(consultor_info)
    activos = sum(1 for c in consultor_info if c['estado'] == 'Activo')
    inactivos = total_consultores - activos
    certificados = sum(1 for c in consultor_info if c['certificado'] == 'SI')

    lineas = [c['linea'] for c in consultor_info if c['linea']]
    conteo_lineas = dict(Counter(lineas))

    modulos = [c['modulo'] for c in consultor_info if c['modulo']]
    conteo_modulos = dict(Counter(modulos))

    lineas_labels = list(conteo_lineas.keys())
    lineas_data = list(conteo_lineas.values())
    modulos_labels = list(conteo_modulos.keys())
    modulos_data = list(conteo_modulos.values())
    estado_labels = ['Activo', 'Inactivo']
    estado_data = [activos, inactivos]

    context = {
        'form': form,
        'consultor_info': consultor_info,      
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'total_consultores': total_consultores,
        'activos': activos,
        'inactivos': inactivos,
        'certificados': certificados,
        'conteo_lineas': conteo_lineas,
        'conteo_modulos': conteo_modulos,
        'lineas_labels': lineas_labels,
        'lineas_data': lineas_data,
        'modulos_labels': modulos_labels,
        'modulos_data': modulos_data,
        'estado_labels': estado_labels,
        'estado_data': estado_data,
        'mensaje': (
            "No se encontraron resultados para los filtros aplicados."
            if busqueda_realizada and not show_data
            else "No se ha realizado ninguna búsqueda aún."
        )
    }

    return render(request, 'informes/informes_consultores_index.html', context)


def exportar_consultores_excel(request):
    """
    Exporta los datos filtrados de consultores a un archivo Excel.
    - Reutiliza lógica de filtrado y transformación.
    - Aplica estilo y formato a los datos exportados usando OpenPyXL.
    """
    consultor_info = []
    
    if request.method == 'GET':
        form = ConsultorFilterForm(request.GET)

        if form.is_valid():
            # Inicializar consultores
            consultores = Consultores.objects.all()

            # Obtener información de los empleados filtrada
            consultores = filtrar_consultores(form, consultores)
            
            # Obtener información de los consultores
            consultor_info = obtener_consultores(consultores)
                        
            if not consultor_info:
                return HttpResponse("No se encontraron datos para exportar.")
                
        else: 
                return HttpResponse("Filtros no válidos.", status=400)

        #Preparar datos para excel 
        data = []
        for consultor in consultor_info:
            fila = {
                'Documento': consultor['documento'],
                'Tipo Documento': consultor['tipo_documento'],
                'Nombre': consultor['nombre'],
                'Empresa': consultor['empresa'],
                'Profesión': consultor['profesion'],
                'Línea': consultor['linea'],
                'Módulo': consultor['modulo'],
                'Perfil': consultor['perfil'],
                'Estado': consultor['estado'],
                'Fecha Ingreso': consultor['fecha_ingreso'],
                'Fecha Retiro': consultor['fecha_retiro'],
                'Fecha Operación': consultor['fecha_operacion'],
                'Certificado': consultor['certificado'],
                'Certificaciones': consultor['certificaciones'],
                'Fecha Nacimiento': consultor['fecha_nacimiento'],
                'Año Evaluación': consultor['anio_evaluacion'],
                'Nota Evaluación': consultor['nota_evaluacion'],
                'Dirección': consultor['direccion'],
                'Teléfono': consultor['telefono'],
            }
            data.append(fila)

        # Crear un libro de trabajo y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe de Consultores"

        # Agregar encabezados
        for col in data[0].keys():
            cell = ws.cell(row=1, column=list(data[0].keys()).index(col) + 1, value=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Agregar datos del DataFrame a la hoja
        for r_idx, row in enumerate(data, 2):
            for c_idx, value in enumerate(row.values(), 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='center')  # Centrar datos

        # Ajustar el ancho de las columnas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        # Crear respuesta de Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"consultores_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Guardar el libro de trabajo en la respuesta  
        wb.save(response)

        return response
    return HttpResponse("Método no permitido", status=405)
        