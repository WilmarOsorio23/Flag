from decimal import Decimal
from django.shortcuts import render
from Modulo.models import Facturacion_Consultores, Linea
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from Modulo.forms import TotalesPorMesFilterForm
from django.db.models import Sum
from collections import defaultdict
from Modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required

def filtrar_datos(form=None):
    facturas = Facturacion_Consultores.objects.all()

    if form and form.is_valid():
        anio = form.cleaned_data.get('Anio')
        lineas = form.cleaned_data.get('LineaId')
        meses = form.cleaned_data.get('Mes')

        if anio:
            facturas = facturas.filter(Anio=anio)
        if lineas:
            # CORRECCIÓN: Usar LineaId directamente en lugar de LineaId__id
            facturas = facturas.filter(LineaId__in=lineas)
        if meses:
            facturas = facturas.filter(Mes__in=[int(m) for m in meses])
    else:
        # Si no hay formulario o no es válido, no filtrar nada (mostrar totales globales)
        anio = None
        lineas = None
        meses = None

    # CORRECCIÓN: Cambiar 'LineaId__Linea' por 'LineaId__nombre_campo_linea'
    # Reemplaza 'Linea' con el nombre real del campo en tu modelo Linea
    facturas_agrupadas = facturas.values('Mes', 'LineaId__Linea').annotate(
        total_factura=Sum('Valor_Fcta_Cliente'),
        total_cobro=Sum('Valor_Cobro'),
        total_diferencia=Sum('Diferencia_Bruta')
    ).order_by('LineaId__Linea', 'Mes')

    # Para devolver líneas y meses usados en la consulta (si filtrados), si no, todos
    if lineas:
        lineas_obj = lineas  # Ya es un queryset de Linea
    else:
        lineas_obj = Linea.objects.all()

    meses_filtrados = None
    if form and form.is_valid():
        meses_filtrados = form.cleaned_data.get('Mes')

    return facturas_agrupadas, lineas_obj, meses_filtrados

@login_required
@verificar_permiso('can_view_informe_facturacion_consultores')
def informe_totales_por_mes(request):
    form = TotalesPorMesFilterForm(request.GET or None)

    meses_completos = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    meses = meses_completos.copy()

    datos = defaultdict(lambda: defaultdict(lambda: {
        'total_factura': 0,
        'total_cobro': 0,
        'total_diferencia': 0,
        'porcentaje_diferencia': 0
    }))
    totales_por_linea = {}
    totales_por_mes = {}
    total_global = {
        'total_factura': 0, 
        'total_cobro': 0, 
        'total_diferencia': 0,
        'porcentaje_diferencia': 0
    }

    facturas, lineas_filtradas, meses_filtrados = filtrar_datos(form)

    if meses_filtrados:
        meses = [par for par in meses_completos if str(par[0]) in meses_filtrados]

    # CORRECCIÓN: Usar lineas_filtradas directamente
    lineas = lineas_filtradas

    for item in facturas:
        linea = item['LineaId__Linea']
        mes = int(item['Mes'])
        factura = item['total_factura'] or 0
        cobro = item['total_cobro'] or 0
        diferencia = item['total_diferencia'] or 0
        
        # Calcular porcentaje (evitando división por cero)
        porcentaje = (diferencia / factura * 100) if factura != 0 else 0

        datos[linea][mes] = {
            'total_factura': factura,
            'total_cobro': cobro,
            'total_diferencia': diferencia,
            'porcentaje_diferencia': porcentaje
        }

        # Totales por línea
        if linea not in totales_por_linea:
            totales_por_linea[linea] = {
                'total_factura': 0, 
                'total_cobro': 0, 
                'total_diferencia': 0,
                'porcentaje_diferencia': 0
            }
        totales_por_linea[linea]['total_factura'] += factura
        totales_por_linea[linea]['total_cobro'] += cobro
        totales_por_linea[linea]['total_diferencia'] += diferencia

        # Totales por mes
        if mes not in totales_por_mes:
            totales_por_mes[mes] = {
                'total_factura': 0, 
                'total_cobro': 0, 
                'total_diferencia': 0,
                'porcentaje_diferencia': 0
            }
        totales_por_mes[mes]['total_factura'] += factura
        totales_por_mes[mes]['total_cobro'] += cobro
        totales_por_mes[mes]['total_diferencia'] += diferencia

        # Total global
        total_global['total_factura'] += factura
        total_global['total_cobro'] += cobro
        total_global['total_diferencia'] += diferencia

    # Calcular porcentajes finales para totales por mes y global
    for mes in totales_por_mes:
        factura = totales_por_mes[mes]['total_factura']
        diferencia = totales_por_mes[mes]['total_diferencia']
        totales_por_mes[mes]['porcentaje_diferencia'] = (diferencia / factura * 100) if factura != 0 else 0

    for linea in totales_por_linea:
        factura = totales_por_linea[linea]['total_factura']
        diferencia = totales_por_linea[linea]['total_diferencia']
        totales_por_linea[linea]['porcentaje_diferencia'] = (diferencia / factura * 100) if factura != 0 else 0

    if total_global['total_factura'] != 0:
        total_global['porcentaje_diferencia'] = (total_global['total_diferencia'] / total_global['total_factura'] * 100)
    else:
        total_global['porcentaje_diferencia'] = 0

    context = {
        'form': form,
        'datos': datos,
        'lineas': lineas,
        'meses': meses,
        'totales_por_linea': totales_por_linea,
        'totales_por_mes': totales_por_mes,
        'total_global': total_global
    }

    return render(request, 'Informes/informes_facturacion_consultores_index.html', context)

@login_required
@verificar_permiso('can_view_informe_facturacion_consultores')
def reporte_excel_totales_por_mes(request):
    form = TotalesPorMesFilterForm(request.GET or None)

    meses_completos = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    if not form.is_valid():
        return HttpResponse("Parámetros inválidos para generar el reporte.")

    facturas, lineas_filtradas, meses_filtrados = filtrar_datos(form)

    meses = meses_completos
    if meses_filtrados:
        meses = [m for m in meses_completos if str(m[0]) in meses_filtrados]

    # Estructuras de datos actualizadas con porcentaje
    datos = defaultdict(lambda: defaultdict(lambda: {
        'total_factura': Decimal('0'),
        'total_cobro': Decimal('0'),
        'total_diferencia': Decimal('0'),
        'porcentaje_diferencia': Decimal('0')
    }))
    
    totales_por_linea = defaultdict(lambda: {
        'total_factura': Decimal('0'),
        'total_cobro': Decimal('0'),
        'total_diferencia': Decimal('0'),
        'porcentaje_diferencia': Decimal('0')
    })
    
    totales_por_mes = defaultdict(lambda: {
        'total_factura': Decimal('0'),
        'total_cobro': Decimal('0'),
        'total_diferencia': Decimal('0'),
        'porcentaje_diferencia': Decimal('0')
    })
    
    total_global = {
        'total_factura': Decimal('0'),
        'total_cobro': Decimal('0'),
        'total_diferencia': Decimal('0'),
        'porcentaje_diferencia': Decimal('0')
    }

    # CORRECCIÓN: Usar lineas_filtradas directamente
    lineas = lineas_filtradas

    for item in facturas:
        linea = item['LineaId__Linea']
        mes = int(item['Mes'])
        factura = Decimal(str(item['total_factura'] or '0'))
        cobro = Decimal(str(item['total_cobro'] or '0'))
        diferencia = Decimal(str(item['total_diferencia'] or '0'))
        porcentaje = (diferencia / factura * 100) if factura != 0 else Decimal('0')

        datos[linea][mes] = {
            'total_factura': factura,
            'total_cobro': cobro,
            'total_diferencia': diferencia,
            'porcentaje_diferencia': porcentaje
        }

        # Actualizar totales
        totales_por_linea[linea]['total_factura'] += factura
        totales_por_linea[linea]['total_cobro'] += cobro
        totales_por_linea[linea]['total_diferencia'] += diferencia

        totales_por_mes[mes]['total_factura'] += factura
        totales_por_mes[mes]['total_cobro'] += cobro
        totales_por_mes[mes]['total_diferencia'] += diferencia

        total_global['total_factura'] += factura
        total_global['total_cobro'] += cobro
        total_global['total_diferencia'] += diferencia

    # Calcular porcentajes finales
    for linea in totales_por_linea:
        factura = totales_por_linea[linea]['total_factura']
        diferencia = totales_por_linea[linea]['total_diferencia']
        totales_por_linea[linea]['porcentaje_diferencia'] = (diferencia / factura * 100) if factura != 0 else Decimal('0')

    for mes in totales_por_mes:
        factura = totales_por_mes[mes]['total_factura']
        diferencia = totales_por_mes[mes]['total_diferencia']
        totales_por_mes[mes]['porcentaje_diferencia'] = (diferencia / factura * 100) if factura != 0 else Decimal('0')

    if total_global['total_factura'] != 0:
        total_global['porcentaje_diferencia'] = (total_global['total_diferencia'] / total_global['total_factura'] * 100)
    else:
        total_global['porcentaje_diferencia'] = Decimal('0')

    # Crear libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Totales por Línea y Mes"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    footer_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alignment_center = Alignment(horizontal="center", vertical="center")

    num_meses = len(meses)

    # Encabezado fila 1
    ws.cell(row=1, column=1, value='Línea')
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    for mes_num, mes_nombre in meses:
        ws.cell(row=1, column=col, value=mes_nombre)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)
        col += 4

    ws.cell(row=1, column=col, value='Total')
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)

    # Encabezado fila 2 (subcolumnas)
    col = 2
    for _ in meses:
        ws.cell(row=2, column=col, value='Factura Cliente')
        ws.cell(row=2, column=col+1, value='Cobro')
        ws.cell(row=2, column=col+2, value='Diferencia')
        ws.cell(row=2, column=col+3, value='% Dif')
        col += 4

    ws.cell(row=2, column=col, value='Factura Cliente')
    ws.cell(row=2, column=col+1, value='Cobro')
    ws.cell(row=2, column=col+2, value='Diferencia')
    ws.cell(row=2, column=col+3, value='% Dif')

    # Aplicar estilo a encabezados
    for row in [1, 2]:
        for col_idx in range(1, 2 + num_meses * 4 + 4):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment_center

    # Datos
    fila_inicio = 3
    for idx_linea, linea in enumerate(lineas):
        fila = fila_inicio + idx_linea
        ws.cell(row=fila, column=1, value=linea.Linea)
        
        col = 2
        for mes_num, _ in meses:
            mes_data = datos[linea.Linea].get(mes_num, {
                'total_factura': Decimal('0'),
                'total_cobro': Decimal('0'),
                'total_diferencia': Decimal('0'),
                'porcentaje_diferencia': Decimal('0')
            })
            
            ws.cell(row=fila, column=col, value=float(mes_data['total_factura']))
            ws.cell(row=fila, column=col+1, value=float(mes_data['total_cobro']))
            ws.cell(row=fila, column=col+2, value=float(mes_data['total_diferencia']))
            percent_cell = ws.cell(row=fila, column=col+3, value=float(mes_data['porcentaje_diferencia']/100))
            percent_cell.number_format = numbers.FORMAT_PERCENTAGE_00
            col += 4
        
        total_linea = totales_por_linea.get(linea.Linea, {
            'total_factura': Decimal('0'),
            'total_cobro': Decimal('0'),
            'total_diferencia': Decimal('0'),
            'porcentaje_diferencia': Decimal('0')
        })
        
        ws.cell(row=fila, column=col, value=float(total_linea['total_factura']))
        ws.cell(row=fila, column=col+1, value=float(total_linea['total_cobro']))
        ws.cell(row=fila, column=col+2, value=float(total_linea['total_diferencia']))
        percent_cell = ws.cell(row=fila, column=col+3, value=float(total_linea['porcentaje_diferencia']/100))
        percent_cell.number_format = numbers.FORMAT_PERCENTAGE_00

    # Fila totales generales
    fila_totales = fila_inicio + len(lineas)
    ws.cell(row=fila_totales, column=1, value='Total General')
    
    col = 2
    for mes_num, _ in meses:
        tot_mes = totales_por_mes.get(mes_num, {
            'total_factura': Decimal('0'),
            'total_cobro': Decimal('0'),
            'total_diferencia': Decimal('0'),
            'porcentaje_diferencia': Decimal('0')
        })
        
        ws.cell(row=fila_totales, column=col, value=float(tot_mes['total_factura']))
        ws.cell(row=fila_totales, column=col+1, value=float(tot_mes['total_cobro']))
        ws.cell(row=fila_totales, column=col+2, value=float(tot_mes['total_diferencia']))
        percent_cell = ws.cell(row=fila_totales, column=col+3, value=float(tot_mes['porcentaje_diferencia']/100))
        percent_cell.number_format = numbers.FORMAT_PERCENTAGE_00
        col += 4

    ws.cell(row=fila_totales, column=col, value=float(total_global['total_factura']))
    ws.cell(row=fila_totales, column=col+1, value=float(total_global['total_cobro']))
    ws.cell(row=fila_totales, column=col+2, value=float(total_global['total_diferencia']))
    percent_cell = ws.cell(row=fila_totales, column=col+3, value=float(total_global['porcentaje_diferencia']/100))
    percent_cell.number_format = numbers.FORMAT_PERCENTAGE_00

    # Estilo fila totales
    for col_idx in range(1, 2 + num_meses * 4 + 4):
        cell = ws.cell(row=fila_totales, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = footer_fill
        cell.border = border
        cell.alignment = alignment_center

    # Formato numérico
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if (cell.column - 2) % 4 != 3 and cell.column != (2 + num_meses * 4 + 3):
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Ajustar ancho columnas
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        adjusted_width = max(max_len + 2, 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"totales_por_linea_mes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response