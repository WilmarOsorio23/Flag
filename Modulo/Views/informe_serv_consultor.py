from decimal import Decimal
from django.shortcuts import render
from Modulo.models import Facturacion_Consultores, Linea
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from Modulo.forms import TotalesPorMesFilterForm
from django.db.models import Sum
from collections import defaultdict
def filtrar_datos(form=None):
    facturas = Facturacion_Consultores.objects.all().select_related(
        'LineaId', 'Documento', 'ClienteId, Documento'
    )

    if form and form.is_valid():
        anio = form.cleaned_data.get('Anio')
        lineas = form.cleaned_data.get('LineaId')
        meses = form.cleaned_data.get('Mes')
        consultor = form.cleaned_data.get('Consultor')

        if anio:
            facturas = facturas.filter(Anio=anio)
        if lineas:
            lineas_ids = [l.id if hasattr(l, 'id') else l for l in lineas]
            facturas = facturas.filter(LineaId__in=lineas_ids)
        if meses:
            facturas = facturas.filter(Mes__in=[int(m) for m in meses])
        if consultor:
            facturas = facturas.filter(Documento__in=consultor)

    facturas_agrupadas = facturas.values(
        'Mes', 
        'LineaId__Linea',
        'Documento__Nombre',  # Nombre del consultor
        'ClienteId__Nombre_Cliente'  # Nombre del cliente
    ).annotate(
        total_factura=Sum('Valor_Fcta_Cliente'),
        total_cobro=Sum('Valor_Cobro'),
        total_diferencia=Sum('Diferencia_Bruta')
    ).order_by('LineaId__Linea', 'Documento__Nombre', 'ClienteId__Nombre_Cliente', 'Mes')

    # Obtener las líneas usadas en la consulta
    lineas_obj = Linea.objects.filter(
        Linea__in=facturas.values_list('LineaId', flat=True).distinct()
    ).distinct()  # Añade .distinct() para asegurar unicidad

    return facturas_agrupadas, lineas_obj, form.cleaned_data.get('Mes') if form and form.is_valid() else None


def informe_totales(request):
    form = TotalesPorMesFilterForm(request.GET or None)
    
    meses_completos = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    facturas, lineas_filtradas, meses_filtrados = filtrar_datos(form)
    
    # Filtrar meses si es necesario
    meses = [par for par in meses_completos if not meses_filtrados or str(par[0]) in meses_filtrados]

    # Estructuras de datos MEJORADAS con porcentaje
    datos = {}
    totales_por_linea = defaultdict(lambda: {'total_factura': 0, 'total_cobro': 0, 'total_diferencia': 0, 'porcentaje_diferencia': 0})
    totales_por_linea_consultor = defaultdict(lambda: defaultdict(lambda: {'total_factura': 0, 'total_cobro': 0, 'total_diferencia': 0, 'porcentaje_diferencia': 0}))
    totales_por_cliente = defaultdict(lambda: {'total_factura': 0, 'total_cobro': 0, 'total_diferencia': 0, 'porcentaje_diferencia': 0})
    totales_por_mes = defaultdict(lambda: {'total_factura': 0, 'total_cobro': 0, 'total_diferencia': 0, 'porcentaje_diferencia': 0})
    totales_por_linea_mes = defaultdict(lambda: defaultdict(lambda: {'total_factura': 0, 'total_cobro': 0, 'total_diferencia': 0, 'porcentaje_diferencia': 0}))
    totales_por_linea_consultor_mes = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'total_factura': 0, 'total_cobro': 0, 'total_diferencia': 0, 'porcentaje_diferencia': 0})))
    
    total_global = {'total_factura': 0, 'total_cobro': 0, 'total_diferencia': 0, 'porcentaje_diferencia': 0}

    for item in facturas:
        linea = item.get('LineaId__Linea', 'Sin línea')
        consultor = item.get('Documento__Nombre', 'Sin consultor')
        cliente = item.get('ClienteId__Nombre_Cliente', 'Sin cliente')
        mes = int(item['Mes']) if item['Mes'] else 0
        
        factura = item['total_factura'] or 0
        cobro = item['total_cobro'] or 0
        diferencia = item['total_diferencia'] or 0
        porcentaje = (diferencia / factura * 100) if factura != 0 else 0

        # Solo procesar si el mes está en los meses filtrados
        if meses_filtrados and str(mes) not in meses_filtrados:
            continue

        # Estructura principal de datos
        if linea not in datos:
            datos[linea] = {}
        if consultor not in datos[linea]:
            datos[linea][consultor] = {}
        if cliente not in datos[linea][consultor]:
            datos[linea][consultor][cliente] = {}
        
        datos[linea][consultor][cliente][mes] = {
            'total_factura': factura,
            'total_cobro': cobro,
            'total_diferencia': diferencia,
            'porcentaje_diferencia': porcentaje
        }

        # Actualizar totales con porcentaje
        def update_totales(totales_dict, factura, cobro, diferencia, *keys):
            for key in keys[:-1]:
                if key not in totales_dict:
                    if len(keys) == 1:  # Es el último nivel
                        totales_dict[key] = {
                            'total_factura': 0, 
                            'total_cobro': 0, 
                            'total_diferencia': 0,
                            'porcentaje_diferencia': 0
                        }
                    else:
                        totales_dict[key] = defaultdict(lambda: {
                            'total_factura': 0, 
                            'total_cobro': 0, 
                            'total_diferencia': 0,
                            'porcentaje_diferencia': 0
                        })
                totales_dict = totales_dict[key]
            
            key = keys[-1]
            if key not in totales_dict:
                totales_dict[key] = {
                    'total_factura': 0, 
                    'total_cobro': 0, 
                    'total_diferencia': 0,
                    'porcentaje_diferencia': 0
                }
            
            totales_dict[key]['total_factura'] += factura
            totales_dict[key]['total_cobro'] += cobro
            totales_dict[key]['total_diferencia'] += diferencia

        # Actualizar todos los totales
        update_totales(totales_por_linea_consultor, factura, cobro, diferencia, linea, consultor)
        update_totales(totales_por_linea_consultor_mes, factura, cobro, diferencia, linea, consultor, mes)
        update_totales(totales_por_linea, factura, cobro, diferencia, linea)
        update_totales(totales_por_cliente, factura, cobro, diferencia, cliente)
        update_totales(totales_por_mes, factura, cobro, diferencia, mes)
        update_totales(totales_por_linea_mes, factura, cobro, diferencia, linea, mes)
        
        total_global['total_factura'] += factura
        total_global['total_cobro'] += cobro
        total_global['total_diferencia'] += diferencia

    # Calcular porcentajes finales
    def calcular_porcentajes(totales_dict):
        if isinstance(totales_dict, defaultdict):
            for key, value in totales_dict.items():
                calcular_porcentajes(value)
        elif isinstance(totales_dict, dict):
            if 'total_factura' in totales_dict and 'total_diferencia' in totales_dict:
                factura = totales_dict['total_factura']
                diferencia = totales_dict['total_diferencia']
                totales_dict['porcentaje_diferencia'] = (diferencia / factura * 100) if factura != 0 else 0

    calcular_porcentajes(totales_por_linea)
    calcular_porcentajes(totales_por_linea_consultor)
    calcular_porcentajes(totales_por_cliente)
    calcular_porcentajes(totales_por_mes)
    calcular_porcentajes(totales_por_linea_mes)
    calcular_porcentajes(totales_por_linea_consultor_mes)
    
    if total_global['total_factura'] != 0:
        total_global['porcentaje_diferencia'] = (total_global['total_diferencia'] / total_global['total_factura'] * 100)
    else:
        total_global['porcentaje_diferencia'] = 0

    # Calcular totales de filas por línea
    lineas_con_totales = {}
    for linea, consultores in datos.items():
        total_filas = 0
        for consultor, clientes in consultores.items():
            total_filas += len(clientes)
        lineas_con_totales[linea] = total_filas

    context = {
        'form': form,
        'datos': datos,
        'lineas': lineas_filtradas,
        'meses': meses,
        'lineas_con_totales': lineas_con_totales,
        'totales_por_linea': totales_por_linea,
        'totales_por_linea_consultor': totales_por_linea_consultor,
        'totales_por_cliente': totales_por_cliente,
        'totales_por_mes': totales_por_mes,
        'totales_por_linea_mes': totales_por_linea_mes,
        'totales_por_linea_consultor_mes': totales_por_linea_consultor_mes,
        'total_global': total_global
    }

    return render(request, 'Informes/informes_serv_consultor_index.html', context) 
def descargar_reporte_excel_totales_por_mes(request):
    form = TotalesPorMesFilterForm(request.GET or None)
    if not form.is_valid():
        return HttpResponse("Parámetros inválidos para generar el reporte.")

    facturas, lineas_filtradas, meses_filtrados = filtrar_datos(form)

    # Definir meses a mostrar
    meses_completos = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    meses = [m for m in meses_completos if not meses_filtrados or str(m[0]) in meses_filtrados]

    # Estructurar datos con Decimal y calcular totales
    datos = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
        'Factura': Decimal('0'),
        'Cobro': Decimal('0'),
        'Diferencia': Decimal('0'),
        'Porcentaje': Decimal('0')  # Nuevo campo para el porcentaje
    })))

    total_global = {
        'Factura': Decimal('0'),
        'Cobro': Decimal('0'),
        'Diferencia': Decimal('0'),
        'Porcentaje': Decimal('0')  # Nuevo campo para el porcentaje
    }

    totales_por_mes = defaultdict(lambda: {
        'Factura': Decimal('0'),
        'Cobro': Decimal('0'),
        'Diferencia': Decimal('0'),
        'Porcentaje': Decimal('0')  # Nuevo campo para el porcentaje
    })

    # Procesar datos convirtiendo a Decimal
    for item in facturas:
        linea = item['LineaId__Linea']
        consultor = item['Documento__Nombre'] or ''
        cliente = item['ClienteId__Nombre_Cliente']
        mes = int(item['Mes'])
        
        factura = Decimal(str(item['total_factura'] or '0'))
        cobro = Decimal(str(item['total_cobro'] or '0'))
        diferencia = Decimal(str(item['total_diferencia'] or '0'))
        porcentaje = (diferencia / factura * 100) if factura != 0 else Decimal('0')
        
        datos[linea][consultor][cliente][mes] = {
            'Factura': factura,
            'Cobro': cobro,
            'Diferencia': diferencia,
            'Porcentaje': porcentaje  # Nuevo campo
        }
        
        # Acumular totales
        total_global['Factura'] += factura
        total_global['Cobro'] += cobro
        total_global['Diferencia'] += diferencia
        
        totales_por_mes[mes]['Factura'] += factura
        totales_por_mes[mes]['Cobro'] += cobro
        totales_por_mes[mes]['Diferencia'] += diferencia

    # Calcular porcentajes finales
    for mes in totales_por_mes:
        factura = totales_por_mes[mes]['Factura']
        diferencia = totales_por_mes[mes]['Diferencia']
        totales_por_mes[mes]['Porcentaje'] = (diferencia / factura * 100) if factura != 0 else Decimal('0')

    if total_global['Factura'] != 0:
        total_global['Porcentaje'] = (total_global['Diferencia'] / total_global['Factura'] * 100)
    else:
        total_global['Porcentaje'] = Decimal('0')

    # Crear libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Consultores"

    # ========== ESTILOS ==========
    header_main_style = {
        'font': Font(bold=True, color="FFFFFF", size=12),
        'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="center", vertical="center")
    }

    header_month_style = {
        'font': Font(bold=True, color="000000", size=10),
        'fill': PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="center", vertical="center")
    }

    header_sub_style = {
        'font': Font(bold=True, color="000000", size=10),
        'fill': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="center", vertical="center")
    }

    data_style = {
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '#,##0.00'
    }

    percent_style = {
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '0.00%'
    }

    total_consultor_style = {
        'font': Font(bold=True, color="000000"),
        'fill': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '#,##0.00'
    }

    total_consultor_percent_style = {
        'font': Font(bold=True, color="000000"),
        'fill': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '0.00%'
    }

    total_linea_style = {
        'font': Font(bold=True, color="000000"),
        'fill': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '#,##0.00'
    }

    total_linea_percent_style = {
        'font': Font(bold=True, color="000000"),
        'fill': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '0.00%'
    }

    total_global_style = {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '#,##0.00'
    }

    total_global_percent_style = {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '0.00%'
    }

    # ========== CONSTRUIR REPORTE ==========
    # Columnas fijas
    columns_fijas = [
        ('Línea', 15),
        ('Consultor', 25),
        ('Cliente', 30)
    ]

    # Crear estructura de columnas para meses
    columnas_meses = []
    for mes_num, mes_nombre in meses:
        # Agregar el mes como grupo principal (ocupará 4 columnas)
        columnas_meses.append((mes_nombre, 4))

    # Columnas de totales
    columnas_totales = [
        ('Total Factura', 15),
        ('Total Cobro', 15),
        ('Total Diferencia', 15),
        ('% Diferencia', 15)  # Nueva columna para el porcentaje
    ]

    # Escribir encabezados principales (fila 1)
    col_num = 1
    for col_title, width in columns_fijas:
        cell = ws.cell(row=1, column=col_num, value=col_title)
        for attr, value in header_main_style.items():
            setattr(cell, attr, value)
        ws.column_dimensions[get_column_letter(col_num)].width = width
        col_num += 1

    # Escribir encabezados de meses (fila 1)
    for mes_nombre, span in columnas_meses:
        cell = ws.cell(row=1, column=col_num, value=mes_nombre)
        for attr, value in header_main_style.items():
            setattr(cell, attr, value)
        # Combinar celdas para el mes
        ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num+span-1)
        col_num += span

    # Escribir encabezados de totales (fila 1)
    for col_title, width in columnas_totales:
        cell = ws.cell(row=1, column=col_num, value=col_title)
        for attr, value in header_main_style.items():
            setattr(cell, attr, value)
        ws.column_dimensions[get_column_letter(col_num)].width = width
        col_num += 1

    # Escribir subencabezados (fila 2)
    # Columnas fijas (dejar vacío ya que ya están en fila 1)
    col_num = 1
    for _ in columns_fijas:
        cell = ws.cell(row=2, column=col_num, value='')
        col_num += 1

    # Subencabezados para cada mes (Factura, Cobro, Diferencia, %)
    for mes_num, mes_nombre in meses:
        for subcol in ['Factura', 'Cobro', 'Diferencia', '% Dif']:
            cell = ws.cell(row=2, column=col_num, value=subcol)
            for attr, value in header_sub_style.items():
                setattr(cell, attr, value)
            ws.column_dimensions[get_column_letter(col_num)].width = 12 if subcol == '% Dif' else 15
            col_num += 1

    # Subencabezados para totales (dejar vacío ya que ya están en fila 1)
    for _ in columnas_totales:
        cell = ws.cell(row=2, column=col_num, value='')
        col_num += 1

    # ------ Datos ------
    row_num = 3  # Empezar en fila 3 porque las filas 1 y 2 son encabezados
    
    for linea, consultores in datos.items():
        linea_start_row = row_num
        total_linea = defaultdict(lambda: {
            'Factura': Decimal('0'),
            'Cobro': Decimal('0'),
            'Diferencia': Decimal('0'),
            'Porcentaje': Decimal('0')
        })
        
        for consultor, clientes in consultores.items():
            consultor_start_row = row_num
            total_consultor = defaultdict(lambda: {
                'Factura': Decimal('0'),
                'Cobro': Decimal('0'),
                'Diferencia': Decimal('0'),
                'Porcentaje': Decimal('0')
            })
            
            for cliente, meses_data in clientes.items():
                # Escribir línea, consultor y cliente (solo en primera fila)
                ws.cell(row=row_num, column=1, value=linea if row_num == linea_start_row else '')
                ws.cell(row=row_num, column=2, value=consultor if row_num == consultor_start_row else '')
                ws.cell(row=row_num, column=3, value=cliente)
                
                col_num = 4
                total_factura = Decimal('0')
                total_cobro = Decimal('0')
                total_diferencia = Decimal('0')
                
                # Datos por mes
                for mes_num, _ in meses:
                    mes_data = meses_data.get(mes_num, {})
                    
                    factura = mes_data.get('Factura', Decimal('0'))
                    cobro = mes_data.get('Cobro', Decimal('0'))
                    diferencia = mes_data.get('Diferencia', Decimal('0'))
                    porcentaje = mes_data.get('Porcentaje', Decimal('0'))
                    
                    ws.cell(row=row_num, column=col_num, value=float(factura))
                    ws.cell(row=row_num, column=col_num+1, value=float(cobro))
                    ws.cell(row=row_num, column=col_num+2, value=float(diferencia))
                    ws.cell(row=row_num, column=col_num+3, value=float(porcentaje/100))  # Convertir a formato decimal para Excel
                    
                    # Aplicar estilo a datos
                    for i in range(3):
                        for attr, value in data_style.items():
                            setattr(ws.cell(row=row_num, column=col_num+i), attr, value)
                    # Estilo para porcentaje
                    for attr, value in percent_style.items():
                        setattr(ws.cell(row=row_num, column=col_num+3), attr, value)
                    
                    total_factura += factura
                    total_cobro += cobro
                    total_diferencia += diferencia
                    
                    # Acumular para totales
                    total_consultor[mes_num]['Factura'] += factura
                    total_consultor[mes_num]['Cobro'] += cobro
                    total_consultor[mes_num]['Diferencia'] += diferencia
                    
                    total_linea[mes_num]['Factura'] += factura
                    total_linea[mes_num]['Cobro'] += cobro
                    total_linea[mes_num]['Diferencia'] += diferencia
                    
                    col_num += 4
                
                # Totales por cliente
                porcentaje_cliente = (total_diferencia / total_factura * 100) if total_factura != 0 else Decimal('0')
                
                ws.cell(row=row_num, column=col_num, value=float(total_factura))
                ws.cell(row=row_num, column=col_num+1, value=float(total_cobro))
                ws.cell(row=row_num, column=col_num+2, value=float(total_diferencia))
                ws.cell(row=row_num, column=col_num+3, value=float(porcentaje_cliente/100))
                
                # Aplicar estilo a totales de cliente
                for i in range(3):
                    for attr, value in data_style.items():
                        setattr(ws.cell(row=row_num, column=col_num+i), attr, value)
                # Estilo para porcentaje
                for attr, value in percent_style.items():
                    setattr(ws.cell(row=row_num, column=col_num+3), attr, value)
                
                row_num += 1
            
            # Total por consultor (si hay más de un cliente)
            if len(clientes) >= 1:
                ws.cell(row=row_num, column=1, value=linea)
                ws.cell(row=row_num, column=2, value=f"Total {consultor}")
                ws.cell(row=row_num, column=3, value="")
                
                col_num = 4
                total_factura_consultor = Decimal('0')
                total_cobro_consultor = Decimal('0')
                total_diferencia_consultor = Decimal('0')
                
                for mes_num, _ in meses:
                    factura = total_consultor[mes_num]['Factura']
                    cobro = total_consultor[mes_num]['Cobro']
                    diferencia = total_consultor[mes_num]['Diferencia']
                    porcentaje = (diferencia / factura * 100) if factura != 0 else Decimal('0')
                    
                    ws.cell(row=row_num, column=col_num, value=float(factura))
                    ws.cell(row=row_num, column=col_num+1, value=float(cobro))
                    ws.cell(row=row_num, column=col_num+2, value=float(diferencia))
                    ws.cell(row=row_num, column=col_num+3, value=float(porcentaje/100))
                    
                    # Aplicar estilo de total consultor
                    for i in range(3):
                        for attr, value in total_consultor_style.items():
                            setattr(ws.cell(row=row_num, column=col_num+i), attr, value)
                    # Estilo para porcentaje
                    for attr, value in total_consultor_percent_style.items():
                        setattr(ws.cell(row=row_num, column=col_num+3), attr, value)
                    
                    total_factura_consultor += factura
                    total_cobro_consultor += cobro
                    total_diferencia_consultor += diferencia
                    
                    col_num += 4
                
                # Total general del consultor
                porcentaje_consultor = (total_diferencia_consultor / total_factura_consultor * 100) if total_factura_consultor != 0 else Decimal('0')
                
                ws.cell(row=row_num, column=col_num, value=float(total_factura_consultor))
                ws.cell(row=row_num, column=col_num+1, value=float(total_cobro_consultor))
                ws.cell(row=row_num, column=col_num+2, value=float(total_diferencia_consultor))
                ws.cell(row=row_num, column=col_num+3, value=float(porcentaje_consultor/100))
                
                for i in range(3):
                    for attr, value in total_consultor_style.items():
                        setattr(ws.cell(row=row_num, column=col_num+i), attr, value)
                # Estilo para porcentaje
                for attr, value in total_consultor_percent_style.items():
                    setattr(ws.cell(row=row_num, column=col_num+3), attr, value)
                
                row_num += 1
        
        # Total por línea (si hay más de un consultor)
        if len(consultores) >= 1:
            ws.cell(row=row_num, column=1, value=f"Total {linea}")
            ws.cell(row=row_num, column=2, value="")
            ws.cell(row=row_num, column=3, value="")
            
            col_num = 4
            total_factura_linea = Decimal('0')
            total_cobro_linea = Decimal('0')
            total_diferencia_linea = Decimal('0')
            
            for mes_num, _ in meses:
                factura = total_linea[mes_num]['Factura']
                cobro = total_linea[mes_num]['Cobro']
                diferencia = total_linea[mes_num]['Diferencia']
                porcentaje = (diferencia / factura * 100) if factura != 0 else Decimal('0')
                
                ws.cell(row=row_num, column=col_num, value=float(factura))
                ws.cell(row=row_num, column=col_num+1, value=float(cobro))
                ws.cell(row=row_num, column=col_num+2, value=float(diferencia))
                ws.cell(row=row_num, column=col_num+3, value=float(porcentaje/100))
                
                # Aplicar estilo de total línea
                for i in range(3):
                    for attr, value in total_linea_style.items():
                        setattr(ws.cell(row=row_num, column=col_num+i), attr, value)
                # Estilo para porcentaje
                for attr, value in total_linea_percent_style.items():
                    setattr(ws.cell(row=row_num, column=col_num+3), attr, value)
                
                total_factura_linea += factura
                total_cobro_linea += cobro
                total_diferencia_linea += diferencia
                
                col_num += 4
            
            # Total general de la línea
            porcentaje_linea = (total_diferencia_linea / total_factura_linea * 100) if total_factura_linea != 0 else Decimal('0')
            
            ws.cell(row=row_num, column=col_num, value=float(total_factura_linea))
            ws.cell(row=row_num, column=col_num+1, value=float(total_cobro_linea))
            ws.cell(row=row_num, column=col_num+2, value=float(total_diferencia_linea))
            ws.cell(row=row_num, column=col_num+3, value=float(porcentaje_linea/100))
            
            for i in range(3):
                for attr, value in total_linea_style.items():
                    setattr(ws.cell(row=row_num, column=col_num+i), attr, value)
            # Estilo para porcentaje
            for attr, value in total_linea_percent_style.items():
                setattr(ws.cell(row=row_num, column=col_num+3), attr, value)
            
            row_num += 1

    # Agregar total global al final del reporte
    if len(datos) > 0:
        ws.cell(row=row_num, column=1, value="TOTAL GLOBAL")
        ws.cell(row=row_num, column=2, value="")
        ws.cell(row=row_num, column=3, value="")
        
        col_num = 4
        
        # Totales por mes
        for mes_num, _ in meses:
            factura = totales_por_mes[mes_num]['Factura']
            cobro = totales_por_mes[mes_num]['Cobro']
            diferencia = totales_por_mes[mes_num]['Diferencia']
            porcentaje = totales_por_mes[mes_num]['Porcentaje']
            
            ws.cell(row=row_num, column=col_num, value=float(factura))
            ws.cell(row=row_num, column=col_num+1, value=float(cobro))
            ws.cell(row=row_num, column=col_num+2, value=float(diferencia))
            ws.cell(row=row_num, column=col_num+3, value=float(porcentaje/100))
            
            # Aplicar estilo
            for i in range(3):
                for attr, value in total_global_style.items():
                    setattr(ws.cell(row=row_num, column=col_num+i), attr, value)
            # Estilo para porcentaje
            for attr, value in total_global_percent_style.items():
                setattr(ws.cell(row=row_num, column=col_num+3), attr, value)
            
            col_num += 4
        
        # Total general global
        ws.cell(row=row_num, column=col_num, value=float(total_global['Factura']))
        ws.cell(row=row_num, column=col_num+1, value=float(total_global['Cobro']))
        ws.cell(row=row_num, column=col_num+2, value=float(total_global['Diferencia']))
        ws.cell(row=row_num, column=col_num+3, value=float(total_global['Porcentaje']/100))
        
        for i in range(3):
            for attr, value in total_global_style.items():
                setattr(ws.cell(row=row_num, column=col_num+i), attr, value)
        # Estilo para porcentaje
        for attr, value in total_global_percent_style.items():
            setattr(ws.cell(row=row_num, column=col_num+3), attr, value)

    # Congelar paneles para ver encabezados al desplazar
    ws.freeze_panes = 'A3'

    # Generar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"reporte_consultores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response