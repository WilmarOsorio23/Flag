from django.shortcuts import render
from datetime import date, datetime
from collections import defaultdict
from Modulo.forms import EmpleadoFilterForm
from Modulo.models import Empleado, Nomina
from django.http import HttpResponse
from django.db.models import Q
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Función para filtrar empleados y nóminas
def filtrar_empleados(form, empleados):
    """
    Aplica filtros a un QuerySet de empleados basado en los datos del formulario.

    :param form: Instancia de EmpleadoFilterForm con datos validados.
    :param empleados: QuerySet de empleados.
    :return: QuerySet filtrado.
    """
    nombre = form.cleaned_data.get('Nombre')
    linea = form.cleaned_data.get('LineaId')
    cargo = form.cleaned_data.get('Cargo')
    certificaciones = form.cleaned_data.get('CertificadoSAP')
    perfil = form.cleaned_data.get('PerfilId')
    activo = form.cleaned_data.get('Activo')

    if nombre:
        empleados = empleados.filter(Nombre=nombre)
    if linea:
        empleados = empleados.filter(LineaId=linea)
    if cargo:
        empleados = empleados.filter(CargoId=cargo)
    if certificaciones:
        empleados = empleados.filter(CertificadoSAP=bool(int(certificaciones)))
        print('Aqui', certificaciones)
    if perfil:
        empleados = empleados.filter(PerfilId=perfil)
    if activo:
        empleados = empleados.filter(Activo=activo == 'True')
    return empleados

# Función para obtener la información del informe
def obtener_informe_empleados(empleados):
    """
    Genera el informe de empleados basado en un queryset dado.

    :param queryset: QuerySet filtrado de empleados.
    :return: Lista de diccionarios con los datos del informe.
    """
    empleado_info = []


    for empleado in empleados: 
        datos_empleado = {
            'TipoDocumento': empleado.TipoDocumento.Nombre, 
            'Documento': empleado.Documento,
            'Nombre': empleado.Nombre, 
            'FechaNacimiento': empleado.FechaNacimiento, 
            'FechaIngreso': empleado.FechaIngreso, 
            'FechaOperacion': empleado.FechaOperacion,
            'Modulo': empleado.ModuloId.Modulo, 
            'Perfil': empleado.PerfilId.Perfil, 
            'Linea': empleado.LineaId.Linea, 
            'CargoId': empleado.CargoId.Cargo, 
            'TituloProfesional': empleado.TituloProfesional, 
            'FechaGrado': empleado.FechaGrado, 
            'Universidad': empleado.Universidad,
            'ProfesionRealizada': empleado.ProfesionRealizada, 
            'AcademiaSAP':  'SI' if empleado.AcademiaSAP else 'NO', 
            'CertificadoSAP': 'SI' if empleado.CertificadoSAP else 'NO',
            'OtrasCertificaciones': 'SI' if empleado.OtrasCertificaciones else 'NO', 
            'Postgrados':  'SI' if empleado.Postgrados else 'NO',             
            'Activo': 'SI' if empleado.Activo else 'NO' ,
            'FechaRetiro': empleado.FechaRetiro,
            'Direccion': empleado.Direccion if empleado.Direccion else '' ,
            'Ciudad': empleado.Ciudad if empleado.Ciudad else '', 
            'Departamento': empleado.Departamento if empleado.Departamento else '', 
            'DireccionAlterna': empleado.DireccionAlterna if empleado.DireccionAlterna else '', 
            'Telefono1': empleado.Telefono1 if empleado.Telefono1 else '', 
            'Telefono2': empleado.Telefono2 if empleado.Telefono2 else '',
        }

        empleado_info.append(datos_empleado)
    return empleado_info

# Vista del informe de empleados
def informe_empleados(request):
    """
    Vista para generar el informe de empleados con filtros opcionales.
    """
    empleados_info = []
    empleados = Empleado.objects.none()
    show_data = False
    busqueda_realizada = False

    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)

        if request.GET:
            busqueda_realizada = True

            if form.is_valid():
                empleados = Empleado.objects.all()

                # Aplicar filtros
                empleados = filtrar_empleados(form, empleados)

                # Obtener la información del informe solo si hay resultados
                if empleados.exists():
                    empleados_info = obtener_informe_empleados(empleados)
                    show_data = True

    else:
        form = EmpleadoFilterForm()

    context = {
        'form': form,
        'empleados_info': empleados_info,
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'mensaje': "No se encontraron resultados para los filtros aplicados." if busqueda_realizada and not show_data else "No se ha realizado ninguna búsqueda aún."
    }

    #Aplicar lógica para cards
    empleados_totales = empleados.count()
    empleados_activos = empleados.filter(Activo=True).count()
    empleados_inactivos = empleados_totales - empleados_activos
    total_empleados = len(empleados_info)
    activos = sum(1 for e in empleados_info if e["Activo"] == "SI")

    # Diccionarios para acumulación por línea
    certificados_sap_por_linea = defaultdict(int)
    otras_certificaciones_por_linea = defaultdict(int)
    postgrados_por_linea = defaultdict(int)

    # Agrupar por línea
    for e in empleados_info:
        linea = e['Linea']
        if e['CertificadoSAP'] == 'SI':
            certificados_sap_por_linea[linea] += 1
        if e['OtrasCertificaciones'] == 'SI':
            otras_certificaciones_por_linea[linea] += 1
        if e['Postgrados'] == 'SI':
            postgrados_por_linea[linea] += 1

    context.update({
        'total_empleados': total_empleados,
        'activos': activos,
        'empleados_totales': empleados_totales,
        'empleados_activos': empleados_activos,
        'empleados_inactivos': empleados_inactivos,

        # Diccionarios originales (para mostrar en cards)
        'certificados_sap_por_linea': dict(certificados_sap_por_linea),
        'otras_certificaciones_por_linea': dict(otras_certificaciones_por_linea),
        'postgrados_por_linea': dict(postgrados_por_linea),

        # Datos para los gráficos en JS
        #'certificados_sap_labels': [k for k in certificados_sap_por_linea.keys()],
        #'certificados_sap_data': [v for v in certificados_sap_por_linea.values()],
        'certificados_sap_labels': list(certificados_sap_por_linea.keys()),
        'certificados_sap_data': list(certificados_sap_por_linea.values()),

        'otras_certificaciones_labels': list(otras_certificaciones_por_linea.keys()),
        'otras_certificaciones_data': list(otras_certificaciones_por_linea.values()),

        'postgrados_labels': list(postgrados_por_linea.keys()),
        'postgrados_data': list(postgrados_por_linea.values()),
    })

    return render(request, 'informes/informes_empleado_index.html', context)

# Funcionalidad para descargar el informe de empleados en Excel
def exportar_empleados_excel(request):
    # Recuperar los mismos datos de filtrado
    meses = "Enero Febrero Marzo Abril Mayo Junio Julio Agosto Septiembre Octubre Noviembre Diciembre".split()
    empleado_info = []  
    
    # Reutilizar la lógica de filtrado de tu vista actual
    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)

        if form.is_valid():
            # Inicializar empleados
            empleados = Empleado.objects.all()

            # Filtrar empleados
            empleados = filtrar_empleados(form, empleados )
            
            # Obtener información de los empleados
            empleado_info = obtener_informe_empleados(empleados)

        if not empleado_info:
            return HttpResponse("No se encontraron datos para exportar.", status=404)
        
        if not form.is_valid(): 
            return HttpResponse("Filtros no válidos. Por favor, revisa los criterios ingresados.", status=400)

    # Preparar datos para Excel
    data = []
    for empleado in empleado_info:
        fila = {
            'Línea': empleado['Linea'],
            'Documento Colaborador': empleado['Documento'],
            'Nombre Colaborador': empleado['Nombre'],
            #'Cargo': empleado['Cargo'],
            'Perfil': empleado['Perfil'],
            'Módulo': empleado['Modulo'],
            'Certificado SAP': empleado['CertificadoSAP'],
            'Fecha Ingreso': empleado['FechaIngreso'],
            #'TipoDocumento': empleado['TipoDocumento.Nombre'], 
            'Documento': empleado['Documento'],
            'Nombre': empleado['Nombre'], 
            'FechaNacimiento': empleado['FechaNacimiento'], 
            'FechaIngreso': empleado['FechaIngreso'], 
            'FechaOperacion': empleado['FechaOperacion'],
            #'Modulo': empleado['ModuloId.Modulo'], 
            #'Perfil': empleado['PerfilId.Perfil'], 
            #'Linea': empleado['LineaId.Linea'], 
            'TituloProfesional': empleado['TituloProfesional'], 
            'FechaGrado': empleado['FechaGrado'], 
            'Universidad': empleado['Universidad'],
            'ProfesionRealizada': empleado['ProfesionRealizada'], 
            'Universidad': empleado['Universidad'], 
            'AcademiaSAP': empleado['AcademiaSAP'],
            'CertificadoSAP': empleado['CertificadoSAP'],
            'OtrasCertificaciones': empleado['OtrasCertificaciones'], 
            'Postgrados': empleado['Postgrados'], 
            #'CargoId': empleado['CargoId.Cargo'], 
            'Activo': empleado['Activo'],
            'FechaRetiro': empleado['FechaRetiro'],
            'Direccion': empleado['Direccion'],
            'Ciudad': empleado['Ciudad'], 
            'Departamento': empleado['Departamento'], 
            'DireccionAlterna': empleado['DireccionAlterna'], 
            'Telefono1': empleado['Telefono1'], 
            'Telefono2': empleado['Telefono2'],
        }
        data.append(fila)

    # Crear un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe de Empleados"

    # Agregar encabezados
    for col in data[0].keys():
        cell = ws.cell(row=1, column=list(data[0].keys()).index(col) + 1, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Agregar datos del DataFrame a la hoja
    for r_idx, row in enumerate(data, 2):
        for c_idx, value in enumerate(row.values(), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')  # Centrar datos

    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Crear respuesta de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"empleados_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Guardar el libro de trabajo en la respuesta  
    wb.save(response)

    return response
