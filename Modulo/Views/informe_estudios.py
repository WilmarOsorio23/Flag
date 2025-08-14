# Informe de Estudios

# Librerías estándar y de terceros
from collections import defaultdict, Counter
from datetime import datetime

# Django
from django.http import HttpResponse
from django.shortcuts import render

# Formularios y modelos del módulo
from Modulo.forms import EstudiosFilterForm
from Modulo.models import Empleado, Empleados_Estudios

# Librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from Modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required

@login_required
@verificar_permiso('can_manage_informe_estudios')

# Función para filtrar Empleado y Empleados_Estudios
def filtrar_empleados_y_estudios(form, empleados, empleados_estudios):
    """
    Aplica filtros del formulario sobre el queryset de empleados.
    Filtros aplicados:
    - Nombre 
    - Línea
    - Cargo
    """
    empleados = empleados.order_by('Documento')
    nombre = form.cleaned_data.get('Nombre')
    linea = form.cleaned_data.get('LineaId')
    cargo = form.cleaned_data.get('Cargo')

    if nombre:
        empleados = empleados.filter(Nombre__icontains=nombre)
    if linea:
        empleados = empleados.filter(LineaId=linea)
    if cargo:
        empleados = empleados.filter(CargoId=cargo)

    documentos_estudios = empleados_estudios.values_list('documentoId', flat=True)
    empleados = empleados.filter(Documento__in=documentos_estudios)

    return empleados, empleados_estudios

# Función para calcular la información de estudio del empleado
def obtener_empleado_estudio(empleados, estudios):
    """
    Agrupa los estudios de cada empleado en una lista de diccionarios
    """
    empleado_estudio_info = []
    estudios_por_documento = defaultdict(list)

    # Organizar los estudios por documentoId para optimizar consultas
    for estudio in estudios:
        estudios_por_documento[estudio.documentoId_id].append(estudio)

    for empleado in empleados:
        # Obtener estudios del empleado
        estudios_empleado = estudios_por_documento.get(empleado.Documento, [])
        
        # Crear el diccionario con los datos del empleado y sus estudios
        datos_empleado = {
            'nombre_linea': empleado.LineaId.Linea,
            'documento_colaborador': empleado.Documento,
            'nombre_colaborador': empleado.Nombre,
            'cargo': empleado.CargoId.Cargo,
            'perfil': empleado.PerfilId.Perfil,
            'modulo': empleado.ModuloId.Modulo,
            'estudios': [{'titulo': estudio.titulo, 'institucion': estudio.institucion, 
                          'fecha_grado': estudio.fecha_Graduacion, 'fecha_inicio': estudio.fecha_Inicio,
                          'fecha_fin': estudio.fecha_Fin} for estudio in estudios_empleado]
        }
        
        empleado_estudio_info.append(datos_empleado)

    return empleado_estudio_info

def empleado_estudio_filtrado(request):
    """
    Vista principal del informe de estudios.
    Permite aplicar filtros, calcular métricas y mostrar los resultados.
    """
    empleado_estudio_info = []  
    show_data = False  
    busqueda_realizada = False

    if request.method == 'GET':
        form = EstudiosFilterForm(request.GET)

        if request.GET:
            busqueda_realizada = True

            if form.is_valid():

                # Inicializar empleados y estudios
                empleados = Empleado.objects.all()
                estudios = Empleados_Estudios.objects.all()

                # Obtener información de los empleados filtrada
                empleados, estudios = filtrar_empleados_y_estudios(form, empleados, estudios)
                
                # Obtener información de los estudios de los empleados
                empleado_estudio_info = obtener_empleado_estudio(empleados, estudios)
                show_data = bool(empleado_estudio_info)  # Mostrar datos si hay resultados

    else: 
        form = EstudiosFilterForm()

    #Aplicar lógica para cards

    # Total de Estudios
    total_estudios = sum(len(e['estudios']) for e in empleado_estudio_info)

    # Estudios por Institución
    instituciones = [estudio['institucion'] for e in empleado_estudio_info for estudio in e['estudios']]
    conteo_instituciones = dict(Counter(instituciones))

    # Estudios por Línea
    lineas = [e['nombre_linea'] for e in empleado_estudio_info if e['nombre_linea']]
    conteo_lineas = dict(Counter(lineas))

    context = {
        'form': form,
        'empleado_estudio_info': empleado_estudio_info,      
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'total_estudios': total_estudios,
        'conteo_instituciones': conteo_instituciones,
        'conteo_lineas': conteo_lineas,
        'lineas_labels': list(conteo_lineas.keys()),
        'lineas_data': list(conteo_lineas.values()),
        'instituciones_labels': list(conteo_instituciones.keys()),
        'instituciones_data': list(conteo_instituciones.values()),
        'mensaje': "No se encontraron resultados para los filtros aplicados." if busqueda_realizada and not show_data else "No se ha realizado ninguna búsqueda aún."
    }

    return render(request, 'informes/informes_estudios_index.html', context)

def exportar_estudio_excel(request):
    """
    Exporta los datos filtrados de estudios de empleados a un archivo Excel.
    """
    empleado_estudio_info = []
    
    if request.method == 'GET':
        form = EstudiosFilterForm(request.GET)

        if form.is_valid():
            # Inicializar empleados y estudios
            empleados = Empleado.objects.all()
            estudios = Empleados_Estudios.objects.all()

            # Obtener información de los empleados filtrada
            empleados, estudios = filtrar_empleados_y_estudios(form, empleados, estudios)
            
            # Obtener información de los estudios de los empleados
            empleado_estudio_info = obtener_empleado_estudio(empleados, estudios)
            
            if not empleado_estudio_info:
                return HttpResponse("No se encontraron datos para exportar.")
              
        else: 
                return HttpResponse("Filtros no válidos.", status=400)

        #Preparar datos para excel 
        data = []
        for empleado in empleado_estudio_info:
            for estudio in empleado['estudios']:    
                fila = {
                    'Nombre Línea': empleado['nombre_linea'],
                    'Documento Colaborador': empleado['documento_colaborador'],
                    'Nombre Colaborador': empleado['nombre_colaborador'],
                    'Cargo': empleado['cargo'],
                    'Perfil': empleado['perfil'],
                    'Módulo': empleado['modulo'],
                    'Titulo': estudio['titulo'],
                    'Institucion': estudio['institucion'],
                    'Fecha Grado': estudio['fecha_grado'],
                    'Fecha Inicio': estudio['fecha_inicio'],
                    'Fecha Fin': estudio['fecha_fin']
                }
                data.append(fila)

        # Crear un libro de trabajo y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe de Estudios"

        # Agregar encabezados
        for col in data[0].keys():
            cell = ws.cell(row=1, column=list(data[0].keys()).index(col) + 1, value=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Agregar datos del DataFrame a la hoja
        for r_idx, row in enumerate(data, 2):
            for c_idx, value in enumerate(row.values(), 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='center')  # Centrar datos

        # Ajustar el ancho de las columnas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        # Crear respuesta de Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"estudio_Empleados_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Guardar el libro de trabajo en la respuesta  
        wb.save(response)

        return response