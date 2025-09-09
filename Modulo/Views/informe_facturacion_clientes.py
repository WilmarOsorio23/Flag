from collections import defaultdict
from decimal import Decimal
from django.shortcuts import render
from Modulo.models import FacturacionClientes, Linea, CentrosCostos, Clientes, Modulo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from Modulo.forms import FacturacionClientesFilterForm
from django.db.models import Sum
from django.db.models import Q
import json
# Importar funciones y modelos necesarios para el cálculo de costos
from Modulo.models import Tiempos_Cliente, Nomina, Tarifa_Consultores, Tarifa_Clientes, IPC, Horas_Habiles
from decimal import Decimal, InvalidOperation
def obtener_horas_habiles(anio: str) -> dict:
    """Obtiene días y horas hábiles por mes para un año específico"""
    registros = Horas_Habiles.objects.filter(Anio=anio)
    return {
        str(r.Mes): {
            'dias': r.Dias_Habiles,
            'horas_mes': r.Dias_Habiles * r.Horas_Laborales,
            'horas_diarias': r.Horas_Laborales
        } for r in registros
    }
def obtener_ipc_anio(anio):
    """Obtiene el IPC de un año (retorna 1 si no existe)"""
    try:
        ipc = IPC.objects.filter(Anio=str(anio)).order_by('-Mes').first()
        return ipc.Indice if ipc else 1
    except Exception:
        return 1
def precargar_datos(anio: str, meses: list, clientes: list, lineas: list):
    """Precarga todos los datos necesarios para optimizar consultas"""
    print(f"PRECARGANDO DATOS - Año: {anio}, Meses: {meses}")
    
    clientes_ids = [c.ClienteId for c in clientes]
    lineas_ids = [l.LineaId for l in lineas]
    
    # Precargar tiempos INCLUYENDO centro de costos
    tiempos = Tiempos_Cliente.objects.filter(
        Anio=anio,
        Mes__in=meses,
        ClienteId__in=clientes_ids,
        LineaId__in=lineas_ids
    ).values('Anio', 'Mes', 'Documento', 'ClienteId', 'LineaId', 'Horas', 'ModuloId', 'centrocostosId')
    
    print(f"TIEMPOS ENCONTRADOS: {len(tiempos)} registros")
    if tiempos:
        print("Ejemplo de tiempo:", tiempos[0])
    
    # Precargar facturación
    facturacion = FacturacionClientes.objects.filter(
        Anio=anio,
        Mes__in=meses
    ).values('ClienteId', 'LineaId', 'Mes', 'HorasFactura', 'DiasFactura', 'MesFactura', 'Valor')
    
    # Precargar nóminas
    nominas = Nomina.objects.all().values('Documento', 'Anio', 'Mes', 'Salario')
    nominas_dict = defaultdict(list)
    for n in nominas:
        nominas_dict[n['Documento']].append((int(n['Anio']), int(n['Mes']), Decimal(str(n['Salario'] or '0'))))
    
    print(f"NÓMINAS ENCONTRADAS: {len(nominas_dict)} documentos")
    
    # Ordenar nominas
    for doc in nominas_dict:
        nominas_dict[doc].sort(key=lambda x: (-x[0], -x[1]))
    
    # Precargar tarifas consultores
    tarifas_consultores = Tarifa_Consultores.objects.all().values(
        'documentoId', 'anio', 'mes', 'valorHora', 'valorDia', 'valorMes', 'moduloId'
    )
    tarifas_consultores_dict = defaultdict(lambda: defaultdict(list))
    for t in tarifas_consultores:
        tarifas_consultores_dict[t['documentoId']][t['moduloId']].append((
            int(t['anio']), 
            int(t['mes']),
            Decimal(str(t['valorHora'] or '0')),
            Decimal(str(t['valorDia'] or '0')),
            Decimal(str(t['valorMes'] or '0'))
        ))
    
    print(f"TARIFAS CONSULTORES: {len(tarifas_consultores_dict)} documentos")
    
    # Ordenar tarifas consultores
    for doc in tarifas_consultores_dict:
        for modulo in tarifas_consultores_dict[doc]:
            tarifas_consultores_dict[doc][modulo].sort(key=lambda x: (-x[0], -x[1]))
    
    # Precargar tarifas clientes
    tarifas_clientes = Tarifa_Clientes.objects.filter(clienteId__in=clientes_ids)
    tarifas_por_cliente = defaultdict(list)
    for tarifa in tarifas_clientes:
        tarifas_por_cliente[tarifa.clienteId_id].append(tarifa)
    
    print(f"TARIFAS CLIENTES: {len(tarifas_por_cliente)} clientes")
    
    return {
        'tiempos_data': list(tiempos),
        'facturacion_data': facturacion,
        'nominas_dict': nominas_dict,
        'tarifas_consultores_dict': tarifas_consultores_dict,
        'tarifas_por_cliente': tarifas_por_cliente
    }
def calcular_costos_por_ceco_modulo_mes(anio: str, meses: list, lineas_ids: list, datos_precargados: dict, horas_habiles: dict):
    """
    Calcula costos agrupados por centro de costos, módulo y mes
    """
    print(f"CALCULANDO COSTOS - Año: {anio}, Meses: {meses}, Líneas: {lineas_ids}")
    print(f"Total tiempos a procesar: {len(datos_precargados['tiempos_data'])}")
    
    # Estructura para incluir mes: {ceco: {modulo: {mes: costo}}}
    costos_por_ceco_modulo_mes = defaultdict(lambda: defaultdict(lambda: defaultdict(Decimal)))
    total_costos_calculados = 0
    tiempos_procesados = 0
    
    for mes in meses:
        if mes not in horas_habiles:
            print(f"Mes {mes} no encontrado en horas hábiles")
            continue
            
        hh_mes = horas_habiles[mes]
        mes_int = int(mes)
        print(f"Procesando mes {mes} - horas hábiles: {hh_mes}")
        
        for tiempo in datos_precargados['tiempos_data']:
            if (str(tiempo['Anio']) != anio or 
                str(tiempo['Mes']) != mes or 
                tiempo['LineaId'] not in lineas_ids):
                continue
            tiempos_procesados += 1
            documento = tiempo['Documento']
            anio_int = int(tiempo['Anio'])
            modulo_id = tiempo['ModuloId']
            ceco_id = tiempo['centrocostosId']
            horas_trabajadas = Decimal(tiempo['Horas'])
            costo_hora = Decimal('0.00')
            
            print(f"Procesando tiempo: doc={documento}, mes={mes}, modulo={modulo_id}, ceco={ceco_id}, horas={horas_trabajadas}")
            
            # Lógica para empleados
            if documento in datos_precargados['nominas_dict']:
                print(f"  Empleado encontrado en nóminas")
                entradas_nomina = datos_precargados['nominas_dict'][documento]
                entrada = next((e for e in entradas_nomina if e[0] <= anio_int and e[1] <= mes_int), None)
                if entrada:
                    salario = entrada[2]
                    ipc = obtener_ipc_anio(anio_int)
                    salario_ajustado = salario * Decimal(str(ipc))
                    horas_mes_total = hh_mes.get('horas_mes', Decimal('1'))
                    costo_hora = salario_ajustado / horas_mes_total if horas_mes_total else Decimal('0')
                    print(f"  Salario: {salario}, IPC: {ipc}, Salario ajustado: {salario_ajustado}, Costo/hora: {costo_hora}")
            
            # Lógica para consultores
            else:
                print(f"  Consultor - buscando tarifas")
                entradas_tarifa = datos_precargados['tarifas_consultores_dict'][documento].get(modulo_id, [])
                entrada = next((e for e in entradas_tarifa if e[0] <= anio_int and e[1] <= mes_int), None)
                if entrada:
                    valorHora, valorDia, valorMes = entrada[2], entrada[3], entrada[4]
                    horas_diarias = hh_mes.get('horas_diarias', Decimal('1'))
                    horas_mes_total = hh_mes.get('horas_mes', Decimal('1'))
                    
                    if valorHora > 0:
                        costo_hora = valorHora
                    elif valorDia > 0:
                        costo_hora = valorDia / horas_diarias
                    elif valorMes > 0:
                        costo_hora = valorMes / horas_mes_total
                    
                    print(f"  Tarifa encontrada: Hora={valorHora}, Dia={valorDia}, Mes={valorMes}, Costo/hora: {costo_hora}")
                else:
                    print(f"  No se encontró tarifa para documento {documento}, módulo {modulo_id}")
            # Acumular costo por centro de costos, módulo y mes
            costo_total = horas_trabajadas * costo_hora
            total_costos_calculados += float(costo_total)
            
            print(f"  Costo calculado: {horas_trabajadas} * {costo_hora} = {costo_total}")
            
            # Obtener código de centro de costos
            try:
                centro_costo = CentrosCostos.objects.get(id=ceco_id)
                codigo_ceco = centro_costo.codigoCeCo
                print(f"  CeCo encontrado: {codigo_ceco}")
            except CentrosCostos.DoesNotExist:
                codigo_ceco = 'SIN_CECO'
                print(f"  CeCo NO encontrado para id {ceco_id}")
            
            # Obtener nombre del módulo
            try:
                modulo_obj = Modulo.objects.get(ModuloId=modulo_id)
                nombre_modulo = modulo_obj.Modulo
                print(f"  Módulo encontrado: {nombre_modulo}")
            except Modulo.DoesNotExist:
                nombre_modulo = 'SIN_MODULO'
                print(f"  Módulo NO encontrado para id {modulo_id}")
            
            # Acumular por mes
            costos_por_ceco_modulo_mes[codigo_ceco][nombre_modulo][mes] += costo_total
            print(f"  Costo acumulado para {codigo_ceco}/{nombre_modulo}/{mes}: {costos_por_ceco_modulo_mes[codigo_ceco][nombre_modulo][mes]}")
    
    print(f"TOTAL TIEMPOS PROCESADOS: {tiempos_procesados}")
    print(f"TOTAL COSTOS CALCULADOS: {total_costos_calculados}")
    print(f"COSTOS POR CECO-MODULO-MES: {dict(costos_por_ceco_modulo_mes)}")
    
    return costos_por_ceco_modulo_mes
def filtrar_datos(form=None):
    facturas = FacturacionClientes.objects.all().select_related('LineaId', 'ModuloId')
    if form and form.is_valid():
        anio = form.cleaned_data.get('Anio')
        lineas = form.cleaned_data.get('LineaId')
        meses = form.cleaned_data.get('Mes')
        ceco = form.cleaned_data.get('Ceco')
        print(f"FILTRO FORM - Año: {anio}, Líneas: {lineas}, Meses: {meses}, CeCo: {ceco}")
        if anio:
            facturas = facturas.filter(Anio=anio)
        if lineas:
            facturas = facturas.filter(LineaId__in=lineas)
        if meses:
            facturas = facturas.filter(Mes__in=[int(m) for m in meses])
        if ceco:
            cecos_filtro = list(ceco) if isinstance(ceco, (list, tuple)) else [ceco]
            cecos_filtro = [str(c).strip() for c in cecos_filtro]
            facturas = facturas.filter(Ceco__in=cecos_filtro)
            if facturas.count() == 0:
                facturas = FacturacionClientes.objects.all()
                facturas = facturas.filter(Ceco__iregex=r'^(%s)$' % '|'.join(cecos_filtro))
    # CALCULAR COSTOS (NUEVO)
    costos_por_ceco_modulo_mes = {}
    if form and form.is_valid() and anio:
        lineas_ids = [l.LineaId for l in (lineas or Linea.objects.all())]
        meses_list = [str(m) for m in (meses or range(1, 13))]
        
        print(f"CALCULANDO COSTOS CON: anio={anio}, lineas_ids={lineas_ids}, meses_list={meses_list}")
        
        # Precargar datos y calcular costos
        horas_habiles = obtener_horas_habiles(anio)
        print(f"HORAS HABILES: {horas_habiles}")
        
        datos_precargados = precargar_datos(
            anio, 
            meses_list, 
            Clientes.objects.all(), 
            lineas or Linea.objects.all()
        )
        
        costos_por_ceco_modulo_mes = calcular_costos_por_ceco_modulo_mes(
            anio, meses_list, lineas_ids, datos_precargados, horas_habiles
        )
    # Obtener todos los códigos de CeCo únicos para buscar sus descripciones
    cecos_codigos = facturas.values_list('Ceco', flat=True).distinct()
    print(f"CECOS ENCONTRADOS EN FACTURAS: {list(cecos_codigos)}")
    
    # Crear un diccionario de {codigoCeCo: descripcionCeCo}
    descripciones_cecos = dict(CentrosCostos.objects.filter(
        codigoCeCo__in=cecos_codigos
    ).values_list('codigoCeCo', 'descripcionCeCo'))
    facturas_agrupadas = facturas.values(
        'ConsecutivoId',
        'Mes', 
        'Ceco',
        'LineaId__Linea',
        'ModuloId__Modulo',
    ).annotate(
        total_valor=Sum('Valor')
    ).order_by('Ceco', 'LineaId__Linea', 'ModuloId__Modulo', 'Mes')
    print(f"FACTURAS AGRUPADAS: {len(facturas_agrupadas)} registros")
    if facturas_agrupadas:
        print("Ejemplo de factura:", facturas_agrupadas[0])
    # Agregar las descripciones a los resultados
    for item in facturas_agrupadas:
        item['descripcion_ceco'] = descripciones_cecos.get(item['Ceco'], '')
    # Obtener las líneas usadas en la consulta
    lineas_obj = Linea.objects.filter(
        Linea__in=facturas.values_list('LineaId', flat=True).distinct()
    ).distinct()
    
    return facturas_agrupadas, lineas_obj, form.cleaned_data.get('Mes') if form and form.is_valid() else None, costos_por_ceco_modulo_mes
def generar_graficos_facturacion(datos):
    """Genera datos para gráficos de facturación"""
    return {}
def convert_to_regular_dict(d):
    """Convierte defaultdict a dict regular para el template"""
    if isinstance(d, defaultdict):
        d = {k: convert_to_regular_dict(v) for k, v in d.items()}
    return d
def informe_facturacion_clientes(request):
    form = FacturacionClientesFilterForm(request.GET or None)
    
    meses_completos = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    print("=" * 50)
    print("INICIANDO INFORME FACTURACIÓN CLIENTES")
    print("=" * 50)
    
    # Ahora recibimos costos_por_ceco_modulo_mes
    facturas, lineas_filtradas, meses_filtrados, costos_por_ceco_modulo_mes = filtrar_datos(form)
    
    print(f"COSTOS POR CECO MODULO MES RECIBIDOS: {dict(costos_por_ceco_modulo_mes)}")
    
    # Filtrar meses si es necesario
    meses = [par for par in meses_completos if not meses_filtrados or str(par[0]) in meses_filtrados]
    # Estructuras de datos
    datos = {}
    totales_por_linea = defaultdict(float)
    totales_por_ceco = defaultdict(float)
    totales_por_modulo = defaultdict(float)
    totales_por_mes = defaultdict(float)
    totales_por_linea_mes = defaultdict(lambda: defaultdict(float))
    totales_por_ceco_linea = defaultdict(lambda: defaultdict(float))
    totales_por_ceco_linea_mes = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    totales_por_ceco_linea_modulo = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    totales_por_ceco_modulo = defaultdict(lambda: defaultdict(float))
    
    # NUEVAS ESTRUCTURAS PARA COSTOS
    costos_por_ceco = defaultdict(float)
    costos_por_modulo = defaultdict(float)
    costos_por_ceco_modulo = defaultdict(lambda: defaultdict(float))
    costos_por_mes = defaultdict(float)
    costo_global = 0.0
    # NUEVAS ESTRUCTURAS PARA COSTOS POR LÍNEA
    costos_por_ceco_linea_mes = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    costos_por_ceco_linea = defaultdict(lambda: defaultdict(float))
    
    total_global = 0.0
    # PRIMERO: Agrupar facturas por ceco + modulo + mes para evitar duplicados
    facturas_agrupadas_por_ceco_modulo_mes = defaultdict(float)
    
    for item in facturas:
        ceco = item.get('Ceco', 'Sin Ceco')
        modulo = item.get('ModuloId__Modulo', 'Sin módulo')
        mes = int(item['Mes']) if item['Mes'] else 0
        mes_str = str(mes)
        
        valor = float(item['total_valor'] or 0)
        # Solo procesar si el mes está en los meses filtrados
        if meses_filtrados and str(mes) not in meses_filtrados:
            continue
        # Agrupar facturas por ceco + modulo + mes
        key = f"{ceco}_{modulo}_{mes_str}"
        facturas_agrupadas_por_ceco_modulo_mes[key] += valor
    # SEGUNDO: Procesar las facturas agrupadas
    for key, valor in facturas_agrupadas_por_ceco_modulo_mes.items():
        ceco, modulo, mes_str = key.split('_')
        mes = int(mes_str)
        
        # Obtener descripción y línea de la primera factura que coincida
        primera_factura = next((item for item in facturas 
                               if item.get('Ceco') == ceco 
                               and item.get('ModuloId__Modulo') == modulo 
                               and int(item.get('Mes', 0)) == mes), None)
        
        if not primera_factura:
            continue
            
        descripcion = primera_factura.get('descripcion_ceco', '')
        linea = primera_factura.get('LineaId__Linea', 'Sin línea')
        # Estructura principal de datos
        if ceco not in datos:
            datos[ceco] = {}
        if descripcion not in datos[ceco]:
            datos[ceco][descripcion] = {}
        if linea not in datos[ceco][descripcion]:
            datos[ceco][descripcion][linea] = {}
        if modulo not in datos[ceco][descripcion][linea]:
            datos[ceco][descripcion][linea][modulo] = {}
        
        # Inicializar si no existe
        if mes not in datos[ceco][descripcion][linea][modulo]:
            datos[ceco][descripcion][linea][modulo][mes] = {'total_valor': 0.0, 'costo': 0.0}
        
        # Sumar el valor (ya está agrupado)
        datos[ceco][descripcion][linea][modulo][mes]['total_valor'] += valor
        # AGREGAR COSTOS - Solo una vez por ceco + modulo + mes
        costo_mes = float(costos_por_ceco_modulo_mes.get(ceco, {}).get(modulo, {}).get(mes_str, Decimal('0.00')))
        datos[ceco][descripcion][linea][modulo][mes]['costo'] = costo_mes  # = en lugar de +=
        # ACUMULAR COSTOS POR LÍNEA Y MES (NUEVO)
        costos_por_ceco_linea_mes[ceco][linea][mes] += costo_mes
        costos_por_ceco_linea[ceco][linea] += costo_mes
        print(f"PROCESANDO: ceco={ceco}, modulo={modulo}, mes={mes}, facturado={valor}, costo={costo_mes}")
        # Actualizar totales
        totales_por_linea[linea] += valor
        totales_por_ceco[ceco] += valor
        totales_por_modulo[modulo] += valor
        totales_por_mes[mes] += valor
        totales_por_linea_mes[linea][mes] += valor
        totales_por_ceco_linea_modulo[ceco][linea][modulo] += valor
        totales_por_ceco_modulo[ceco][modulo] += valor
        totales_por_ceco_linea[ceco][linea] += valor
        totales_por_ceco_linea_mes[ceco][linea][mes] += valor
        
        total_global += valor
        # Actualizar totales de costos
        costos_por_ceco[ceco] += costo_mes
        costos_por_modulo[modulo] += costo_mes
        costos_por_ceco_modulo[ceco][modulo] += costo_mes
        costos_por_mes[mes] += costo_mes
        costo_global += costo_mes
    print(f"TOTAL GLOBAL FACTURADO: {total_global}")
    print(f"TOTAL GLOBAL COSTO: {costo_global}")
    print(f"MARGEN GLOBAL: {total_global - costo_global}")
    # Calcular totales de filas por ceco
    cecos_con_totales = {}
    for ceco, descripciones in datos.items():
        total_filas = 0
        for descripcion, lineas in descripciones.items():
            for linea, modulos in lineas.items():
                total_filas += len(modulos)
        cecos_con_totales[ceco] = total_filas
    graficos = generar_graficos_facturacion(datos)
    context = {
        'form': form,
        'datos': datos,
        'lineas': lineas_filtradas,
        'meses': meses,
        'cecos_con_totales': cecos_con_totales,
        'totales_por_linea': convert_to_regular_dict(totales_por_linea),
        'totales_por_ceco': convert_to_regular_dict(totales_por_ceco),
        'totales_por_modulo': convert_to_regular_dict(totales_por_modulo),
        'totales_por_mes': convert_to_regular_dict(totales_por_mes),
        'totales_por_linea_mes': convert_to_regular_dict(totales_por_linea_mes),
        'graficos': graficos,
        'totales_por_ceco_linea': convert_to_regular_dict(totales_por_ceco_linea),
        'totales_por_ceco_linea_mes': convert_to_regular_dict(totales_por_ceco_linea_mes),
        'totales_por_ceco_linea_modulo': convert_to_regular_dict(totales_por_ceco_linea_modulo),
        'totales_por_ceco_modulo': convert_to_regular_dict(totales_por_ceco_modulo),
        'total_global': total_global,
        
        # NUEVOS DATOS DE COSTOS
        'costos_por_ceco': convert_to_regular_dict(costos_por_ceco),
        'costos_por_modulo': convert_to_regular_dict(costos_por_modulo),
        'costos_por_ceco_modulo': convert_to_regular_dict(costos_por_ceco_modulo),
        'costos_por_mes': convert_to_regular_dict(costos_por_mes),
        'costo_global': costo_global,
        'margen_global': total_global - costo_global,
        # NUEVOS: Costos por línea
        'costos_por_ceco_linea': convert_to_regular_dict(costos_por_ceco_linea),
        'costos_por_ceco_linea_mes': convert_to_regular_dict(costos_por_ceco_linea_mes),
    }
    print("=" * 50)
    print("FINALIZANDO INFORME FACTURACIÓN CLIENTES")
    print("=" * 50)
    
    return render(request, 'Informes/informes_facturacion_clientes_index.html', context)


def descargar_reporte_excel_facturacion_clientes(request):
    # 1. Limpieza de parámetros GET
    get_params = request.GET.copy()
    
    # Eliminar parámetros vacíos o None
    for key in list(get_params.keys()):
        if get_params[key] in ['None', '', None, 'Todos']:
            del get_params[key]
    
    # 2. Manejo del formulario
    if not get_params:
        # Caso sin filtros - formulario sin validar
        form = FacturacionClientesFilterForm()
    else:
        # Caso con filtros - validar formulario
        form = FacturacionClientesFilterForm(get_params)
        if not form.is_valid():
            return HttpResponse("Parámetros inválidos para generar el reporte.")

    # 3. Definición de meses
    meses_completos = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    # 4. Obtención de datos
    try:
        facturas, lineas_filtradas, meses_filtrados = filtrar_datos(form)
    except Exception as e:
        return HttpResponse(f"Error al obtener datos: {str(e)}")

    if not facturas:
        return HttpResponse("No hay datos disponibles para los filtros seleccionados.")

    # 5. Determinar meses a mostrar
    meses = meses_completos
    if meses_filtrados:  # Solo si hay meses filtrados específicamente
        meses = [par for par in meses_completos if str(par[0]) in meses_filtrados]

    # 6. Procesamiento de datos
    datos = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict))))
    totales_por_mes = defaultdict(float)
    total_global = 0.0

    for item in facturas:
        ceco = item.get('Ceco', 'Sin Ceco')
        descripcion = item.get('descripcion_ceco', '')
        linea = item.get('LineaId__Linea', 'Sin línea')
        modulo = item.get('ModuloId__Modulo', 'Sin módulo')
        
        try:
            mes = int(item['Mes']) if item['Mes'] else 0
        except (ValueError, TypeError):
            mes = 0
            
        valor = float(item['total_valor'] or 0)

        # Solo filtrar por mes si hay meses_filtrados definidos
        if meses_filtrados and str(mes) not in meses_filtrados:
            continue

        if mes not in datos[ceco][descripcion][linea][modulo]:
            datos[ceco][descripcion][linea][modulo][mes] = 0.0
        
        datos[ceco][descripcion][linea][modulo][mes] += valor
        totales_por_mes[mes] += valor
        total_global += valor

    # 7. Creación del libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturación Clientes"

    # ========== ESTILOS ==========
    header_style = {
        'font': Font(bold=True, color="FFFFFF", size=12),
        'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="center", vertical="center")
    }

    data_style = {
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '#,##0.00'
    }

    total_style = {
        'font': Font(bold=True, color="000000"),
        'fill': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '#,##0.00'
    }

    global_total_style = {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'number_format': '#,##0.00'
    }

    # ========== ESTRUCTURA DEL REPORTE ==========
    # Columnas fijas
    columns_fijas = [
        ('Ceco', 15),
        ('Descripción', 30),
        ('Línea', 20),
        ('Módulo', 20)
    ]

    # Columnas de meses
    columnas_meses = [(mes_nombre, 15) for mes_num, mes_nombre in meses]

    # Escribir encabezados
    col_num = 1
    for col_title, width in columns_fijas:
        cell = ws.cell(row=1, column=col_num, value=col_title)
        for attr, value in header_style.items():
            setattr(cell, attr, value)
        ws.column_dimensions[get_column_letter(col_num)].width = width
        col_num += 1

    for mes_nombre, width in columnas_meses:
        cell = ws.cell(row=1, column=col_num, value=mes_nombre)
        for attr, value in header_style.items():
            setattr(cell, attr, value)
        ws.column_dimensions[get_column_letter(col_num)].width = width
        col_num += 1

    # Columna de total
    cell = ws.cell(row=1, column=col_num, value="Total")
    for attr, value in header_style.items():
        setattr(cell, attr, value)
    ws.column_dimensions[get_column_letter(col_num)].width = 15

    # ------ Escribir datos ------
    row_num = 2
    
    for ceco, descripciones in datos.items():
        ceco_start_row = row_num
        
        for descripcion, lineas in descripciones.items():
            desc_start_row = row_num
            
            for linea, modulos in lineas.items():
                linea_start_row = row_num
                
                for modulo, meses_data in modulos.items():
                    # Escribir datos fijos
                    ws.cell(row=row_num, column=1, value=ceco if row_num == ceco_start_row else '')
                    ws.cell(row=row_num, column=2, value=descripcion if row_num == desc_start_row else '')
                    ws.cell(row=row_num, column=3, value=linea if row_num == linea_start_row else '')
                    ws.cell(row=row_num, column=4, value=modulo)
                    
                    # Aplicar estilo a las celdas fijas
                    for i in range(1, 5):
                        cell = ws.cell(row=row_num, column=i)
                        for attr, value in data_style.items():
                            if attr != 'number_format':
                                setattr(cell, attr, value)
                    
                    # Escribir valores por mes
                    col_num = 5
                    total_fila = 0.0
                    
                    for mes_num, mes_nombre in meses:
                        valor = meses_data.get(mes_num, 0.0)
                        cell = ws.cell(row=row_num, column=col_num, value=valor)
                        for attr, value in data_style.items():
                            setattr(cell, attr, value)
                        
                        total_fila += valor
                        col_num += 1
                    
                    # Escribir total de la fila
                    cell = ws.cell(row=row_num, column=col_num, value=total_fila)
                    for attr, value in data_style.items():
                        setattr(cell, attr, value)
                    
                    row_num += 1
                
                # Total por grupo (mostrar siempre, incluso si hay un solo módulo)
                ws.cell(row=row_num, column=3, value=f"Total {descripcion}")
                
                col_num = 5
                total_linea = 0.0
                
                for mes_num, mes_nombre in meses:
                    total_mes = sum(modulo_data.get(mes_num, 0.0) for modulo_data in modulos.values())
                    cell = ws.cell(row=row_num, column=col_num, value=total_mes)
                    for attr, value in total_style.items():
                        setattr(cell, attr, value)
                    
                    total_linea += total_mes
                    col_num += 1
                
                cell = ws.cell(row=row_num, column=col_num, value=total_linea)
                for attr, value in total_style.items():
                    setattr(cell, attr, value)
                
                row_num += 1
            
            # Total por descripción si hay más de una línea
            if len(lineas) > 1:
                ws.cell(row=row_num, column=2, value=f"Total {descripcion}")
                
                col_num = 5
                total_desc = 0.0
                
                for mes_num, mes_nombre in meses:
                    total_mes = sum(
                        sum(modulo_data.get(mes_num, 0.0) for modulo_data in lineas[linea].values())
                        for linea in lineas
                    )
                    cell = ws.cell(row=row_num, column=col_num, value=total_mes)
                    for attr, value in total_style.items():
                        setattr(cell, attr, value)
                    
                    total_desc += total_mes
                    col_num += 1
                
                cell = ws.cell(row=row_num, column=col_num, value=total_desc)
                for attr, value in total_style.items():
                    setattr(cell, attr, value)
                
                row_num += 1
        
        # Total por ceco si hay más de una descripción
        if len(descripciones) > 1:
            ws.cell(row=row_num, column=1, value=f"Total {ceco}")
            
            col_num = 5
            total_ceco = 0.0
            
            for mes_num, mes_nombre in meses:
                total_mes = sum(
                    sum(
                        sum(modulo_data.get(mes_num, 0.0) for modulo_data in lineas[linea].values())
                        for linea in lineas
                    )
                    for lineas in descripciones.values()
                )
                cell = ws.cell(row=row_num, column=col_num, value=total_mes)
                for attr, value in total_style.items():
                    setattr(cell, attr, value)
                
                total_ceco += total_mes
                col_num += 1
            
            cell = ws.cell(row=row_num, column=col_num, value=total_ceco)
            for attr, value in total_style.items():
                setattr(cell, attr, value)
            
            row_num += 1

    # Total global
    if datos:
        ws.cell(row=row_num, column=1, value="TOTAL GLOBAL")
        
        col_num = 5
        for mes_num, mes_nombre in meses:
            cell = ws.cell(row=row_num, column=col_num, value=totales_por_mes.get(mes_num, 0.0))
            for attr, value in global_total_style.items():
                setattr(cell, attr, value)
            col_num += 1
        
        cell = ws.cell(row=row_num, column=col_num, value=total_global)
        for attr, value in global_total_style.items():
            setattr(cell, attr, value)

    # Congelar paneles
    ws.freeze_panes = 'A2'

    # Generar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"facturacion_clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def generar_graficos_facturacion(datos):
    """Genera SOLO 2 gráficos de barras: totales por CeCo y totales por módulo"""
    graficos = []
    
    # --- 1. Gráfico de TOTALES POR CECO ---
    cecos_labels = []
    cecos_data = []
    
    for ceco, descripciones in datos.items():
        descripcion = next(iter(descripciones.keys())) if descripciones else ""
        total_ceco = 0.0
        
        for linea, modulos in descripciones[descripcion].items():
            for modulo, meses_data in modulos.items():
                for mes, valores in meses_data.items():
                    total_ceco += float(valores.get('total_valor', 0))
        
        if total_ceco > 0:
            cecos_labels.append(f"{ceco} - {descripcion}")
            cecos_data.append(total_ceco)
    
    if cecos_data:
        graficos.append({
            'tipo': 'cecos',
            'nombre': "Totales por Centro de Costos",
            'config': {
                'tipo_grafico': 'bar',
                'labels': cecos_labels,
                'datasets': [{
                    'label': 'Total Facturado',
                    'data': cecos_data,
                    'backgroundColor': 'rgba(54, 162, 235, 0.7)',
                    'borderColor': 'rgba(54, 162, 235, 1)',
                    'borderWidth': 1
                }]
            }
        })
    
    # --- 2. Gráfico de TOTALES POR MÓDULO CON COLORES POR DESCRIPCIÓN ---
    modulos_data = []
    modulos_labels = []
    modulos_descripciones = []
    descripciones_colores = {}
    colores_disponibles = [
        'rgba(255, 99, 132, 0.7)',  # Rojo
        'rgba(54, 162, 235, 0.7)',   # Azul
        'rgba(255, 206, 86, 0.7)',   # Amarillo
        'rgba(75, 192, 192, 0.7)',   # Verde
        'rgba(153, 102, 255, 0.7)',  # Morado
        'rgba(255, 159, 64, 0.7)',   # Naranja
        'rgba(199, 199, 199, 0.7)',  # Gris
    ]
    
    # Primero recolectamos todas las descripciones únicas
    descripciones_unicas = set()
    for ceco, descripciones in datos.items():
        descripcion = next(iter(descripciones.keys())) if descripciones else "Sin descripción"
        descripciones_unicas.add(descripcion)
    
    # Asignamos colores a cada descripción
    for i, descripcion in enumerate(descripciones_unicas):
        descripciones_colores[descripcion] = colores_disponibles[i % len(colores_disponibles)]
    
    # Recolectamos datos de módulos
    background_colors = []
    for ceco, descripciones in datos.items():
        descripcion = next(iter(descripciones.keys())) if descripciones else "Sin descripción"

        if descripcion not in descripciones_colores:
            # Asignar color si no existe
            idx = len(descripciones_colores) % len(colores_disponibles)
            descripciones_colores[descripcion] = colores_disponibles[idx]
        
        for linea, modulos in descripciones.get(descripcion, {}).items():
            for modulo, meses_data in modulos.items():
                total = sum(float(mes_data.get('total_valor', 0)) for mes_data in meses_data.values())
                if total > 0:
                    modulos_labels.append(modulo)
                    modulos_data.append(total)
                    modulos_descripciones.append(descripcion)
                    background_colors.append(descripciones_colores[descripcion])
    
    if modulos_data:
        # Crear un dataset por cada descripción para la leyenda
        datasets = [{
            'label': 'Total Facturado',  # Esto aparecerá en el tooltip
            'data': modulos_data,
            'backgroundColor': background_colors,
            'borderColor': [color.replace('0.7', '1') for color in background_colors],
            'borderWidth': 1,
            'descripciones': modulos_descripciones
        }]
        
        # Agregar datasets ficticios solo para la leyenda
        legend_datasets = []
        for descripcion, color in descripciones_colores.items():
            legend_datasets.append({
                'label': descripcion,
                'data': [],
                'backgroundColor': color,
                'borderColor': color.replace('0.7', '1'),
                'borderWidth': 1
            })
        
        graficos.append({
            'tipo': 'modulos',
            'nombre': "Totales por Módulo (Colores por Descripción)",
            'config': {
                'tipo_grafico': 'bar',
                'labels': modulos_labels,
                'descripciones_colores': descripciones_colores,
                'datasets': datasets,
                'legend_datasets': legend_datasets  # Datasets para la leyenda
            }
        })
    
    return graficos