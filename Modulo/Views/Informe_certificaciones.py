# Informe de Certificación de Empleados
from datetime import datetime
from msilib.schema import Font
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from Modulo.forms import EmpleadoFilterForm
from Modulo.models import Certificacion, Detalle_Certificacion, Empleado
from django.views.decorators.csrf import csrf_exempt

def filtrar_empleado(form, empleados, certificaciones, detalles):
    linea = form.cleaned_data.get('LineaId')
    certificacion = form.cleaned_data.get('Certificacion')
    nombre = form.cleaned_data.get('Nombre')

    if nombre:
        empleados = empleados.filter(Nombre__icontains=nombre)
    if linea:
        empleados = empleados.filter(LineaId=linea)
    if certificacion:
        detalles = detalles.filter(CertificacionId=certificacion)

    empleados_con_certificacion = empleados.filter(
        Documento__in=detalles.values_list('Documento__Documento', flat=True)
    )
    
    return empleados_con_certificacion, certificaciones, detalles

def obtener_informe_certificacion(certificaciones, empleados, detalles):
    certificaciones_info = []
    detalles = detalles.select_related('Documento', 'CertificacionId')
    empleados_dict = {empleado.Documento: empleado for empleado in empleados}

    for detalle in detalles:
        empleado = empleados_dict.get(detalle.Documento.Documento)
        if empleado:
            certificaciones_info.append({
                'Documento': empleado.Documento,
                'Nombre': empleado.Nombre,
                'Linea': empleado.LineaId.Linea if empleado.LineaId else '',
                'Cargo': empleado.CargoId.Cargo if empleado.CargoId else '',
                'Perfil': empleado.PerfilId.Perfil if empleado.PerfilId else '',
                'Modulo': empleado.ModuloId.Modulo if empleado.ModuloId else '',
                'CertificacionId': detalle.CertificacionId.CertificacionId,
                'Certificacion': detalle.CertificacionId.Certificacion,
                'Fecha_Certificacion': detalle.Fecha_Certificacion,
            })
    return certificaciones_info

def empleado_filtrado(request):
    empleados = Empleado.objects.all()
    certificaciones = Certificacion.objects.all()
    detalles = Detalle_Certificacion.objects.all()
    certificaciones_info = []
    show_data = False

    if request.method == 'GET':
        form = EmpleadoFilterForm(request.GET)
        if form.is_valid():
            empleados, certificaciones, detalles = filtrar_empleado(form, empleados, certificaciones, detalles)
            certificaciones_info = obtener_informe_certificacion(certificaciones, empleados, detalles)
            show_data = bool(certificaciones_info)
    else:
        form = EmpleadoFilterForm()

    return render(request, 'informes/informes_certificacion_index.html', {
        'form': form,
        'certificaciones_info': certificaciones_info,
        'show_data': show_data,
        'mensaje': "No se encontraron resultados para los filtros aplicados"
    })

@csrf_exempt
def exportar_certificaciones_excel(request):
    """
    Exporta el informe de certificaciones en formato Excel.
    """
    print('LLEGO AQUÍ')
    certificaciones_info = []  
    

    if request.method == 'GET':
        print('LLEGO AQUI 2')
        form = EmpleadoFilterForm(request.GET)
        print('LLEGO AQUÍ 3')
        
        if form.is_valid():
            certificaciones = Certificacion.objects.all()
            empleados = Empleado.objects.all()
            detalles = Detalle_Certificacion.objects.all()

            print('LLEGO AQUÍ', empleados)
            print('LLEGO AQUI 2',certificaciones)
            print('LLEGO AQUI 3', detalles)
            # Aplicar filtros
            certificaciones = filtrar_empleado(form, certificaciones)
            empleados = filtrar_empleado(form, empleados)
            detalles = filtrar_empleado(form, detalles)

            # Obtener los datos del informe
            certificaciones_info = obtener_informe_certificacion(certificaciones, empleados, detalles)

        if not certificaciones_info:
            return HttpResponse("No se encontraron datos para exportar.", status=404)

        if not form.is_valid():
            return HttpResponse("Filtros no válidos. Por favor, revisa los criterios ingresados.", status=400)

    # Preparar los datos en formato de lista de diccionarios
    data = []
    for cert in certificaciones_info:
        fila = {
            'Nombre Colaborador': cert['Nombre'],
            'Documento': cert['Documento'],
            'Módulo': cert['Modulo'],
            'Perfil': cert['Perfil'],
            'Línea': cert['Linea'],
            'Cargo': cert['Cargo'],
            'ID Certificación': cert['CertificacionId'],
            'Certificación': cert['Certificacion'],
            'Fecha Certificación': cert['Fecha_Certificacion'],
        }
        data.append(fila)

    # Crear el archivo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificaciones"

    # Agregar encabezados
    for col in data[0].keys():
        cell = ws.cell(row=1, column=list(data[0].keys()).index(col) + 1, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Agregar datos al Excel
    for r_idx, row in enumerate(data, 2):
        for c_idx, value in enumerate(row.values(), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')

    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    # Agregar bordes
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Crear respuesta para descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"certificaciones_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Guardar el archivo en la respuesta
    wb.save(response)

    return response