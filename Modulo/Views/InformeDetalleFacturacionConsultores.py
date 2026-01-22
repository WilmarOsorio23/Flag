from django import forms
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Q
from modulo.models import Facturacion_Consultores, Empleado, Linea
from modulo.forms import FacturacionConsultoresDetalleFilterForm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from modulo.decorators import verificar_permiso
from django.contrib.auth.decorators import login_required

@login_required
@verificar_permiso('can_view_informe_detalle_facturacion_consultores')
def filtrar_facturacion_consultores(request):
    # Obtener los filtros de la URL (GET)
    anio = request.GET.get('Anio')
    meses = request.GET.getlist('Mes')
    consultores = request.GET.getlist('Consultor')
    lineas = request.GET.getlist('LineaId')

    # Construir la consulta con los filtros aplicados
    filtro = Q()
    if anio:
        filtro &= Q(Anio=anio)
    if meses:
        filtro &= Q(Mes__in=meses)
    if consultores:
        filtro &= Q(Documento_id__in=consultores)
    if lineas:
        filtro &= Q(LineaId_id__in=lineas)

    # Obtener los datos de facturación con los filtros aplicados
    facturacion = Facturacion_Consultores.objects.filter(filtro).select_related('Documento', 'ClienteId', 'ModuloId', 'LineaId')
    return facturacion


# Función para filtrar informe de facturación de consultores
def filtrar_facturacion_consultores_form(form, facturacion):
    anio = form.cleaned_data.get('Anio')
    meses = form.cleaned_data.get('Mes')
    consultores = form.cleaned_data.get('Consultor')
    lineas = form.cleaned_data.get('LineaId')

    if anio:
        facturacion = facturacion.filter(Anio=anio)
    if meses:
        facturacion = facturacion.filter(Mes__in=meses)
    if consultores:
        facturacion = facturacion.filter(Documento_id__in=consultores)
    if lineas:
        facturacion = facturacion.filter(LineaId_id__in=lineas)

    return facturacion

@login_required
@verificar_permiso('can_view_informe_detalle_facturacion_consultores')
# Vista para mostrar el informe de facturación de consultores
def informe_detalle_facturacion_consultores(request):
    resultados = []
    show_data = False  
    busqueda_realizada = False

    if request.method == 'GET':
        form = FacturacionConsultoresDetalleFilterForm(request.GET)
        
        if request.GET:
            busqueda_realizada = True
        
            if form.is_valid():
                # Primero obtener el queryset base
                facturacion = Facturacion_Consultores.objects.all()
                
                # Luego aplicar los filtros (pasando ambos argumentos)
                facturacion = filtrar_facturacion_consultores_form(form, facturacion)
                
                # Hacer el select_related después de filtrar
                facturacion = facturacion.select_related('Documento', 'ClienteId', 'ModuloId', 'LineaId')
                
                # Obtener directamente los valores necesarios
                resultados = list(facturacion.values(
                    'Anio', 'Mes', 'Empresa', 'Cta_Cobro', 'Periodo_Cobrado',
                    'Aprobado_Por', 'Fecha_Cobro', 'Fecha_Pago', 'Horas',
                    'Valor_Unitario', 'Valor_Cobro', 'IVA', 'Valor_Neto',
                    'Retencion_Fuente', 'Valor_Pagado', 'Factura',
                    'Valor_Fcta_Cliente', 'Fecha', 'Deuda_Tecnica',
                    'Factura_Pendiente', 'Dif', 'Diferencia_Bruta', 'Observaciones',
                    'Documento__Documento', 'Documento__Nombre',
                    'LineaId__Linea', 'ClienteId__Nombre_Cliente', 'ModuloId__Modulo'
                ))
                
                show_data = bool(resultados)
    else:
        form = FacturacionConsultoresDetalleFilterForm()
        
    context = {
        'form': form,
        'resultados': resultados,
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'mensaje': "No se encontraron resultados para los filtros aplicados." if busqueda_realizada and not show_data else "No se ha realizado ninguna búsqueda aún."
    }

    return render(request, 'Informes/Informedetallefacturacionconsultores.html', context)

@login_required
@verificar_permiso('can_view_informe_detalle_facturacion_consultores')
def exportar_detalle_facturacion_consultores_excel(request):
    form = FacturacionConsultoresDetalleFilterForm(request.GET or None)
    if not form.is_valid():
        return HttpResponse("Parámetros inválidos para generar el reporte.")

    qs = Facturacion_Consultores.objects.select_related('Documento', 'ClienteId', 'ModuloId', 'LineaId')
    qs = filtrar_facturacion_consultores_form(form, qs)

    if qs.count() == 0:
        return HttpResponse("No se encontraron resultados para los filtros aplicados.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle Facturación"

    headers = [
        'Año', 'Mes', 'Consultor', 'Empresa', 'Línea', 'N° Factura', 'Periodo Cobrado',
        'Aprobado Por', 'Fecha Cobro', 'Fecha Pago', 'Cliente', 'Módulo', 'Horas',
        'Valor Unitario', 'Valor Cobro', 'IVA', 'Valor Neto', 'Retención Fuente',
        'Valor Pagado', 'Factura', 'Valor Factura Cliente', 'Fecha', 'Deuda Técnica',
        'Factura Pendiente', '% Dif', 'Diferencia Bruta', 'Observaciones'
    ]

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row_idx, r in enumerate(qs, start=2):
        row_values = [
            r.Anio,
            r.Mes,
            getattr(r.Documento, 'Nombre', str(r.Documento)),
            getattr(r, 'Empresa', ''),
            getattr(r.LineaId, 'Linea', str(r.LineaId) if r.LineaId else ''),
            getattr(r, 'Cta_Cobro', ''),
            getattr(r, 'Periodo_Cobrado', ''),
            getattr(r, 'Aprobado_Por', ''),
            r.Fecha_Cobro.strftime('%Y-%m-%d') if getattr(r, 'Fecha_Cobro', None) else '',
            r.Fecha_Pago.strftime('%Y-%m-%d') if getattr(r, 'Fecha_Pago', None) else '',
            getattr(r.ClienteId, 'Nombre_Cliente', str(r.ClienteId) if r.ClienteId else ''),
            getattr(r.ModuloId, 'Modulo', str(r.ModuloId) if r.ModuloId else ''),
            r.Horas,
            r.Valor_Unitario,
            r.Valor_Cobro,
            r.IVA,
            r.Valor_Neto,
            r.Retencion_Fuente,
            r.Valor_Pagado,
            r.Factura,
            r.Valor_Fcta_Cliente,
            r.Fecha.strftime('%Y-%m-%d') if getattr(r, 'Fecha', None) else '',
            r.Deuda_Tecnica,
            r.Factura_Pendiente,
            r.Dif,
            r.Diferencia_Bruta,
            r.Observaciones,
        ]
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, (int, float)) and col_idx in {14, 15, 16, 17, 18, 21, 26}:  # Dinero
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Ajustar anchos
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="detalle_facturacion_consultores.xlsx"'
    wb.save(response)
    return response


