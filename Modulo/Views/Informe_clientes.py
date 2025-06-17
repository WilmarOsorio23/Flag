from django.shortcuts import render
from collections import defaultdict, Counter
from django.http import HttpResponse
from Modulo.forms import ClienteFilterForm
from Modulo.models import Clientes
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

# Función para filtrar clientes
def filtrar_clientes(form, clientes):
    # Obtener los parámetros de búsqueda del request
    clientes = clientes.order_by('Nombre_Cliente')
    nombre = form.cleaned_data.get('Nombre_Cliente')
    activo = form.cleaned_data.get('Activo')
    pais = form.cleaned_data.get('Pais')
    ciudad = form.cleaned_data.get('Ciudad')
    tipo_cliente = form.cleaned_data.get('TipoCliente')
    nacional = form.cleaned_data.get('Nacional')

    # Filtrar los clientes según los parámetros
    if nombre:
        nombre = nombre.Nombre_Cliente
        # Manejar tanto objetos Cliente como IDs
        if isinstance(nombre, str):
            try:
                clientes = clientes.filter(Nombre_Cliente=nombre)
            except (ValueError, TypeError):
                pass
        else:
            clientes = clientes.filter(Nombre_Cliente=nombre)
            print(nombre)
    if activo:
        clientes = clientes.filter(Activo=activo == 'True')
    if pais:
        clientes = clientes.filter(Pais=pais)
    if ciudad:
        clientes = clientes.filter(Ciudad=ciudad)
    if tipo_cliente:
        clientes = clientes.filter(TipoCliente=tipo_cliente)
    if nacional:
        clientes = clientes.filter(Nacional=nacional == 'True')

    return clientes

# Función para obtener información de los clientes
def obtener_info_clientes(clientes):
    clientes_info = []
    for cliente in clientes:
        datos_cliente = {
            'DocumentoId': cliente.DocumentoId,
            'TipoDocumentoID': cliente.TipoDocumentoID,
            'Nombre_Cliente': cliente.Nombre_Cliente,
            'Activo': 'SI' if cliente.Activo else 'NO' ,
            'Fecha_Inicio': cliente.Fecha_Inicio,
            'Fecha_Retiro': cliente.Fecha_Retiro,
            'Direccion': cliente.Direccion,
            'Telefono': cliente.Telefono,
            'CorreoElectronico': cliente.CorreoElectronico,
            #'ContactoID': cliente.ContactoID,
            'ContactoID': cliente.ContactoID.Nombre if cliente.ContactoID else '',
            'BuzonFacturacion': cliente.BuzonFacturacion,
            'TipoCliente': cliente.TipoCliente,
            'Ciudad': cliente.Ciudad,
            'Departamento': cliente.Departamento,
            'Pais': cliente.Pais,  
            'Nacional': 'SI' if cliente.Nacional else 'NO',          
        }
        clientes_info.append(datos_cliente)
    return clientes_info

# Vista para mostrar el informe de clientes
def clientes_filtrado(request):
    clientes_info = []
    show_data = False  
    busqueda_realizada = False

    if request.method == 'GET':
        form = ClienteFilterForm(request.GET)
        clientes = Clientes.objects.all()  # Definir clientes antes de la validación

        if request.GET:
            busqueda_realizada = True
            if form.is_valid():
                clientes = filtrar_clientes(form, clientes)
                clientes_info = obtener_info_clientes(clientes)
                show_data = bool(clientes_info)
    else: 
        form = ClienteFilterForm()

    #Aplicar lógica para cards
    
    # Estadísticas de clientes
    total_clientes = len(clientes_info)
    clientes_activos = sum(1 for c in clientes_info if c['Activo'] == 'SI')
    clientes_inactivos = total_clientes - clientes_activos

    # Conteo por tipo de cliente y país
    conteo_tipo_cliente = Counter(c['TipoCliente'] for c in clientes_info if c['TipoCliente'])
    conteo_paises = Counter(c['Pais'] for c in clientes_info if c['Pais'])

    # Nacionales / Internacionales
    clientes_nacionales = sum(1 for c in clientes_info if c['Nacional'] == 'SI')
    clientes_internacionales = total_clientes - clientes_nacionales

    if not conteo_tipo_cliente:
        conteo_tipo_cliente = {}

    if not conteo_paises:
        conteo_paises = {}

    context = {
        'form': form,
        'clientes_info': clientes_info,      
        'show_data': show_data,
        'busqueda_realizada': busqueda_realizada,
        'total_clientes': total_clientes,
        'clientes_activos': clientes_activos,
        'clientes_inactivos': clientes_inactivos,
        'conteo_tipo_cliente': dict(conteo_tipo_cliente),
        'conteo_paises': dict(conteo_paises),
        'clientes_nacionales': clientes_nacionales,
        'clientes_internacionales': clientes_internacionales,
        # Datos para gráficos de barras
        'labels_tipo_cliente': list(conteo_tipo_cliente.keys()),
        'values_tipo_cliente': list(conteo_tipo_cliente.values()),

        'labels_paises': list(conteo_paises.keys()),
        'values_paises': list(conteo_paises.values()),
        # Para etiquetas y datos de gráficos
        'labels_nacionalidad': ['Nacionales', 'Internacionales'],
        'values_nacionalidad': [clientes_nacionales, clientes_internacionales],
        'labels_activos': ['Activos', 'Inactivos'],
        'values_activos': [clientes_activos, clientes_inactivos],
        'mensaje': "No se encontraron resultados para los filtros aplicados." if busqueda_realizada and not show_data else "No se ha realizado ninguna búsqueda aún."
    }

    return render(request, 'informes/informes_clientes_index.html', context)


# Vista para exportar el informe a Excel
def exportar_clientes_excel(request):
    clientes_info = []

    if request.method == 'GET':
        form = ClienteFilterForm(request.GET)

        if form.is_valid():
            clientes = Clientes.objects.all()
            clientes = filtrar_clientes(form, clientes)
            clientes_info = obtener_info_clientes(clientes)
            if not clientes_info:
                return HttpResponse("No se encontraron resultados para los filtros aplicados.")
    
        else:
            return HttpResponse("Error en el formulario.")

        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe de Clientes"

        # Agregar encabezados
        encabezados = ["Documento", "Tipo Documento", "Nombre Cliente", "Activo", "Fecha Inicio", "Fecha Retiro", "Dirección",
                       "Telefono", "Correo", "Contacto", "Buzon Facturación", "Tipo Cliente", "Ciudad", "Departamento", "País", "Nacional"]
        for col_num, header in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Agregar datos
        for row_num, cliente in enumerate(clientes_info, 2):
            ws.cell(row=row_num, column=1, value=str(cliente['DocumentoId']))
            ws.cell(row=row_num, column=2, value=str(cliente['TipoDocumentoID'].Nombre) if cliente['TipoDocumentoID'] else '')
            ws.cell(row=row_num, column=3, value=str(cliente['Nombre_Cliente']))
            ws.cell(row=row_num, column=4, value=cliente['Activo'])
            ws.cell(row=row_num, column=5, value=cliente['Fecha_Inicio'].strftime('%Y-%m-%d') if cliente['Fecha_Inicio'] else '')
            ws.cell(row=row_num, column=6, value=cliente['Fecha_Retiro'].strftime('%Y-%m-%d') if cliente['Fecha_Retiro'] else '')
            ws.cell(row=row_num, column=7, value=str(cliente['Direccion']) if cliente['Direccion'] else '')
            ws.cell(row=row_num, column=8, value=str(cliente['Telefono']) if cliente['Telefono'] else '')
            ws.cell(row=row_num, column=9, value=str(cliente['CorreoElectronico']) if cliente['CorreoElectronico'] else '')
            ws.cell(row=row_num, column=10, value=str(cliente['ContactoID']) if cliente['ContactoID'] else '')
            ws.cell(row=row_num, column=11, value=str(cliente['BuzonFacturacion']) if cliente['BuzonFacturacion'] else '')
            ws.cell(row=row_num, column=12, value=str(cliente['TipoCliente']) if cliente['TipoCliente'] else '')
            ws.cell(row=row_num, column=13, value=str(cliente['Ciudad']) if cliente['Ciudad'] else '')
            ws.cell(row=row_num, column=14, value=str(cliente['Departamento']) if cliente['Departamento'] else '')
            ws.cell(row=row_num, column=15, value=str(cliente['Pais']) if cliente['Pais'] else '')
            ws.cell(row=row_num, column=16, value=cliente['Nacional'])

        # Aplicar estilos a las celdas
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Respuesta HTTP con archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"informe_clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response